import torch.nn as nn
import os
import torch.optim as optim
import random
from collections import deque
import numpy as np
import torch
from datetime import datetime
import time
import psutil
from statistics import mean
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# === 设备设置 ===
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"

# === 归一化配置 ===
# 根据新的变量范围更新
MIN_X = 1
MAX_X = 100
MIN_Y = 1
MAX_Y = 100
MIN_Z = 1
MAX_Z = 60

# 天气和时间段范围
MIN_WEATHER = 1
MAX_WEATHER = 6
MIN_TIMEPERIOD = 1
MAX_TIMEPERIOD = 6


def normalize_state(state):
    """将状态归一化到[0,1]"""
    x, y, z = state
    normalized_x = (x - MIN_X) / (MAX_X - MIN_X)
    normalized_y = (y - MIN_Y) / (MAX_Y - MIN_Y)
    normalized_z = (z - MIN_Z) / (MAX_Z - MIN_Z)
    return [normalized_x, normalized_y, normalized_z]


def denormalize_state(normalized_state):
    """将归一化状态还原到原始范围"""
    norm_x, norm_y, norm_z = normalized_state
    x = int(norm_x * (MAX_X - MIN_X) + MIN_X)
    y = int(norm_y * (MAX_Y - MIN_Y) + MIN_Y)
    z = int(norm_z * (MAX_Z - MIN_Z) + MIN_Z)
    return [x, y, z]


# === 简化的奖励函数 ===
def compute_reward(state, target_path, triggered, prev_triggered=None, prev_state=None):
    sim = jaccard_similarity(triggered, target_path)
    reward = sim * 10
    if target_path.issubset(triggered):
        reward += 1
    return reward


def execute_validation_rules_block4(weather, time_period, z):
    """验证规则函数 - weather, time_period, z组合"""
    triggered = set()

    # 将z映射到合适的范围以匹配条件逻辑
    x = z  # 直接使用z作为x
    y = (weather * time_period * 10 + z) % 100 + 1  # 基于输入参数计算y值

    # 1-7: 早高峰组合（time_period == 1）
    if time_period == 1:
        if x < 60 and y > 75:
            triggered.add(1)
        if x > 60 and y > 70:
            triggered.add(2)
        if x < 50 and y < 40:
            triggered.add(3)
        if x > 78 and 45 < y < 70:
            triggered.add(4)
        if 45 < x < 70 and y > 78:
            triggered.add(5)
        if x < 55 and 50 < y < 75:
            triggered.add(6)
        if 50 < x < 75 and y < 55:
            triggered.add(7)

    # 8-14: 晚高峰组合（time_period == 2）
    if time_period == 2:
        if x < 60 and y > 75:
            triggered.add(8)
        if x > 60 and y > 70:
            triggered.add(9)
        if x < 55 and y < 45:
            triggered.add(10)
        if 45 < x < 70 and y > 78:
            triggered.add(11)
        if x > 78 and 45 < y < 70:
            triggered.add(12)
        if 55 < x < 75 and y < 50:
            triggered.add(13)
        if x < 50 and 55 < y < 75:
            triggered.add(14)

    # 15-19: 午餐时间组合（time_period == 3）
    if time_period == 3:
        if x > 60 and 40 < y < 65:
            triggered.add(15)
        if 40 < x < 65 and y > 60:
            triggered.add(16)
        if 45 < x < 70 and 45 < y < 60:
            triggered.add(17)
        if x < 50 and y < 40:
            triggered.add(18)
        if x > 65 and y < 45:
            triggered.add(19)

    # 20-25: 夜间组合（time_period == 4）
    if time_period == 4:
        if x < 45 and y < 35:
            triggered.add(20)
        if x > 60 and y < 40:
            triggered.add(21)
        if x < 50 and y > 70:
            triggered.add(22)
        if 45 < x < 70 and 45 < y < 60:
            triggered.add(23)
        if x < 35 and y < 25:
            triggered.add(24)
        if 40 < x < 65 and y < 45:
            triggered.add(25)

    # 26-28: 周末组合（time_period == 5）
    if time_period == 5:
        if x < 60 and y < 50:
            triggered.add(26)
        if x > 65 and y > 75:
            triggered.add(27)
        if x > 60 and y < 45:
            triggered.add(28)

    # 29-33: 假日组合（time_period == 6）
    if time_period == 6:
        if 40 < x < 70 and 40 < y < 60:
            triggered.add(29)
        if x < 55 and y < 45:
            triggered.add(30)
        if x > 60 and y < 50:
            triggered.add(31)
        if x < 60 and y > 70:
            triggered.add(32)
        if x > 65 and y > 75:
            triggered.add(33)

    # 34-68: 天气相关扩展规则
    if weather == 1:  # 晴天
        if time_period in [1, 2] and x > 70:
            triggered.add(34)
        if time_period in [1, 2] and y > 70:
            triggered.add(35)
        if time_period in [3, 4] and x < 50:
            triggered.add(36)
        if time_period in [3, 4] and y < 50:
            triggered.add(37)
        if time_period in [5, 6] and 40 < x < 80:
            triggered.add(38)
        if time_period in [5, 6] and 40 < y < 80:
            triggered.add(39)

    if weather == 2:  # 雨天
        if time_period in [1, 2] and x > 75:
            triggered.add(40)
        if time_period in [1, 2] and y < 60:
            triggered.add(41)
        if time_period in [3, 4] and x < 45:
            triggered.add(42)
        if time_period in [3, 4] and y > 65:
            triggered.add(43)
        if time_period in [5, 6] and 35 < x < 75:
            triggered.add(44)
        if time_period in [5, 6] and 35 < y < 75:
            triggered.add(45)

    if weather == 3:  # 雾天
        if time_period in [1, 2] and x > 60:
            triggered.add(46)
        if time_period in [1, 2] and y > 65:
            triggered.add(47)
        if time_period in [3, 4] and x < 55:
            triggered.add(48)
        if time_period in [3, 4] and y < 55:
            triggered.add(49)
        if time_period in [5, 6] and 30 < x < 70:
            triggered.add(50)
        if time_period in [5, 6] and 30 < y < 70:
            triggered.add(51)

    if weather == 4:  # 雪天
        if time_period in [1, 2] and x > 65:
            triggered.add(52)
        if time_period in [1, 2] and y < 55:
            triggered.add(53)
        if time_period in [3, 4] and x < 40:
            triggered.add(54)
        if time_period in [3, 4] and y > 60:
            triggered.add(55)
        if time_period in [5, 6] and 25 < x < 65:
            triggered.add(56)
        if time_period in [5, 6] and 25 < y < 65:
            triggered.add(57)

    if weather == 5:  # 风天
        if time_period in [1, 2] and x > 70:
            triggered.add(58)
        if time_period in [1, 2] and y > 60:
            triggered.add(59)
        if time_period in [3, 4] and x < 35:
            triggered.add(60)
        if time_period in [3, 4] and y < 40:
            triggered.add(61)
        if time_period in [5, 6] and 20 < x < 60:
            triggered.add(62)
        if time_period in [5, 6] and 20 < y < 60:
            triggered.add(63)

    if weather == 6:  # 暴雨
        if time_period in [1, 2] and x > 55:
            triggered.add(64)
        if time_period in [1, 2] and y > 55:
            triggered.add(65)
        if time_period in [3, 4] and x < 45:
            triggered.add(66)
        if time_period in [3, 4] and y < 45:
            triggered.add(67)
        if time_period in [5, 6] and 15 < x < 55:
            triggered.add(68)

    # 69-78: 复合条件（多参数组合）
    if weather + time_period > 6:
        if x > 50 and y > 50:
            triggered.add(69)
        if x < 50 and y < 50:
            triggered.add(70)
        if x > y:
            triggered.add(71)
        if x < y:
            triggered.add(72)
        if abs(x - y) < 20:
            triggered.add(73)

    if weather + time_period <= 6:
        if x > 60 or y > 60:
            triggered.add(74)
        if x < 40 or y < 40:
            triggered.add(75)
        if x + y > 100:
            triggered.add(76)
        if x + y < 80:
            triggered.add(77)
        if abs(x - y) > 30:
            triggered.add(78)

    # 79-88: 数值关系条件
    if weather % 2 == time_period % 2:  # 同奇偶性
        if x % 10 < 5:
            triggered.add(79)
        if y % 10 >= 5:
            triggered.add(80)
        if (x + y) % 3 == 0:
            triggered.add(81)
        if (x * y) % 7 == 0:
            triggered.add(82)
        if x // 10 == y // 10:
            triggered.add(83)

    if weather % 2 != time_period % 2:  # 不同奇偶性
        if x > 75 or y > 75:
            triggered.add(84)
        if x < 25 or y < 25:
            triggered.add(85)
        if max(x, y) - min(x, y) > 40:
            triggered.add(86)
        if (x + y) // 2 > 50:
            triggered.add(87)
        if weather * time_period > 15:
            triggered.add(88)

    # 89-95: 高级组合条件（奇数天气）
    if weather in [1, 3, 5]:  # 奇数天气
        if time_period in [1, 3, 5] and x > 40:
            triggered.add(89)
        if time_period in [2, 4, 6] and y > 40:
            triggered.add(90)
        if x % 20 < 10 and y % 20 < 10:
            triggered.add(91)
        if x + weather * 10 > 50:
            triggered.add(92)
        if y + time_period * 10 > 50:
            triggered.add(93)
        if time_period in [1, 3, 5] and x < 60:
            triggered.add(94)
        if time_period in [2, 4, 6] and y < 60:
            triggered.add(95)

    # 96-98: 偶数天气条件
    if weather in [2, 4, 6]:  # 偶数天气
        if (x + y) % weather == 0:
            triggered.add(96)
        if x * weather > 100:
            triggered.add(97)
        if y * time_period > 100:
            triggered.add(98)

    # 99-100: 最后的复杂条件
    if (weather * time_period + z) % 7 == 0:
        triggered.add(99)
    if max(weather, time_period) * min(x, y) > 150:
        triggered.add(100)

    return triggered


def execute_Tr(weather, time_period, z):
    """执行验证规则的包装函数"""
    return execute_validation_rules_block4(weather, time_period, z)


# === 目标路径定义 ===
target_paths = [
    [15, 16, 48, 49, 74, 75, 76, 77, 78, 79, 80, 81, 82, 83, 89, 91, 92, 93, 94, 99, 100],
    [16, 18, 19, 60, 61, 70, 71, 72, 73, 80, 81, 82, 83, 89, 91, 92, 93, 94, 99, 100],
    [1, 4, 6, 46, 47, 74, 75, 76, 77, 78, 80, 81, 82, 83, 89, 92, 93, 94, 99, 100],
    [30, 31, 50, 51, 70, 71, 72, 73, 84, 85, 86, 87, 88, 91, 92, 93, 95, 99, 100],
    [18, 19, 36, 37, 74, 76, 77, 78, 79, 80, 81, 82, 83, 89, 92, 93, 94, 99, 100],
    [20, 24, 25, 36, 37, 76, 77, 78, 84, 86, 87, 88, 90, 91, 92, 93, 95, 99, 100],
    [8, 12, 34, 35, 74, 75, 76, 77, 78, 84, 86, 87, 88, 90, 91, 92, 93, 95, 100],
    [8, 10, 58, 59, 70, 71, 72, 73, 84, 85, 86, 87, 88, 91, 92, 93, 95, 99, 100],
    [8, 14, 46, 47, 75, 76, 77, 78, 84, 85, 86, 87, 88, 90, 92, 93, 95, 99, 100],
    [1, 2, 6, 46, 47, 75, 76, 77, 78, 79, 80, 81, 82, 83, 89, 92, 93, 94, 100],
    [39, 74, 75, 76, 77, 78, 79, 80, 81, 82, 83, 89, 91, 92, 93, 94, 99, 100],
    [20, 21, 60, 61, 70, 71, 72, 73, 84, 85, 86, 87, 88, 90, 92, 93, 95, 99],
    [8, 9, 11, 13, 40, 41, 75, 76, 77, 78, 79, 80, 81, 83, 96, 97, 98, 100],
    [18, 19, 54, 55, 70, 71, 72, 73, 84, 86, 87, 88, 96, 97, 98, 99, 100],
    [27, 75, 76, 77, 78, 79, 80, 81, 82, 83, 89, 91, 92, 93, 94, 99, 100],
    [25, 48, 49, 69, 71, 72, 73, 84, 85, 86, 87, 88, 90, 92, 93, 95, 100],
    [26, 28, 62, 70, 71, 72, 73, 80, 81, 82, 83, 89, 91, 92, 93, 94, 100],
    [32, 33, 68, 69, 71, 72, 73, 79, 80, 81, 82, 83, 96, 97, 98, 99, 100],
    [1, 52, 53, 74, 75, 76, 77, 78, 84, 85, 86, 87, 88, 97, 98, 99, 100],
    [8, 12, 14, 64, 65, 69, 71, 72, 73, 80, 81, 82, 83, 96, 97, 98, 100],
    [1, 3, 64, 65, 70, 71, 72, 73, 84, 86, 87, 88, 96, 97, 98, 99, 100],
    [22, 36, 37, 76, 77, 78, 85, 86, 87, 88, 90, 91, 93, 95, 100],
    [31, 45, 70, 71, 72, 73, 79, 80, 81, 83, 96, 97, 98, 99, 100],
    [22, 66, 67, 69, 71, 72, 73, 79, 80, 82, 83, 97, 98, 100],
    [44, 45, 69, 71, 72, 73, 79, 80, 83, 96, 97, 98, 99, 100],
    [57, 71, 72, 73, 79, 80, 83, 97, 98, 100],
    [15, 16, 17, 48, 49, 74, 75, 76, 77, 78, 79, 80, 82, 83, 89, 91, 92, 93, 94, 100],
    [1, 2, 5, 46, 47, 75, 76, 77, 78, 79, 80, 81, 82, 83, 89, 91, 92, 93, 94, 100],
    [20, 21, 25, 42, 43, 74, 76, 77, 78, 79, 80, 81, 82, 83, 96, 97, 98, 99, 100],
    [2, 5, 7, 40, 41, 75, 76, 77, 78, 84, 85, 86, 87, 88, 96, 97, 98, 99, 100],
    [26, 28, 56, 57, 70, 71, 72, 73, 84, 85, 86, 87, 88, 96, 97, 98, 99, 100],
    [26, 28, 38, 74, 76, 77, 78, 80, 81, 82, 83, 89, 91, 92, 93, 94, 100],
    [30, 31, 62, 63, 70, 71, 72, 73, 84, 86, 87, 88, 90, 91, 92, 93, 95],
    [29, 62, 63, 71, 72, 73, 84, 85, 86, 87, 88, 90, 92, 93, 95, 100],
    [23, 25, 60, 61, 71, 72, 73, 84, 85, 86, 87, 88, 90, 92, 93, 95, 100]
]

# 转换为集合
target_paths = [set(path) for path in target_paths]


def jaccard_similarity(set1, set2):
    intersection = len(set1 & set2)
    union = len(set1 | set2)
    if set2.issubset(set1):
        return 1.0
    return intersection / union if union != 0 else 0.0


# === 路径相似度矩阵计算和自动分组算法 ===
def compute_path_similarity_matrix(paths):
    n = len(paths)
    matrix = np.zeros((n, n))
    for i in range(n):
        for j in range(n):
            inter = len(paths[i] & paths[j])
            union = len(paths[i] | paths[j])
            matrix[i][j] = inter / union if union > 0 else 0.0
    return matrix


def group_paths_by_similarity(paths):
    sim_matrix = compute_path_similarity_matrix(paths)
    avg_sim_scores = np.mean(sim_matrix, axis=1)
    threshold = np.mean(avg_sim_scores)

    center_idx = np.argmax(avg_sim_scores)
    similar_group = [center_idx]
    for i in range(len(paths)):
        if i != center_idx and sim_matrix[center_idx][i] > threshold:
            similar_group.append(i)

    isolated_group = [i for i in range(len(paths)) if i not in similar_group]
    return similar_group, isolated_group


# === 优化的样本生成（分组差异化筛选）===
def compute_robustness(state, path):
    base = execute_Tr(state[0], state[1], state[2])
    if not base:
        return 0.0

    rob, neighbors = 0.0, 0
    for dx in [-1, 0, 1]:
        for dy in [-1, 0, 1]:
            for dz in [-1, 0, 1]:
                if dx == dy == dz == 0:
                    continue
                neighbor = [state[0] + dx, state[1] + dy, state[2] + dz]
                # 确保邻居状态在有效范围内
                neighbor[0] = max(MIN_X, min(MAX_X, neighbor[0]))
                neighbor[1] = max(MIN_Y, min(MAX_Y, neighbor[1]))
                neighbor[2] = max(MIN_Z, min(MAX_Z, neighbor[2]))

                n_trig = execute_Tr(neighbor[0], neighbor[1], neighbor[2])
                if not n_trig:
                    continue
                rob += jaccard_similarity(base, n_trig)
                neighbors += 1
    return rob / neighbors if neighbors > 0 else 0.0


def compute_q_value_score(state, similar_model):
    """保持原始Q值逻辑：返回1-归一化Q值"""
    if similar_model is None:
        return 0.0

    try:
        # 归一化状态后输入模型
        normalized_state = normalize_state(state)
        state_tensor = torch.tensor(normalized_state, dtype=torch.float32).unsqueeze(0).to(device)
        with torch.no_grad():
            q_values = similar_model(state_tensor)
            max_q_value = torch.max(q_values).item()
            normalized_q = max_q_value / 20.0
            normalized_q = max(0.0, min(1.0, normalized_q))
            return 1.0 - normalized_q
    except:
        return 0.0


def generate_samples_for_similar_paths(similar_group, num_candidates=2000, top_k=200, run_id=1):
    """为相似路径组生成样本（使用3个标准）"""
    SIMILAR_WEIGHTS = [0.55, 0.39, 0.06]

    def save_samples(path_id, samples, base_dir):
        os.makedirs(base_dir, exist_ok=True)
        filepath = os.path.join(base_dir, f"path{path_id}_similar.txt")
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(f"Similar Group Path {path_id} - Run {run_id}\n")
            f.write("weather time_period z\tScore\tSimilarity\tLengthDiff\tRobustness\n")
            for s in samples:
                weather, time_period, z = s['state']
                f.write(
                    f"{weather} {time_period} {z}\t{s['score']:.4f}\t{s['similarity']:.4f}\t{s['length_diff']:.4f}\t{s['robustness']:.4f}\n")

    base_dir = r"D:\实验\CNN\DQNNEW\path_samples_grouped"

    for path_idx in similar_group:
        path = target_paths[path_idx]
        path_id = path_idx + 1
        candidate_samples = []
        attempts = 0

        while len(candidate_samples) < num_candidates and attempts < num_candidates * 10:
            attempts += 1
            # 根据新的变量范围生成随机状态
            weather = np.random.randint(MIN_WEATHER, MAX_WEATHER + 1)
            time_period = np.random.randint(MIN_TIMEPERIOD, MAX_TIMEPERIOD + 1)
            z = np.random.randint(MIN_Z, MAX_Z + 1)
            state = (weather, time_period, z)
            triggered = execute_Tr(weather, time_period, z)

            if not triggered:
                continue

            sim = jaccard_similarity(triggered, path)
            len_diff = 1 - abs(len(triggered) - len(path)) / max(len(triggered), len(path))
            rob = compute_robustness(state, path)

            candidate_samples.append({
                'state': state,
                'similarity': sim,
                'length_diff': len_diff,
                'robustness': rob,
                'triggered': triggered
            })

        if candidate_samples:
            for sample in candidate_samples:
                score = (SIMILAR_WEIGHTS[0] * sample['similarity'] +
                         SIMILAR_WEIGHTS[1] * sample['length_diff'] +
                         SIMILAR_WEIGHTS[2] * sample['robustness'])
                sample['score'] = score

            candidate_samples.sort(key=lambda x: x['score'], reverse=True)
            selected_samples = candidate_samples[:top_k]
            save_samples(path_id=path_id, samples=selected_samples, base_dir=base_dir)


def generate_samples_for_isolated_paths(isolated_group, similar_model, num_candidates=2000, top_k=200, run_id=1):
    """为孤立路径组生成样本（使用4个标准，包括Q值预测）"""
    ISOLATED_WEIGHTS = [0.18, 0.21, 0.32, 0.29]

    def save_samples(path_id, samples, base_dir):
        os.makedirs(base_dir, exist_ok=True)
        filepath = os.path.join(base_dir, f"path{path_id}_isolated.txt")
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(f"Isolated Group Path {path_id} - Run {run_id}\n")
            f.write("weather time_period z\tScore\tSimilarity\tLengthDiff\tRobustness\tQValueScore\n")
            for s in samples:
                weather, time_period, z = s['state']
                f.write(
                    f"{weather} {time_period} {z}\t{s['score']:.4f}\t{s['similarity']:.4f}\t{s['length_diff']:.4f}\t{s['robustness']:.4f}\t{s['q_value_score']:.4f}\n")

    base_dir = r"D:\实验\CNN\DQNNEW\path_samples_grouped"

    for path_idx in isolated_group:
        path = target_paths[path_idx]
        path_id = path_idx + 1
        candidate_samples = []
        attempts = 0

        while len(candidate_samples) < num_candidates and attempts < num_candidates * 10:
            attempts += 1
            # 根据新的变量范围生成随机状态
            weather = np.random.randint(MIN_WEATHER, MAX_WEATHER + 1)
            time_period = np.random.randint(MIN_TIMEPERIOD, MAX_TIMEPERIOD + 1)
            z = np.random.randint(MIN_Z, MAX_Z + 1)
            state = (weather, time_period, z)
            triggered = execute_Tr(weather, time_period, z)

            if not triggered:
                continue

            sim = jaccard_similarity(triggered, path)
            len_diff = 1 - abs(len(triggered) - len(path)) / max(len(triggered), len(path))
            rob = compute_robustness(state, path)
            q_score = compute_q_value_score(state, similar_model)

            candidate_samples.append({
                'state': state,
                'similarity': sim,
                'length_diff': len_diff,
                'robustness': rob,
                'q_value_score': q_score,
                'triggered': triggered
            })

        if candidate_samples:
            for sample in candidate_samples:
                score = (ISOLATED_WEIGHTS[0] * sample['similarity'] +
                         ISOLATED_WEIGHTS[1] * sample['length_diff'] +
                         ISOLATED_WEIGHTS[2] * sample['robustness'] +
                         ISOLATED_WEIGHTS[3] * sample['q_value_score'])
                sample['score'] = score

            candidate_samples.sort(key=lambda x: x['score'], reverse=True)
            selected_samples = candidate_samples[:top_k]
            save_samples(path_id=path_id, samples=selected_samples, base_dir=base_dir)


# === 组别经验回放池（修改为支持不重复抽取）===
class GroupExperienceReplay:
    def __init__(self, capacity=20000):
        self.capacity = capacity
        self.buffer = deque(maxlen=self.capacity)
        self.priorities = deque(maxlen=self.capacity)

    def append(self, experience):
        self.buffer.append(experience)
        self.priorities.append(experience[-1])

    def sample(self, batch_size, alpha=0.6):
        """不重复抽样"""
        priorities = np.array(self.priorities) ** alpha
        probabilities = priorities / np.sum(priorities)

        # 不重复抽样
        batch_size = min(batch_size, len(self.buffer))
        batch_indices = np.random.choice(len(self.buffer), batch_size, replace=False, p=probabilities)
        batch = [self.buffer[idx] for idx in batch_indices]
        return batch, batch_indices, probabilities[batch_indices]

    def update_priorities(self, batch_indices, td_errors):
        for idx, td_error in zip(batch_indices, td_errors):
            if idx < len(self.priorities):
                self.priorities[idx] = max(td_error, 1e-6)

    def __len__(self):
        return len(self.buffer)

    def get_high_reward_samples(self, target_path, num_samples=20):
        """获取高奖励样本（不重复）"""
        if len(self.buffer) == 0:
            return []

        samples_with_recalculated_scores = []
        seen_states = set()  # 用于去重

        for experience in self.buffer:
            state_tensor = experience[0]
            # 还原归一化状态
            normalized_state = state_tensor.cpu().numpy().flatten()
            state_tuple = tuple(denormalize_state(normalized_state))

            # 去重检查
            if state_tuple in seen_states:
                continue
            seen_states.add(state_tuple)

            triggered = execute_Tr(state_tuple[0], state_tuple[1], state_tuple[2])
            new_reward = compute_reward(state_tuple, target_path, triggered, None, None)
            sim = jaccard_similarity(triggered, target_path)
            samples_with_recalculated_scores.append((state_tuple, new_reward, sim, triggered))

        samples_with_recalculated_scores.sort(key=lambda x: x[1], reverse=True)
        return samples_with_recalculated_scores[:num_samples]


def load_path_data(file_path):
    path_data = []
    with open(file_path, "r", encoding="utf-8") as f:
        lines = f.readlines()
        for line in lines[2:]:
            parts = line.strip().split("\t")
            state = tuple(map(int, parts[0].split()))
            path_data.append(state)
    return path_data


# === DQN网络 ===
class DQN(nn.Module):
    def __init__(self, state_dim, action_dim):
        super(DQN, self).__init__()
        self.fc1 = nn.Linear(state_dim, 128)
        self.fc2 = nn.Linear(128, 64)
        self.fc3 = nn.Linear(64, action_dim)

    def forward(self, state):
        x = torch.relu(self.fc1(state))
        x = torch.relu(self.fc2(x))
        return self.fc3(x)


# === DQN Agent（修改为使用归一化）===
class DQNAgentWithPER:
    def __init__(self, state_dim, action_dim, replay_buffer, gamma=0.99, epsilon=1.0, epsilon_decay=0.995,
                 epsilon_min=0.1, learning_rate=0.001, alpha=0.6, beta=0.4):
        self.state_dim = state_dim
        self.action_dim = action_dim
        self.gamma = gamma
        self.epsilon = epsilon
        self.epsilon_decay = epsilon_decay
        self.epsilon_min = epsilon_min
        self.learning_rate = learning_rate
        self.replay_buffer = replay_buffer
        self.alpha = alpha
        self.beta = beta

        self.model = DQN(state_dim, action_dim).to(device)
        self.target_model = DQN(state_dim, action_dim).to(device)
        self.optimizer = optim.Adam(self.model.parameters(), lr=self.learning_rate)
        self.target_model.load_state_dict(self.model.state_dict())

    def decode_action(self, action_idx):
        delta_values = [1, 2, 3, 5, -1, -2, -3, -5]
        dim = action_idx // 8
        delta_idx = action_idx % 8
        delta = delta_values[delta_idx]
        if dim == 0:
            return (delta, 0, 0)
        elif dim == 1:
            return (0, delta, 0)
        elif dim == 2:
            return (0, 0, delta)

    def act(self, state):
        """使用归一化状态进行决策"""
        if random.random() < self.epsilon:
            return random.randrange(self.action_dim)

        normalized_state = normalize_state(state)
        state_tensor = torch.tensor(normalized_state, dtype=torch.float32).unsqueeze(0).to(device)
        with torch.no_grad():
            q_values = self.model(state_tensor)
        return torch.argmax(q_values, dim=1).item()

    def store_transition(self, state, action, reward, next_state, done):
        """存储时使用归一化状态"""
        normalized_state = normalize_state(state)
        normalized_next_state = normalize_state(next_state)

        state_tensor = torch.tensor(normalized_state, dtype=torch.float32).unsqueeze(0).to(device)
        next_state_tensor = torch.tensor(normalized_next_state, dtype=torch.float32).unsqueeze(0).to(device)

        with torch.no_grad():
            q_values = self.model(state_tensor)
            next_q_values = self.target_model(next_state_tensor)
            max_next_q_values = next_q_values.max(1)[0]
            target_q_values = reward + (self.gamma * max_next_q_values * (1 - done))
            td_error = torch.abs(q_values[0][action] - target_q_values).item()

        self.replay_buffer.append((state_tensor, action, reward, next_state_tensor, done, td_error))
        return td_error

    def train(self, batch_size=32):
        if len(self.replay_buffer) < batch_size:
            return

        batch, batch_indices, probabilities = self.replay_buffer.sample(batch_size, alpha=self.alpha)
        states, actions, rewards, next_states, dones, _ = zip(*batch)

        weights = (len(self.replay_buffer) * probabilities) ** (-self.beta)
        weights = weights / weights.max()
        weights = torch.tensor(weights, dtype=torch.float32).to(device)

        states = torch.tensor(np.array([s.cpu().numpy().flatten() for s in states]), dtype=torch.float32).to(device)
        actions = torch.tensor(actions, dtype=torch.long).to(device)
        rewards = torch.tensor(rewards, dtype=torch.float32).to(device)
        next_states = torch.tensor(np.array([ns.cpu().numpy().flatten() for ns in next_states]),
                                   dtype=torch.float32).to(device)
        dones = torch.tensor(dones, dtype=torch.float32).to(device)

        current_q_values = self.model(states).gather(1, actions.unsqueeze(1)).squeeze(1)
        next_max_q_values = self.target_model(next_states).max(1)[0].detach()
        target_q_values = rewards + (self.gamma * next_max_q_values * (1 - dones))

        td_errors = current_q_values - target_q_values
        weighted_loss = (td_errors.pow(2) * weights).mean()

        self.optimizer.zero_grad()
        weighted_loss.backward()
        self.optimizer.step()

        new_priorities = torch.abs(td_errors).detach().cpu().numpy()
        self.replay_buffer.update_priorities(batch_indices, new_priorities)

        if self.epsilon > self.epsilon_min:
            self.epsilon *= self.epsilon_decay

    def update_target_model(self):
        self.target_model.load_state_dict(self.model.state_dict())


# === 修改后的训练函数 ===
def train_group(group_paths, path_documents, replay_buffer, batch_size=32, group_name=""):
    """训练指定组的路径 - 按新流程：4批次×50样本×3步×5轮重复"""
    state_dim = 3
    action_dim = 24  # 3维 × 8个delta值

    agent = DQNAgentWithPER(state_dim, action_dim, replay_buffer)

    global_steps = 0
    path_rewards = {}

    print(f"开始训练{group_name}，包含路径: {[idx + 1 for idx in group_paths]}")
    start_time = time.time()

    N_SAMPLES = 200
    BATCH_SIZE = 50
    N_BATCHES = 4
    N_STEPS = 3
    N_REPEATS = 5
    TARGET_UPDATE_EVERY_N_BATCHES = 2

    for path_idx in group_paths:
        file_path = os.path.join(path_documents,
                                 f"path{path_idx + 1}_{'similar' if group_name == '相似组' else 'isolated'}.txt")
        if not os.path.exists(file_path):
            print(f"  警告: 路径{path_idx + 1}的样本文件不存在，跳过")
            continue

        path_data = load_path_data(file_path)
        target_path = target_paths[path_idx]

        if path_idx not in path_rewards:
            path_rewards[path_idx] = 0

        print(f"\n  开始训练路径{path_idx + 1}，样本数: {len(path_data)}")

        for repeat in range(N_REPEATS):
            print(f"    第{repeat + 1}/{N_REPEATS}轮重复训练")

            batch_count = 0

            for batch_idx in range(N_BATCHES):
                batch_start = batch_idx * BATCH_SIZE
                batch_end = min(batch_start + BATCH_SIZE, N_SAMPLES)

                if batch_start >= len(path_data):
                    print(f"      第{batch_idx + 1}批: 样本不足，跳过")
                    break

                print(f"      第{batch_idx + 1}/{N_BATCHES}批 (样本 {batch_start}-{batch_end})")

                for sample_idx in range(batch_start, batch_end):
                    if sample_idx >= len(path_data):
                        break

                    state = path_data[sample_idx]
                    prev_state = None
                    prev_triggered = None

                    for step in range(N_STEPS):
                        legal_actions = []
                        for a in range(agent.action_dim):
                            dw, dt, dz = agent.decode_action(a)
                            cand_next = (state[0] + dw, state[1] + dt, state[2] + dz)
                            # 检查新状态是否在有效范围内
                            if (MIN_WEATHER <= cand_next[0] <= MAX_WEATHER and
                                    MIN_TIMEPERIOD <= cand_next[1] <= MAX_TIMEPERIOD and
                                    MIN_Z <= cand_next[2] <= MAX_Z):
                                legal_actions.append(a)

                        if not legal_actions:
                            break

                        if random.random() < agent.epsilon:
                            action = random.choice(legal_actions)
                        else:
                            normalized_state = normalize_state(state)
                            state_tensor = torch.tensor(normalized_state, dtype=torch.float32).unsqueeze(0).to(device)
                            with torch.no_grad():
                                q_values = agent.model(state_tensor)[0]
                            action = legal_actions[torch.argmax(q_values[legal_actions]).item()]

                        dw, dt, dz = agent.decode_action(action)
                        next_state = (state[0] + dw, state[1] + dt, state[2] + dz)

                        triggered = execute_Tr(next_state[0], next_state[1], next_state[2])
                        reward = compute_reward(next_state, target_path, triggered,
                                                prev_triggered, prev_state)

                        done = (step == N_STEPS - 1)

                        td_error = agent.store_transition(state, action, reward, next_state, done)

                        prev_state = state
                        prev_triggered = triggered
                        state = next_state

                        path_rewards[path_idx] += reward
                        global_steps += 1

                print(f"        批次{batch_idx + 1}完成，开始回放训练（更新策略网络参数）...")
                if len(agent.replay_buffer) >= batch_size:
                    agent.train(batch_size)

                batch_count += 1

                if batch_count % TARGET_UPDATE_EVERY_N_BATCHES == 0:
                    agent.update_target_model()
                    print(f"        已完成{batch_count}批，更新目标网络参数")

        print(f"  路径{path_idx + 1}训练完成，累积奖励: {path_rewards[path_idx]:.2f}")

    training_time = time.time() - start_time
    print(f"\n{group_name}训练完成，用时: {training_time:.2f}秒")
    print(f"经验池大小: {len(replay_buffer)}")

    return agent, path_rewards, training_time


# === 分阶段训练流程 ===
def generate_and_train_grouped_paths_staged(path_documents, similar_group, isolated_group, batch_size=32, run_id=1):
    """分阶段训练：先训练相似组，再用相似组模型为孤立组生成样本，最后训练孤立组"""

    print(f"\n=== 运行 {run_id}/20 分阶段训练开始 ===")
    similar_group_paths = [idx + 1 for idx in similar_group]
    isolated_group_paths = [idx + 1 for idx in isolated_group]

    print(f"相似组路径: {similar_group_paths}")
    print(f"孤立组路径: {isolated_group_paths}")

    total_start_time = time.time()

    print(f"\n[阶段1] 为相似组生成样本...")
    generate_samples_for_similar_paths(similar_group, num_candidates=2000, top_k=200, run_id=run_id)

    print(f"\n[阶段2] 训练相似组...")
    similar_replay_buffer = GroupExperienceReplay(capacity=20000)
    similar_agent, similar_path_rewards, similar_training_time = train_group(
        similar_group, path_documents, similar_replay_buffer, batch_size=batch_size, group_name="相似组"
    )

    print(f"\n[阶段3] 使用相似组模型为孤立组生成样本...")
    generate_samples_for_isolated_paths(isolated_group, similar_agent.model, num_candidates=2000, top_k=200,
                                        run_id=run_id)

    print(f"\n[阶段4] 训练孤立组...")
    isolated_replay_buffer = GroupExperienceReplay(capacity=20000)
    isolated_agent, isolated_path_rewards, isolated_training_time = train_group(
        isolated_group, path_documents, isolated_replay_buffer, batch_size=batch_size, group_name="孤立组"
    )

    total_path_rewards = {**similar_path_rewards, **isolated_path_rewards}
    total_cumulative_reward = sum(total_path_rewards.values())
    total_training_time = time.time() - total_start_time

    print(f"\n=== 运行 {run_id}/20 分阶段训练完成，总用时: {total_training_time:.2f}秒 ===")

    return similar_agent, isolated_agent, similar_replay_buffer, isolated_replay_buffer, total_cumulative_reward, total_path_rewards, total_training_time


# 创建Excel报告
def create_consolidated_excel_report(all_runs_data, similar_group, isolated_group, output_dir):
    """创建包含20次运行数据的Excel报告"""
    os.makedirs(output_dir, exist_ok=True)

    similar_group_paths = [idx + 1 for idx in similar_group]
    isolated_group_paths = [idx + 1 for idx in isolated_group]

    wb = Workbook()

    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    header_color = "4472C4"
    similar_group_color = "E2EFDA"
    isolated_group_color = "FCE4D6"
    stats_color = "FFF2CC"

    # === 工作表1: 各路径详细表现 ===
    ws_paths = wb.active
    ws_paths.title = "各路径详细表现"

    path_headers = ['路径编号', '分组类型'] + [f'第{i}次' for i in range(1, 21)] + ['平均相似度', '最高相似度',
                                                                                    '最低相似度', '标准差']
    for col, header in enumerate(path_headers, 1):
        cell = ws_paths.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    ws_paths.row_dimensions[1].height = 30

    for path_id in range(1, len(target_paths) + 1):
        row = path_id + 1

        if path_id in similar_group_paths:
            group_type = "高关联度路径组"
            row_color = similar_group_color
        elif path_id in isolated_group_paths:
            group_type = "低关联度路径组"
            row_color = isolated_group_color
        else:
            group_type = "未分组"
            row_color = "FFFFFF"

        cell = ws_paths.cell(row=row, column=1, value=f"路径{path_id}")
        cell.font = Font(bold=True, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type="solid")
        cell.border = thin_border

        cell = ws_paths.cell(row=row, column=2, value=group_type)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type="solid")
        cell.border = thin_border

        path_similarities = []
        for run_idx, run_data in enumerate(all_runs_data):
            sim = run_data['path_similarities'].get(path_id, {}).get('avg_similarity', 0.0)
            path_similarities.append(sim)

            cell = ws_paths.cell(row=row, column=3 + run_idx, value=round(sim, 4))
            cell.number_format = '0.0000'
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        stats_values = [
            round(np.mean(path_similarities), 4),
            round(np.max(path_similarities), 4),
            round(np.min(path_similarities), 4),
            round(np.std(path_similarities), 4)
        ]

        for i, value in enumerate(stats_values):
            cell = ws_paths.cell(row=row, column=23 + i, value=value)
            cell.number_format = '0.0000'
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill(start_color=stats_color, end_color=stats_color, fill_type="solid")
            cell.font = Font(bold=True, size=10)
            cell.border = thin_border

    ws_paths.column_dimensions['A'].width = 13
    ws_paths.column_dimensions['B'].width = 16
    for col in range(3, 23):
        ws_paths.column_dimensions[get_column_letter(col)].width = 10
    for col in range(23, 27):
        ws_paths.column_dimensions[get_column_letter(col)].width = 13

    # === 工作表2: 分组统计 ===
    ws_groups = wb.create_sheet("分组统计")

    group_headers = ['分组名称', '包含路径'] + [f'第{i}次' for i in range(1, 21)] + ['平均相似度', '标准差']
    for col, header in enumerate(group_headers, 1):
        cell = ws_groups.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    ws_groups.row_dimensions[1].height = 30

    row = 2

    cell = ws_groups.cell(row=row, column=1, value="高关联度路径组")
    cell.font = Font(bold=True, size=11)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = PatternFill(start_color=similar_group_color, end_color=similar_group_color, fill_type="solid")
    cell.border = thin_border

    cell = ws_groups.cell(row=row, column=2, value=','.join(map(str, similar_group_paths)))
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.fill = PatternFill(start_color=similar_group_color, end_color=similar_group_color, fill_type="solid")
    cell.border = thin_border

    group_similarities = []
    for run_idx, run_data in enumerate(all_runs_data):
        group_sim = np.mean(
            [run_data['path_similarities'].get(p, {}).get('avg_similarity', 0.0) for p in similar_group_paths])
        group_similarities.append(group_sim)

        cell = ws_groups.cell(row=row, column=3 + run_idx, value=round(group_sim, 4))
        cell.number_format = '0.0000'
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    cell = ws_groups.cell(row=row, column=23, value=round(np.mean(group_similarities), 4))
    cell.number_format = '0.0000'
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = PatternFill(start_color=stats_color, end_color=stats_color, fill_type="solid")
    cell.font = Font(bold=True, size=11)
    cell.border = thin_border

    cell = ws_groups.cell(row=row, column=24, value=round(np.std(group_similarities), 4))
    cell.number_format = '0.0000'
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = PatternFill(start_color=stats_color, end_color=stats_color, fill_type="solid")
    cell.font = Font(bold=True, size=11)
    cell.border = thin_border

    row += 1

    if isolated_group_paths:
        cell = ws_groups.cell(row=row, column=1, value="低关联度路径组")
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color=isolated_group_color, end_color=isolated_group_color, fill_type="solid")
        cell.border = thin_border

        cell = ws_groups.cell(row=row, column=2, value=','.join(map(str, isolated_group_paths)))
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.fill = PatternFill(start_color=isolated_group_color, end_color=isolated_group_color, fill_type="solid")
        cell.border = thin_border

        isolated_similarities = []
        for run_idx, run_data in enumerate(all_runs_data):
            iso_sim = np.mean(
                [run_data['path_similarities'].get(p, {}).get('avg_similarity', 0.0) for p in isolated_group_paths])
            isolated_similarities.append(iso_sim)

            cell = ws_groups.cell(row=row, column=3 + run_idx, value=round(iso_sim, 4))
            cell.number_format = '0.0000'
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        cell = ws_groups.cell(row=row, column=23, value=round(np.mean(isolated_similarities), 4))
        cell.number_format = '0.0000'
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color=stats_color, end_color=stats_color, fill_type="solid")
        cell.font = Font(bold=True, size=11)
        cell.border = thin_border

        cell = ws_groups.cell(row=row, column=24, value=round(np.std(isolated_similarities), 4))
        cell.number_format = '0.0000'
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color=stats_color, end_color=stats_color, fill_type="solid")
        cell.font = Font(bold=True, size=11)
        cell.border = thin_border

    ws_groups.column_dimensions['A'].width = 16
    ws_groups.column_dimensions['B'].width = 22
    for col in range(3, 23):
        ws_groups.column_dimensions[get_column_letter(col)].width = 10
    ws_groups.column_dimensions[get_column_letter(23)].width = 14
    ws_groups.column_dimensions[get_column_letter(24)].width = 12

    # === 工作表3: 详细样本数据 ===
    ws_samples = wb.create_sheet("详细样本数据")

    sample_headers = ['运行次数', '路径编号', '样本序号', 'Weather', 'TimePeriod', 'Z', '相似度', '触发规则集合']
    for col, header in enumerate(sample_headers, 1):
        cell = ws_samples.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    ws_samples.row_dimensions[1].height = 30

    sample_row = 2
    for run_idx, run_data in enumerate(all_runs_data, 1):
        for path_id in range(1, len(target_paths) + 1):
            samples = run_data['path_samples'].get(path_id, [])

            if path_id in similar_group_paths:
                path_color = similar_group_color
            elif path_id in isolated_group_paths:
                path_color = isolated_group_color
            else:
                path_color = "FFFFFF"

            for sample_idx, (state_tuple, reward, sim, triggered) in enumerate(samples, 1):
                weather, time_period, z = state_tuple
                triggered_str = ','.join(map(str, sorted(triggered)))

                cell = ws_samples.cell(row=sample_row, column=1, value=f"第{run_idx}次")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = PatternFill(start_color=path_color, end_color=path_color, fill_type="solid")
                cell.border = thin_border

                cell = ws_samples.cell(row=sample_row, column=2, value=f"路径{path_id}")
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = PatternFill(start_color=path_color, end_color=path_color, fill_type="solid")
                cell.border = thin_border

                cell = ws_samples.cell(row=sample_row, column=3, value=sample_idx)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

                for col_offset, value in enumerate([weather, time_period, z]):
                    cell = ws_samples.cell(row=sample_row, column=4 + col_offset, value=value)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = thin_border

                cell = ws_samples.cell(row=sample_row, column=7, value=round(sim, 4))
                cell.number_format = '0.0000'
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

                cell = ws_samples.cell(row=sample_row, column=8, value=f"{{{triggered_str}}}")
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border = thin_border

                sample_row += 1

    sample_widths = [13, 13, 11, 10, 12, 8, 12, 45]
    for i, width in enumerate(sample_widths, 1):
        ws_samples.column_dimensions[get_column_letter(i)].width = width

    output_path = os.path.join(output_dir, "20次运行综合报告_交通场景.xlsx")
    wb.save(output_path)
    print(f"\n✅ 综合Excel报告已生成: {output_path}")


def run_20_times_training():
    """运行20次完整的训练流程（交通场景版本）"""
    model_path_base = r"D:\实验\CNN\DQNNEW\saved_models_traffic"
    path_documents = r"D:\实验\CNN\DQNNEW\path_samples_grouped"
    output_dir = r"D:\实验\CNN\对比实验二\excel_reports_traffic"

    os.makedirs(model_path_base, exist_ok=True)
    os.makedirs(output_dir, exist_ok=True)

    similar_group, isolated_group = group_paths_by_similarity(target_paths)
    similar_group_display = [idx + 1 for idx in similar_group]
    isolated_group_display = [idx + 1 for idx in isolated_group]

    print("=" * 60)
    print("开始20次连续训练实验 - 交通场景版本")
    print(
        f"变量范围: weather∈[{MIN_WEATHER},{MAX_WEATHER}], time_period∈[{MIN_TIMEPERIOD},{MAX_TIMEPERIOD}], z∈[{MIN_Z},{MAX_Z}]")
    print("训练流程: 归一化输入 → 训练 → 抽取时还原 → 不重复抽样")
    print("=" * 60)
    print(f"\n自动分组结果:")
    print(f"相似路径组: {similar_group_display}")
    print(f"孤立路径组: {isolated_group_display}")
    print("\n" + "=" * 60)

    all_runs_data = []
    total_start_time = time.time()

    for run_id in range(1, 21):
        print(f"\n{'=' * 60}")
        print(f"开始第 {run_id}/20 次运行")
        print(f"{'=' * 60}")

        similar_agent, isolated_agent, similar_buffer, isolated_buffer, total_cumulative_reward, path_rewards, training_time = \
            generate_and_train_grouped_paths_staged(path_documents, similar_group, isolated_group, batch_size=32,
                                                    run_id=run_id)

        similar_model_path = os.path.join(model_path_base, f"similar_group_model_run_{run_id}.pth")
        isolated_model_path = os.path.join(model_path_base, f"isolated_group_model_run_{run_id}.pth")

        torch.save({
            'model_state_dict': similar_agent.model.state_dict(),
            'optimizer_state_dict': similar_agent.optimizer.state_dict(),
            'epsilon': similar_agent.epsilon,
            'normalization': {
                'x_range': (MIN_WEATHER, MAX_WEATHER),
                'y_range': (MIN_TIMEPERIOD, MAX_TIMEPERIOD),
                'z_range': (MIN_Z, MAX_Z)
            },
            'run_id': run_id,
            'group_type': 'similar_group',
            'group_paths': similar_group_display,
            'pool_size': len(similar_buffer),
            'pool_capacity': 20000,
        }, similar_model_path)

        torch.save({
            'model_state_dict': isolated_agent.model.state_dict(),
            'optimizer_state_dict': isolated_agent.optimizer.state_dict(),
            'epsilon': isolated_agent.epsilon,
            'normalization': {
                'x_range': (MIN_WEATHER, MAX_WEATHER),
                'y_range': (MIN_TIMEPERIOD, MAX_TIMEPERIOD),
                'z_range': (MIN_Z, MAX_Z)
            },
            'run_id': run_id,
            'group_type': 'isolated_group',
            'group_paths': isolated_group_display,
            'pool_size': len(isolated_buffer),
            'pool_capacity': 20000,
        }, isolated_model_path)

        print(f"[第{run_id}次] 模型已保存（包含归一化参数）")

        run_data = {
            'run_id': run_id,
            'training_time': training_time,
            'total_reward': total_cumulative_reward,
            'path_rewards': path_rewards,
            'path_similarities': {},
            'path_samples': {}
        }

        all_similarities = []
        for path_idx in range(len(target_paths)):
            target_path = target_paths[path_idx]
            path_id = path_idx + 1

            if path_id in similar_group_display:
                buffer = similar_buffer
            elif path_id in isolated_group_display:
                buffer = isolated_buffer
            else:
                continue

            high_reward_samples = buffer.get_high_reward_samples(target_path, num_samples=20)

            if high_reward_samples:
                similarities = [sim for _, _, sim, _ in high_reward_samples]
                run_data['path_similarities'][path_idx + 1] = {
                    'avg_similarity': np.mean(similarities),
                    'max_similarity': np.max(similarities),
                    'min_similarity': np.min(similarities),
                    'sample_count': len(similarities)
                }
                run_data['path_samples'][path_idx + 1] = high_reward_samples
                all_similarities.extend(similarities)
            else:
                run_data['path_similarities'][path_idx + 1] = {
                    'avg_similarity': 0.0,
                    'max_similarity': 0.0,
                    'min_similarity': 0.0,
                    'sample_count': 0
                }
                run_data['path_samples'][path_idx + 1] = []

        if all_similarities:
            run_data['overall_avg_similarity'] = np.mean(all_similarities)
            run_data['max_similarity'] = np.max(all_similarities)
            run_data['min_similarity'] = np.min(all_similarities)
        else:
            run_data['overall_avg_similarity'] = 0.0
            run_data['max_similarity'] = 0.0
            run_data['min_similarity'] = 0.0

        all_runs_data.append(run_data)

        print(f"[第{run_id}次] 完成! 总体平均相似度: {run_data['overall_avg_similarity']:.4f}")
        print(f"{'=' * 60}\n")

    total_time = time.time() - total_start_time

    print("\n正在生成综合Excel报告...")
    create_consolidated_excel_report(all_runs_data, similar_group, isolated_group, output_dir)

    print("\n" + "=" * 60)
    print("20次训练全部完成! - 交通场景版本")
    print("=" * 60)
    print(f"变量范围:")
    print(f"  weather (天气): [{MIN_WEATHER}, {MAX_WEATHER}]")
    print(f"  time_period (时间段): [{MIN_TIMEPERIOD}, {MAX_TIMEPERIOD}]")
    print(f"  z (行人数量): [{MIN_Z}, {MAX_Z}]")
    print(f"\n总耗时: {total_time:.2f}秒 ({total_time / 60:.2f}分钟)")
    print(f"平均每次耗时: {total_time / 20:.2f}秒")
    print(f"\n平均相似度统计:")
    avg_similarities = [r['overall_avg_similarity'] for r in all_runs_data]
    print(f"  总体平均: {np.mean(avg_similarities):.4f}")
    print(f"  最高: {np.max(avg_similarities):.4f}")
    print(f"  最低: {np.min(avg_similarities):.4f}")
    print(f"  标准差: {np.std(avg_similarities):.4f}")
    print(f"\n所有结果已保存到: {output_dir}")
    print("=" * 60)


if __name__ == "__main__":
    run_20_times_training()