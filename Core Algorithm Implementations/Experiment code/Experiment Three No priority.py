import torch.nn as nn
import os
import torch.optim as optim
import random
from collections import deque
import numpy as np
import torch
from datetime import datetime
import time
import psutil
from statistics import mean
import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

# === device setup ===
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"

# X, Y, Z - 150
X_MIN = 1
X_MAX = 50
Y_MIN = 1
Y_MAX = 50
Z_MIN = 1
Z_MAX = 50


# ===  ===
def normalize_state(state):
    """
    (1-50)(0-1)
    state: (x, y, z) tuple or array
    returns: normalized state (0-1 range)
    """
    x, y, z = state
    norm_x = (x - X_MIN) / (X_MAX - X_MIN)
    norm_y = (y - Y_MIN) / (Y_MAX - Y_MIN)
    norm_z = (z - Z_MIN) / (Z_MAX - Z_MIN)
    return (norm_x, norm_y, norm_z)


def denormalize_state(norm_state):
    """
    (0-1)(1-50)
    norm_state: normalized (x, y, z) tuple or array
    returns: denormalized state (1-50 range)
    """
    norm_x, norm_y, norm_z = norm_state
    x = int(round(norm_x * (X_MAX - X_MIN) + X_MIN))
    y = int(round(norm_y * (Y_MAX - Y_MIN) + Y_MIN))
    z = int(round(norm_z * (Z_MAX - Z_MIN) + Z_MIN))
    # 
    x = max(X_MIN, min(X_MAX, x))
    y = max(Y_MIN, min(Y_MAX, y))
    z = max(Z_MIN, min(Z_MAX, z))
    return (x, y, z)


def is_valid_state(state):
    """"""
    x, y, z = state
    return (X_MIN <= x <= X_MAX and
            Y_MIN <= y <= Y_MAX and
            Z_MIN <= z <= Z_MAX)


def clip_state(state):
    """"""
    x, y, z = state
    return (
        max(X_MIN, min(X_MAX, x)),
        max(Y_MIN, min(Y_MAX, y)),
        max(Z_MIN, min(Z_MAX, z))
    )


# === Metric() ===
class EnhancedStandardMetricsCollector:
    def __init__(self, experiment_name="Enhanced_Standard_DQN"):
        self.experiment_name = experiment_name
        self.start_time = None
        self.end_time = None

        # Metric
        self.total_reward = 0
        self.td_errors = []
        self.final_output_similarities = []
        self.action_improvements = []
        self.total_memory_usage = 0
        self.memory_check_count = 0
        self.step_count = 0

        # Metric()
        self.episode_rewards = []  # episode
        self.episode_similarities = []  # episodeAverage Similarity
        self.episode_td_errors = []  # episodeTD
        self.episode_epsilon_values = []  # epsilon
        self.episode_memory_usage = []  # episode

        # Path Metric
        self.similar_paths_performance = []
        self.isolated_paths_performance = []

        # ()
        self.milestone_data = {}  # episode 50, 100, 150, 200

        # 
        self.convergence_window = 20  # 
        self.convergence_threshold = 0.02  # 
        self.convergence_detected_episode = None

        # Metric
        self.sample_efficiency_data = []  # (episode, )
        self.performance_milestones = [0.6, 0.7, 0.75, 0.8]  # 

        # 
        self.learning_curve_characteristics = {}
        self.early_vs_late_performance = {}

        # X, Y, Z
        self.xyz_coordinate_stats = {
            'x_values': [],
            'y_values': [],
            'z_values': [],
            'x_distribution': {},
            'y_distribution': {},
            'z_distribution': {}
        }

    def reset(self):
        """Metric, """
        self.start_time = None
        self.end_time = None
        self.total_reward = 0
        self.td_errors = []
        self.final_output_similarities = []
        self.action_improvements = []
        self.total_memory_usage = 0
        self.memory_check_count = 0
        self.step_count = 0
        self.episode_rewards = []
        self.episode_similarities = []
        self.episode_td_errors = []
        self.episode_epsilon_values = []
        self.episode_memory_usage = []
        self.similar_paths_performance = []
        self.isolated_paths_performance = []
        self.milestone_data = {}
        self.convergence_detected_episode = None
        self.sample_efficiency_data = []
        self.learning_curve_characteristics = {}
        self.early_vs_late_performance = {}
        self.xyz_coordinate_stats = {
            'x_values': [],
            'y_values': [],
            'z_values': [],
            'x_distribution': {},
            'y_distribution': {},
            'z_distribution': {}
        }

    def start_training(self):
        self.start_time = time.time()

    def end_training(self):
        self.end_time = time.time()

    def record_xyz_coordinates(self, x_val, y_val, z_val):
        """X, Y, Z"""
        self.xyz_coordinate_stats['x_values'].append(x_val)
        self.xyz_coordinate_stats['y_values'].append(y_val)
        self.xyz_coordinate_stats['z_values'].append(z_val)

        self.xyz_coordinate_stats['x_distribution'][x_val] = self.xyz_coordinate_stats['x_distribution'].get(x_val,
                                                                                                             0) + 1
        self.xyz_coordinate_stats['y_distribution'][y_val] = self.xyz_coordinate_stats['y_distribution'].get(y_val,
                                                                                                             0) + 1
        self.xyz_coordinate_stats['z_distribution'][z_val] = self.xyz_coordinate_stats['z_distribution'].get(z_val,
                                                                                                             0) + 1

    def record_step_metrics(self, reward, td_error, triggered, target_path, x_coord=None, y_coord=None, z_coord=None):
        """Metric"""
        self.step_count += 1
        self.total_reward += reward
        self.td_errors.append(td_error)

        if x_coord is not None and y_coord is not None and z_coord is not None:
            self.record_xyz_coordinates(x_coord, y_coord, z_coord)

        process = psutil.Process(os.getpid())
        current_memory = process.memory_info().rss / 1024 / 1024
        self.total_memory_usage += current_memory
        self.memory_check_count += 1

    def record_episode_metrics(self, episode, episode_reward, avg_similarity, avg_td_error, epsilon,
                               path_group="similar"):
        """episodeMetric"""
        self.episode_rewards.append(episode_reward)
        self.episode_similarities.append(avg_similarity)
        self.episode_td_errors.append(avg_td_error)
        self.episode_epsilon_values.append(epsilon)

        process = psutil.Process(os.getpid())
        current_memory = process.memory_info().rss / 1024 / 1024
        self.episode_memory_usage.append(current_memory)

        if path_group == "similar":
            self.similar_paths_performance.append({
                'episode': episode,
                'reward': episode_reward,
                'similarity': avg_similarity,
                'td_error': avg_td_error
            })
        else:
            self.isolated_paths_performance.append({
                'episode': episode,
                'reward': episode_reward,
                'similarity': avg_similarity,
                'td_error': avg_td_error
            })

        if episode in [50, 100, 150, 200, 250, 300, 400, 450, 500]:
            self.milestone_data[episode] = {
                'avg_reward': np.mean(self.episode_rewards[-10:]) if len(
                    self.episode_rewards) >= 10 else episode_reward,
                'avg_similarity': avg_similarity,
                'avg_td_error': avg_td_error,
                'epsilon': epsilon,
                'memory_usage': current_memory,
                'total_steps': self.step_count
            }

        self._check_convergence(episode)
        self._check_performance_milestones(episode, avg_similarity)

    def _check_convergence(self, episode):
        """"""
        if len(self.episode_similarities) >= self.convergence_window and self.convergence_detected_episode is None:
            recent_similarities = self.episode_similarities[-self.convergence_window:]
            if np.std(recent_similarities) < self.convergence_threshold:
                self.convergence_detected_episode = episode

    def _check_performance_milestones(self, episode, similarity):
        """"""
        for milestone in self.performance_milestones:
            if similarity >= milestone and not any(data[1] == milestone for data in self.sample_efficiency_data):
                self.sample_efficiency_data.append((episode, milestone, self.step_count))

    def record_final_output_sample(self, triggered, target_path):
        """final samplesSimilarity"""
        if len(triggered | target_path) > 0:
            similarity = len(triggered & target_path) / len(triggered | target_path)
        else:
            similarity = 0.0
        self.final_output_similarities.append(similarity)

    def record_action_improvement(self, current_reward, prev_reward):
        """"""
        if prev_reward is not None:
            improvement = current_reward - prev_reward
            self.action_improvements.append(1 if improvement > 0 else 0)


# Metric
enhanced_standard_metrics = EnhancedStandardMetricsCollector("Enhanced_Standard_DQN_No_Priority")


# === reward function ===
def compute_reward(state, target_path, triggered, prev_triggered=None, prev_state=None):
    sim = jaccard_similarity(triggered, target_path)
    reward = sim * 10
    if target_path.issubset(triggered):
        reward += 1
    return reward


def execute_Tr(dx: int, dy: int, dz: int):
    # --- 1. constants and configuration ---
    MAX_GRID_SIZE = 500.0  # ,  500.0
    INITIAL_BATTERY = 1000.0  # , Path 
    BATTERY_PER_STEP = 1.0  # , 
    SAFE_DISTANCE = 5.0  #  ()
    CRITICAL_BATTERY_LEVEL = 100.0  #  ()
    TARGET_X, TARGET_Y, TARGET_Z = 450.0, 450.0, 200.0  #  ()

    MIN_PLANNING_X = 10.0
    MIN_PLANNING_Y = 15.0
    MIN_PLANNING_Z = 8.0
    CRITICAL_X_VELOCITY = 20.0
    CRITICAL_Y_VELOCITY = 25.0
    CRITICAL_Z_VELOCITY = 15.0

    triggered = set()

    # , 
    # , 
    current_x = random.uniform(0.0, MAX_GRID_SIZE)
    current_y = random.uniform(0.0, MAX_GRID_SIZE)
    current_z = random.uniform(0.0, MAX_GRID_SIZE)

    # '''', 
    # Run 10-15branch 'self.y' .
    simulated_y = current_y  #  current_y  self.y 

    # --- branch 1-4 ---
    if abs(dx) < MIN_PLANNING_X != abs(dy) < MIN_PLANNING_X: triggered.add(1)
    if abs(dx) < MIN_PLANNING_X != abs(dz) < MIN_PLANNING_X: triggered.add(2)
    if abs(dx) < MIN_PLANNING_X != abs(dx) < MIN_PLANNING_Y: triggered.add(3)
    if abs(dx) < MIN_PLANNING_X != abs(dx) < MIN_PLANNING_Z: triggered.add(4)

    # --- branch 5-9 ---
    if abs(dz) > MIN_PLANNING_Z * 2 != abs(dx) > MIN_PLANNING_Z * 2: triggered.add(5)
    if abs(dz) > MIN_PLANNING_Z * 2 != abs(dy) > MIN_PLANNING_Z * 2: triggered.add(6)
    if abs(dz) > MIN_PLANNING_Z * 2 != abs(dz) > MIN_PLANNING_X * 2: triggered.add(7)
    if abs(dz) > MIN_PLANNING_Z * 2 != abs(dz) > MIN_PLANNING_Y * 2: triggered.add(8)
    if abs(dz) > MIN_PLANNING_Z * 2 != abs(dz) > MIN_PLANNING_Z: triggered.add(9)

    # --- branch 10-15 --- ( simulated_y  self.y)
    if TARGET_Y > simulated_y and dy < 20 != TARGET_Y > simulated_y and dy < 10: triggered.add(10)
    if TARGET_Y > simulated_y and dy < 20 != TARGET_Y > simulated_y and dy < 30: triggered.add(11)
    if TARGET_Y > simulated_y and dy < 20 != TARGET_Y > simulated_y and dy < 40: triggered.add(12)
    if TARGET_Y > simulated_y and dy < 20 != TARGET_Y > simulated_y and dy < 50: triggered.add(13)
    if TARGET_Y > simulated_y and dy < 20 != TARGET_Y > simulated_y and dx < 20: triggered.add(14)
    if TARGET_Y > simulated_y and dy < 20 != TARGET_Y > simulated_y and dz < 20: triggered.add(15)

    # --- branch 16-21 ---
    if abs(dy) > CRITICAL_X_VELOCITY * 1.5 != abs(dx) > CRITICAL_X_VELOCITY * 1.5: triggered.add(16)
    if abs(dy) > CRITICAL_X_VELOCITY * 1.5 != abs(dz) > CRITICAL_X_VELOCITY * 1.5: triggered.add(17)
    if abs(dy) > CRITICAL_X_VELOCITY * 1.5 != abs(dy) > CRITICAL_X_VELOCITY: triggered.add(18)
    if abs(dy) > CRITICAL_X_VELOCITY * 1.5 != abs(dy) > CRITICAL_X_VELOCITY * 2: triggered.add(19)
    if abs(dy) > CRITICAL_X_VELOCITY * 1.5 != abs(dy) > CRITICAL_Z_VELOCITY * 1.5: triggered.add(20)
    if abs(dy) > CRITICAL_X_VELOCITY * 1.5 != abs(dy) > CRITICAL_Y_VELOCITY * 1.5: triggered.add(21)

    # --- branch 22-29 --- ( current_x, current_y, current_z )
    if TARGET_Z < current_z and dz > CRITICAL_Z_VELOCITY != TARGET_X < current_z and dz > CRITICAL_Z_VELOCITY: triggered.add(
        22)
    if TARGET_Z < current_z and dz > CRITICAL_Z_VELOCITY != TARGET_Y < current_z and dz > CRITICAL_Z_VELOCITY: triggered.add(
        23)
    if TARGET_Z < current_z and dz > CRITICAL_Z_VELOCITY != TARGET_Z < current_x and dz > CRITICAL_Z_VELOCITY: triggered.add(
        24)
    if TARGET_Z < current_z and dz > CRITICAL_Z_VELOCITY != TARGET_Z < current_y and dz > CRITICAL_Z_VELOCITY: triggered.add(
        25)
    if TARGET_Z < current_z and dz > CRITICAL_Z_VELOCITY != TARGET_Z < current_z and dx > CRITICAL_Z_VELOCITY: triggered.add(
        26)
    if TARGET_Z < current_z and dz > CRITICAL_Z_VELOCITY != TARGET_Z < current_z and dy > CRITICAL_Z_VELOCITY: triggered.add(
        27)
    if TARGET_Z < current_z and dz > CRITICAL_Z_VELOCITY != TARGET_Z < current_z and dz > CRITICAL_X_VELOCITY: triggered.add(
        28)
    if TARGET_Z < current_z and dz > CRITICAL_Z_VELOCITY != TARGET_Z < current_z and dz > CRITICAL_Y_VELOCITY: triggered.add(
        29)

    return triggered


def jaccard_similarity(set1, set2):
    intersection = len(set1 & set2)
    union = len(set1 | set2)
    if set2.issubset(set1):
        return 1.0
    return intersection / union if union != 0 else 0.0


# === Path Similarity ===
def compute_path_similarity_matrix(paths):
    n = len(paths)
    matrix = np.zeros((n, n))
    for i in range(n):
        for j in range(n):
            inter = len(paths[i] & paths[j])
            union = len(paths[i] | paths[j])
            matrix[i][j] = inter / union if union > 0 else 0.0
    return matrix


targetPaths = [
    {1, 2, 3, 4, 10, 11, 12, 13, 14, 15, 24, 25, 26, 27, 28, 29},
    {5, 6, 7, 8, 9, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25},
    {5, 6, 7, 8, 9, 17, 18, 19, 20, 21, 24, 25, 26, 27, 28, 29}
]


# === Path  ===
def group_paths_by_similarity(paths):
    sim_matrix = compute_path_similarity_matrix(paths)
    avg_sim_scores = np.mean(sim_matrix, axis=1)
    threshold = np.mean(avg_sim_scores)

    center_idx = np.argmax(avg_sim_scores)
    similar_group = [center_idx]
    for i in range(len(paths)):
        if i != center_idx and sim_matrix[center_idx][i] > threshold:
            similar_group.append(i)

    isolated_group = [i for i in range(len(paths)) if i not in similar_group]
    return similar_group, isolated_group


# === (, )===
class StandardExperienceReplay:
    def __init__(self, capacity=10000):
        self.capacity = capacity
        self.buffer = deque(maxlen=self.capacity)

    def append(self, experience):
        """"""
        self.buffer.append(experience[:5])

    def sample(self, batch_size):
        """
        
        , 
        """
        if len(self.buffer) < batch_size:
            return [], [], None
        # replace=False 
        batch_indices = np.random.choice(len(self.buffer), batch_size, replace=False)
        batch = [self.buffer[idx] for idx in batch_indices]
        return batch, batch_indices, None

    def __len__(self):
        return len(self.buffer)

    def get_high_reward_samples(self, target_path, num_samples=20):
        """, """
        if len(self.buffer) == 0:
            return []
        samples_with_recalculated_scores = []
        for experience in self.buffer:
            # , 
            norm_state_tensor = experience[0]
            norm_state = tuple(norm_state_tensor.cpu().numpy().flatten())
            # 
            state_tuple = denormalize_state(norm_state)
            triggered = execute_Tr(state_tuple[0], state_tuple[1], state_tuple[2])
            new_reward = compute_reward(state_tuple, target_path, triggered, None, None)
            sim = jaccard_similarity(triggered, target_path)
            samples_with_recalculated_scores.append((state_tuple, new_reward, sim, triggered))
        samples_with_recalculated_scores.sort(key=lambda x: x[1], reverse=True)
        return samples_with_recalculated_scores[:num_samples]


def load_path_data(file_path):
    path_data = []
    with open(file_path, "r", encoding="utf-8") as f:
        lines = f.readlines()
        for line in lines[2:]:
            parts = line.strip().split("\t")
            state = tuple(map(int, parts[0].split()))
            path_data.append(state)
    return path_data


# === DQN ===
class DQN(nn.Module):
    def __init__(self, state_dim, action_dim):
        super(DQN, self).__init__()
        self.fc1 = nn.Linear(state_dim, 128)
        self.fc2 = nn.Linear(128, 64)
        self.fc3 = nn.Linear(64, action_dim)

    def forward(self, state):
        x = torch.relu(self.fc1(state))
        x = torch.relu(self.fc2(x))
        return self.fc3(x)


# === DQN Agent()===
class StandardDQNAgent:
    def __init__(self, state_dim, action_dim, replay_buffer, gamma=0.99, epsilon=1.0,
                 epsilon_decay=0.995, epsilon_min=0.1, learning_rate=0.001):
        self.state_dim = state_dim
        self.action_dim = action_dim
        self.gamma = gamma
        self.epsilon = epsilon
        self.epsilon_decay = epsilon_decay
        self.epsilon_min = epsilon_min
        self.learning_rate = learning_rate
        self.replay_buffer = replay_buffer

        self.model = DQN(state_dim, action_dim).to(device)
        self.target_model = DQN(state_dim, action_dim).to(device)
        self.optimizer = optim.Adam(self.model.parameters(), lr=self.learning_rate)
        self.target_model.load_state_dict(self.model.state_dict())

    def decode_action(self, action_idx):
        delta_values = [10, 5, 3, 2, 1, -1, -2, -3, -5, -10]
        dim = action_idx // 10
        delta_idx = action_idx % 10
        delta = delta_values[delta_idx]
        if dim == 0:
            return (delta, 0, 0)
        elif dim == 1:
            return (0, delta, 0)
        elif dim == 2:
            return (0, 0, delta)

    def act(self, state):
        """
        
        :  (1-50)
        """
        if random.random() < self.epsilon:
            return random.randrange(self.action_dim)
        # 
        norm_state = normalize_state(state)
        state_tensor = torch.tensor(norm_state, dtype=torch.float32).unsqueeze(0).to(device)
        with torch.no_grad():
            q_values = self.model(state_tensor)
        return torch.argmax(q_values, dim=1).item()

    def store_transition(self, state, action, reward, next_state, done):
        """
        , 
        :  (1-50)
        :  (0-1)
        """
        # 
        norm_state = normalize_state(state)
        norm_next_state = normalize_state(next_state)

        # tensor
        norm_state_tensor = torch.tensor(norm_state, dtype=torch.float32).unsqueeze(0).to(device)
        norm_next_state_tensor = torch.tensor(norm_next_state, dtype=torch.float32).unsqueeze(0).to(device)

        # TD
        with torch.no_grad():
            q_values = self.model(norm_state_tensor)
            next_q_values = self.target_model(norm_next_state_tensor)
            max_next_q_values = next_q_values.max(1)[0]
            target_q_values = reward + (self.gamma * max_next_q_values * (1 - done))
            td_error = torch.abs(q_values[0][action] - target_q_values).item()

        # 
        self.replay_buffer.append((norm_state_tensor, action, reward, norm_next_state_tensor, done))
        return td_error

    def train(self, batch_size=32):
        """
        
        , 
        """
        if len(self.replay_buffer) < batch_size:
            return

        # ()
        batch, batch_indices, _ = self.replay_buffer.sample(batch_size)
        states, actions, rewards, next_states, dones = zip(*batch)

        # tensor(, )
        states = torch.tensor(np.array([s.cpu().numpy().flatten() for s in states]), dtype=torch.float32).to(device)
        actions = torch.tensor(actions, dtype=torch.long).to(device)
        rewards = torch.tensor(rewards, dtype=torch.float32).to(device)
        next_states = torch.tensor(np.array([ns.cpu().numpy().flatten() for ns in next_states]),
                                   dtype=torch.float32).to(device)
        dones = torch.tensor(dones, dtype=torch.float32).to(device)

        # Q
        current_q_values = self.model(states).gather(1, actions.unsqueeze(1)).squeeze(1)
        next_max_q_values = self.target_model(next_states).max(1)[0].detach()
        target_q_values = rewards + (self.gamma * next_max_q_values * (1 - dones))

        loss = nn.MSELoss()(current_q_values, target_q_values)

        self.optimizer.zero_grad()
        loss.backward()
        self.optimizer.step()

        if self.epsilon > self.epsilon_min:
            self.epsilon *= self.epsilon_decay

    def update_target_model(self):
        self.target_model.load_state_dict(self.model.state_dict())


# === Sample generation===
def generate_samples_for_similar_paths(similar_group_indices, num_total=2000, top_k=200):
    def jaccard_similarity_local(a, b):
        if not a and not b: return 1.0
        return len(a & b) / len(a | b) if a | b else 0.0

    def compute_robustness(state, path):
        base = execute_Tr(state[0], state[1], state[2])
        if not base: return 0.0
        rob, neighbors = 0.0, 0
        for dx in [-1, 0, 1]:
            for dy in [-1, 0, 1]:
                for dz in [-1, 0, 1]:
                    if dx == dy == dz == 0: continue
                    neighbor_state = (state[0] + dx, state[1] + dy, state[2] + dz)
                    if not is_valid_state(neighbor_state): continue
                    neighbor = clip_state(neighbor_state)
                    n_trig = execute_Tr(neighbor[0], neighbor[1], neighbor[2])
                    if not n_trig: continue
                    rob += jaccard_similarity_local(base, n_trig)
                    neighbors += 1
        return rob / neighbors if neighbors > 0 else 0.0

    def save_samples(path_id, samples, base_dir):
        os.makedirs(base_dir, exist_ok=True)
        filepath = os.path.join(base_dir, f"enhanced_standard_path{path_id}.txt")
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(f"Enhanced Standard Path {path_id}\n")
            f.write("x y z\tScore\tSimilarity\tLengthDiff\tRobustness\n")
            for s in samples:
                x, y, z = s[0]
                f.write(f"{x} {y} {z}\t{s[1]:.4f}\t{s[2]:.4f}\t{s[3]:.4f}\t{s[4]:.4f}\n")

    base_dir = r"D:\Experiment\CNN\DQNNEW\path_samples_enhanced_standard"
    for path_idx in similar_group_indices:
        path = targetPaths[path_idx]
        samples = []
        attempts = 0
        while len(samples) < top_k and attempts < num_total * 5:
            attempts += 1
            state = (
                np.random.randint(X_MIN, X_MAX + 1),
                np.random.randint(Y_MIN, Y_MAX + 1),
                np.random.randint(Z_MIN, Z_MAX + 1)
            )
            triggered = execute_Tr(state[0], state[1], state[2])
            if not triggered: continue
            sim = jaccard_similarity_local(triggered, path)
            len_diff = 1 - abs(len(triggered) - len(path)) / max(len(triggered), len(path))
            rob = compute_robustness(state, path)
            score = 0.55 * sim + 0.25 * len_diff + 0.2 * rob
            samples.append((state, score, sim, len_diff, rob))
        if samples:
            samples.sort(key=lambda x: x[1], reverse=True)
            save_samples(path_id=path_idx + 1, samples=samples[:top_k], base_dir=base_dir)


# === Run ===
def enhanced_standard_generate_and_train_for_similar_paths(agent, similar_group, path_documents, episodes=500,
                                                           batch_size=32, steps_per_test=5, replay_times=10,
                                                           is_isolated=False):
    trained_paths = set()
    update_target_every = 100
    global_steps = 0

    for episode in range(episodes):
        episode_reward = 0
        episode_similarities = []
        episode_td_errors = []

        for path_idx in similar_group:
            if path_idx in trained_paths:
                continue

            file_path = os.path.join(path_documents,
                                     f"enhanced_standard_path{path_idx + 1}{'_isolated' if is_isolated else ''}.txt")
            path_data = load_path_data(file_path)
            target_path = targetPaths[path_idx]

            BATCH_SIZE = 50
            N_SAMPLES = 200
            N_STEPS = 10
            REPLAY_TIMES = 3

            for batch_start in range(0, N_SAMPLES, BATCH_SIZE):
                batch_end = min(batch_start + BATCH_SIZE, N_SAMPLES)
                for test_data in range(batch_start, batch_end):
                    if test_data >= len(path_data):
                        break

                    step_count = 0
                    state = path_data[test_data]  #  (1-50)
                    prev_state = None
                    prev_triggered = None
                    prev_reward = None

                    for step in range(N_STEPS):
                        legal_actions = []
                        for a in range(agent.action_dim):
                            dx, dy, dz = agent.decode_action(a)
                            cand_next = (state[0] + dx, state[1] + dy, state[2] + dz)
                            if is_valid_state(cand_next):
                                legal_actions.append(a)
                        if not legal_actions:
                            break

                        if random.random() < agent.epsilon:
                            action = random.choice(legal_actions)
                        else:
                            # act
                            norm_state = normalize_state(state)
                            state_tensor = torch.tensor(norm_state, dtype=torch.float32).unsqueeze(0).to(device)
                            with torch.no_grad():
                                q_values = agent.model(state_tensor)[0]
                            action = legal_actions[torch.argmax(q_values[legal_actions]).item()]

                        dx, dy, dz = agent.decode_action(action)
                        next_state = clip_state((state[0] + dx, state[1] + dy, state[2] + dz))

                        triggered = execute_Tr(next_state[0], next_state[1], next_state[2])
                        reward = compute_reward(next_state, target_path, triggered, prev_triggered, prev_state)
                        done = (step_count == N_STEPS - 1)

                        # store_transition
                        td_error = agent.store_transition(state, action, reward, next_state, done)
                        enhanced_standard_metrics.record_step_metrics(reward, td_error, triggered, target_path,
                                                                      next_state[0], next_state[1], next_state[2])

                        episode_similarities.append(jaccard_similarity(triggered, target_path))
                        episode_td_errors.append(td_error)

                        if prev_reward is not None:
                            enhanced_standard_metrics.record_action_improvement(reward, prev_reward)

                        prev_state = state
                        prev_triggered = triggered
                        prev_reward = reward
                        state = next_state
                        step_count += 1
                        episode_reward += reward
                        global_steps += 1

                        if global_steps % update_target_every == 0:
                            agent.update_target_model()

                for _ in range(REPLAY_TIMES):
                    if len(agent.replay_buffer) >= batch_size:
                        agent.train(batch_size)

            trained_paths.add(path_idx)

        avg_similarity = np.mean(episode_similarities) if episode_similarities else 0
        avg_td_error = np.mean(episode_td_errors) if episode_td_errors else 0
        enhanced_standard_metrics.record_episode_metrics(episode, episode_reward, avg_similarity, avg_td_error,
                                                         agent.epsilon, "similar")

        if episode % 10 == 0:
            agent.update_target_model()

        if len(trained_paths) == len(similar_group):
            break

    return agent


def generate_samples_for_isolated_paths_standard(agent_similar, isolated_group_indices, num_total=2000, top_k=200):
    def compute_q_value(state, agent):
        # Q
        norm_state = normalize_state(state)
        state_tensor = torch.tensor(norm_state, dtype=torch.float32).unsqueeze(0).to(device)
        with torch.no_grad():
            q_values = agent.model(state_tensor)
        return q_values.max().item()

    def compute_robustness(state, path):
        base = execute_Tr(state[0], state[1], state[2])
        if not base: return 0.0
        rob, neighbors = 0.0, 0
        for dx in [-1, 0, 1]:
            for dy in [-1, 0, 1]:
                for dz in [-1, 0, 1]:
                    if dx == dy == dz == 0: continue
                    neighbor_state = (state[0] + dx, state[1] + dy, state[2] + dz)
                    if not is_valid_state(neighbor_state): continue
                    neighbor = clip_state(neighbor_state)
                    n_trig = execute_Tr(neighbor[0], neighbor[1], neighbor[2])
                    if not n_trig: continue
                    rob += jaccard_similarity(base, n_trig)
                    neighbors += 1
        return rob / neighbors if neighbors > 0 else 0.0

    def save_samples(path_id, samples, base_dir):
        os.makedirs(base_dir, exist_ok=True)
        filepath = os.path.join(base_dir, f"enhanced_standard_path{path_id}_isolated.txt")
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(f"Enhanced Standard Isolated Path {path_id}\n")
            f.write("x y z\tScore\tSimilarity\tLengthDiff\tRobustness\tQ_complement\n")
            for s in samples:
                x, y, z = s[0]
                f.write(f"{x} {y} {z}\t{s[1]:.4f}\t{s[2]:.4f}\t{s[3]:.4f}\t{s[4]:.4f}\t{s[5]:.4f}\n")

    base_dir = r"D:\Experiment\CNN\DQNNEW\path_samples_enhanced_standard"
    for path_idx in isolated_group_indices:
        path = targetPaths[path_idx]
        samples_raw = []  # Q
        attempts = 0

        # Run : candidatesQ
        print(f"Path  {path_idx + 1} ...")
        while len(samples_raw) < num_total and attempts < num_total * 5:
            attempts += 1
            state = (
                np.random.randint(X_MIN, X_MAX + 1),
                np.random.randint(Y_MIN, Y_MAX + 1),
                np.random.randint(Z_MIN, Z_MAX + 1)
            )
            triggered = execute_Tr(state[0], state[1], state[2])
            if not triggered: continue
            sim = jaccard_similarity(triggered, path)
            len_diff = 1 - abs(len(triggered) - len(path)) / max(len(triggered), len(path))
            rob = compute_robustness(state, path)
            q_value = compute_q_value(state, agent_similar)
            samples_raw.append((state, sim, len_diff, rob, q_value))

        # Run : Q
        if not samples_raw:
            print(f"Path  {path_idx + 1}: ")
            continue

        q_values_list = [s[4] for s in samples_raw]
        q_min = min(q_values_list)
        q_max = max(q_values_list)

        print(f"Path  {path_idx + 1}:  {len(samples_raw)} ")
        print(f"  Q: [{q_min:.4f}, {q_max:.4f}]")

        # Run : Q, 
        samples_final = []
        for state, sim, len_diff, rob, q_value in samples_raw:
            # : (q - q_min) / (q_max - q_min)
            if q_max - q_min > 1e-6:  # 
                q_normalized = (q_value - q_min) / (q_max - q_min)
            else:
                q_normalized = 0.5  # Q, 0.5

            # : 1 - q_normalized
            q_complement = 1.0 - q_normalized

            # final samples, QRun Metric
            score = 0.28 * sim + 0.1 * len_diff + 0.19 * rob + 0.43 * q_complement

            samples_final.append((state, score, sim, len_diff, rob, q_complement))

        # Run : top_k
        if samples_final:
            samples_final.sort(key=lambda x: x[1], reverse=True)
            save_samples(path_id=path_idx + 1, samples=samples_final[:top_k], base_dir=base_dir)
            print(f"Path  {path_idx + 1}:  {min(top_k, len(samples_final))} ")
            print(f"  Q_complement : [{samples_final[-1][5]:.4f}, {samples_final[0][5]:.4f}]")


# === Run ===
def enhanced_standard_generate_and_train_for_isolated_paths(agent_similar, agent_isolated, similar_group,
                                                            isolated_group, path_documents, episodes=500, batch_size=32,
                                                            is_isolated=True):
    trained_paths = set()
    update_target_every = 100
    global_steps = 0

    stage1_samples_pool = {}

    for path_idx in isolated_group:
        target_path = targetPaths[path_idx]
        # get_high_reward_samples
        high_reward_samples = agent_similar.replay_buffer.get_high_reward_samples(target_path, num_samples=100)
        stage1_samples_pool[path_idx] = high_reward_samples

    for episode in range(episodes):
        episode_reward = 0
        episode_similarities = []
        episode_td_errors = []

        for path_idx in isolated_group:
            if path_idx in trained_paths:
                continue

            file_path = os.path.join(path_documents, f"enhanced_standard_path{path_idx + 1}_isolated.txt")
            stage2_path_data = load_path_data(file_path)
            stage1_samples = stage1_samples_pool.get(path_idx, [])
            target_path = targetPaths[path_idx]

            BATCH_SIZE = 50
            N_SAMPLES_STAGE2 = min(140, len(stage2_path_data))
            N_SAMPLES_STAGE1 = min(60, len(stage1_samples))
            N_STEPS = 10
            REPLAY_TIMES = 3

            for batch_start in range(0, N_SAMPLES_STAGE2, BATCH_SIZE):
                batch_end = min(batch_start + BATCH_SIZE, N_SAMPLES_STAGE2)

                for test_data_idx in range(batch_start, batch_end):
                    if test_data_idx >= len(stage2_path_data):
                        break

                    step_count = 0
                    state = stage2_path_data[test_data_idx]  # 
                    prev_state = None
                    prev_triggered = None
                    prev_reward = None

                    for step in range(N_STEPS):
                        legal_actions = []
                        for a in range(agent_isolated.action_dim):
                            dx, dy, dz = agent_isolated.decode_action(a)
                            cand_next = (state[0] + dx, state[1] + dy, state[2] + dz)
                            if is_valid_state(cand_next):
                                legal_actions.append(a)

                        if not legal_actions:
                            break

                        if random.random() < agent_isolated.epsilon:
                            action = random.choice(legal_actions)
                        else:
                            # 
                            norm_state = normalize_state(state)
                            state_tensor = torch.tensor(norm_state, dtype=torch.float32).unsqueeze(0).to(device)
                            with torch.no_grad():
                                q_values = agent_isolated.model(state_tensor)[0]
                            action = legal_actions[torch.argmax(q_values[legal_actions]).item()]

                        dx, dy, dz = agent_isolated.decode_action(action)
                        next_state = clip_state((state[0] + dx, state[1] + dy, state[2] + dz))

                        triggered = execute_Tr(next_state[0], next_state[1], next_state[2])
                        reward = compute_reward(next_state, target_path, triggered, prev_triggered, prev_state)

                        if target_path.issubset(triggered):
                            reward += 2.0

                        done = (step_count == N_STEPS - 1)

                        td_error = agent_isolated.store_transition(state, action, reward, next_state, done)
                        enhanced_standard_metrics.record_step_metrics(reward, td_error, triggered, target_path,
                                                                      next_state[0], next_state[1], next_state[2])

                        episode_similarities.append(jaccard_similarity(triggered, target_path))
                        episode_td_errors.append(td_error)

                        if prev_reward is not None:
                            enhanced_standard_metrics.record_action_improvement(reward, prev_reward)

                        prev_state = state
                        prev_triggered = triggered
                        prev_reward = reward
                        state = next_state
                        step_count += 1
                        episode_reward += reward
                        global_steps += 1

                        if global_steps % update_target_every == 0:
                            agent_isolated.update_target_model()

            if stage1_samples:
                for sample_idx in range(N_SAMPLES_STAGE1):
                    if sample_idx >= len(stage1_samples):
                        break

                    # stage1_samples
                    stage1_state_tuple, _, _, _ = stage1_samples[sample_idx]
                    step_count = 0
                    state = stage1_state_tuple  # 
                    prev_state = None
                    prev_triggered = None
                    prev_reward = None

                    for step in range(N_STEPS):
                        legal_actions = []
                        for a in range(agent_isolated.action_dim):
                            dx, dy, dz = agent_isolated.decode_action(a)
                            cand_next = (state[0] + dx, state[1] + dy, state[2] + dz)
                            if is_valid_state(cand_next):
                                legal_actions.append(a)

                        if not legal_actions:
                            break

                        if random.random() < agent_isolated.epsilon:
                            action = random.choice(legal_actions)
                        else:
                            # 
                            norm_state = normalize_state(state)
                            state_tensor = torch.tensor(norm_state, dtype=torch.float32).unsqueeze(0).to(device)
                            with torch.no_grad():
                                q_values = agent_isolated.model(state_tensor)[0]
                            action = legal_actions[torch.argmax(q_values[legal_actions]).item()]

                        dx, dy, dz = agent_isolated.decode_action(action)
                        next_state = clip_state((state[0] + dx, state[1] + dy, state[2] + dz))

                        triggered = execute_Tr(next_state[0], next_state[1], next_state[2])
                        reward = compute_reward(next_state, target_path, triggered, prev_triggered, prev_state)
                        reward *= 0.8

                        done = (step_count == N_STEPS - 1)

                        td_error = agent_isolated.store_transition(state, action, reward, next_state, done)
                        enhanced_standard_metrics.record_step_metrics(reward, td_error, triggered, target_path,
                                                                      next_state[0], next_state[1], next_state[2])

                        episode_similarities.append(jaccard_similarity(triggered, target_path))
                        episode_td_errors.append(td_error)

                        if prev_reward is not None:
                            enhanced_standard_metrics.record_action_improvement(reward, prev_reward)

                        prev_state = state
                        prev_triggered = triggered
                        prev_reward = reward
                        state = next_state
                        step_count += 1
                        episode_reward += reward
                        global_steps += 1

                        if global_steps % update_target_every == 0:
                            agent_isolated.update_target_model()

            for replay_round in range(REPLAY_TIMES):
                if len(agent_isolated.replay_buffer) >= batch_size:
                    agent_isolated.train(batch_size)

            trained_paths.add(path_idx)

        avg_similarity = np.mean(episode_similarities) if episode_similarities else 0
        avg_td_error = np.mean(episode_td_errors) if episode_td_errors else 0
        enhanced_standard_metrics.record_episode_metrics(episode, episode_reward, avg_similarity, avg_td_error,
                                                         agent_isolated.epsilon, "isolated")

        if episode % 10 == 0:
            agent_isolated.update_target_model()

        if len(trained_paths) == len(isolated_group):
            break

    return agent_isolated


# === Excel ===
def append_metrics_to_combined_excel(metrics_collector, agent_similar, agent_isolated, similar_group, isolated_group,
                                     targetPaths, filepath, run_number):
    """Metricfinal samplesExcelsheet"""

    # ===== Metric =====
    training_time = metrics_collector.end_time - metrics_collector.start_time if metrics_collector.end_time else 0
    avg_memory = metrics_collector.total_memory_usage / metrics_collector.memory_check_count if metrics_collector.memory_check_count > 0 else 0
    avg_similarity = np.mean(
        metrics_collector.final_output_similarities) if metrics_collector.final_output_similarities else 0
    avg_td_error = np.mean(metrics_collector.td_errors) if metrics_collector.td_errors else 0
    action_improve_rate = np.mean(metrics_collector.action_improvements) if metrics_collector.action_improvements else 0

    performance_row = {
        'Run': run_number,
        'Average Similarity': f"{avg_similarity:.4f}",
        'TD': f"{avg_td_error:.4f}",
        '': f"{action_improve_rate:.4f}",
        'Training Time( seconds)': f"{training_time:.2f}",
        '(MB)': f"{avg_memory:.2f}"
    }

    # ===== final samples =====
    sample_rows = []

    # 
    for path_idx in similar_group:
        target_path = targetPaths[path_idx]
        high_reward_samples = agent_similar.replay_buffer.get_high_reward_samples(target_path, num_samples=20)

        for state_tuple, reward, sim, triggered in high_reward_samples:
            sample_rows.append({
                'Run': run_number,
                'Path ': '',
                'Path ID': path_idx + 1,
                'X': state_tuple[0],
                'Y': state_tuple[1],
                'Z': state_tuple[2],
                'Similarity': f"{sim:.4f}",
                '': f"{reward:.2f}",
                '': len(triggered),
                '': len(target_path),
                '': str(sorted(triggered)),
                '': str(sorted(target_path))
            })

    # 
    for path_idx in isolated_group:
        target_path = targetPaths[path_idx]
        high_reward_samples = agent_isolated.replay_buffer.get_high_reward_samples(target_path, num_samples=20)

        for state_tuple, reward, sim, triggered in high_reward_samples:
            sample_rows.append({
                'Run': run_number,
                'Path ': '',
                'Path ID': path_idx + 1,
                'X': state_tuple[0],
                'Y': state_tuple[1],
                'Z': state_tuple[2],
                'Similarity': f"{sim:.4f}",
                '': f"{reward:.2f}",
                '': len(triggered),
                '': len(target_path),
                '': str(sorted(triggered)),
                '': str(sorted(target_path))
            })

    # ===== Excel =====
    os.makedirs(os.path.dirname(filepath), exist_ok=True)

    # 
    if os.path.exists(filepath):
        # Metricsheet
        try:
            df_performance = pd.read_excel(filepath, sheet_name='Metric')
            df_performance = pd.concat([df_performance, pd.DataFrame([performance_row])], ignore_index=True)
        except:
            df_performance = pd.DataFrame([performance_row])

        # sheet
        try:
            df_samples = pd.read_excel(filepath, sheet_name='final samples')
            df_samples = pd.concat([df_samples, pd.DataFrame(sample_rows)], ignore_index=True)
        except:
            df_samples = pd.DataFrame(sample_rows)
    else:
        df_performance = pd.DataFrame([performance_row])
        df_samples = pd.DataFrame(sample_rows)

    # ===== Excel(sheet) =====
    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        # Metricsheet
        df_performance.to_excel(writer, sheet_name='Metric', index=False)

        # sheet
        df_samples.to_excel(writer, sheet_name='final samples', index=False)

        workbook = writer.book

        # ===== Metricsheet =====
        ws_performance = writer.sheets['Metric']

        # 
        ws_performance.column_dimensions['A'].width = 15
        for col in ['B', 'C', 'D', 'E', 'F']:
            ws_performance.column_dimensions[col].width = 20

        # 
        header_font = Font(bold=True, size=11)
        for cell in ws_performance[1]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 
        for row in ws_performance.iter_rows(min_row=2, max_row=ws_performance.max_row):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # ===== sheet =====
        ws_samples = writer.sheets['final samples']

        # 
        column_widths = {
            'A': 12, 'B': 15, 'C': 12, 'D': 10, 'E': 10, 'F': 10,
            'G': 12, 'H': 12, 'I': 12, 'J': 12, 'K': 40, 'L': 40
        }
        for col, width in column_widths.items():
            ws_samples.column_dimensions[col].width = width

        # 
        for cell in ws_samples[1]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 
        for row in ws_samples.iter_rows(min_row=2, max_row=ws_samples.max_row):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    print(f"Run {run_number} run: {filepath}")
    print(f"  - Metricsheet: ")
    print(f"  - final samplessheet: ")


# ===  run ===
def run_single_experiment(run_number, results_save_dir):
    """"""
    print(f"\n{'=' * 80}")
    print(f"Start run  {run_number}  run")
    print(f"{'=' * 80}\n")

    # Metric
    enhanced_standard_metrics.reset()
    enhanced_standard_metrics.start_training()

    model_path_similar = os.path.join(results_save_dir, f"run{run_number}_similar.pth")
    model_path_isolated = os.path.join(results_save_dir, f"run{run_number}_isolated.pth")
    path_documents = r"D:\Experiment\CNN\DQNNEW\path_samples_enhanced_standard"

    similar_group, isolated_group = group_paths_by_similarity(targetPaths)

    # Run : Path 
    if run_number == 1:
        generate_samples_for_similar_paths(similar_group, num_total=2000, top_k=200)

    replay_buffer = StandardExperienceReplay(capacity=10000)
    state_dim = 3
    action_dim = 30
    agent = StandardDQNAgent(state_dim, action_dim, replay_buffer)

    agent = enhanced_standard_generate_and_train_for_similar_paths(agent, similar_group, path_documents, episodes=500,
                                                                   batch_size=32, is_isolated=False)

    os.makedirs(os.path.dirname(model_path_similar), exist_ok=True)
    torch.save({
        'model_state_dict': agent.model.state_dict(),
        'optimizer_state_dict': agent.optimizer.state_dict(),
        'epsilon': agent.epsilon
    }, model_path_similar)

    # Run : Path 
    if run_number == 1:
        generate_samples_for_isolated_paths_standard(agent, isolated_group, num_total=2000, top_k=200)

    # Run : Path 
    isolated_replay_buffer = StandardExperienceReplay(capacity=15000)
    agent_isolated = StandardDQNAgent(state_dim, action_dim, isolated_replay_buffer)

    try:
        checkpoint = torch.load(model_path_similar)
        agent_isolated.model.load_state_dict(checkpoint['model_state_dict'])
        agent_isolated.target_model.load_state_dict(checkpoint['model_state_dict'])
        agent_isolated.epsilon = checkpoint.get('epsilon', 0.5)
    except Exception as e:
        pass

    agent_isolated = enhanced_standard_generate_and_train_for_isolated_paths(
        agent_similar=agent,
        agent_isolated=agent_isolated,
        similar_group=similar_group,
        isolated_group=isolated_group,
        path_documents=path_documents,
        episodes=500,
        batch_size=32,
        is_isolated=True
    )

    os.makedirs(os.path.dirname(model_path_isolated), exist_ok=True)
    torch.save({
        'model_state_dict': agent_isolated.model.state_dict(),
        'optimizer_state_dict': agent_isolated.optimizer.state_dict(),
        'epsilon': agent_isolated.epsilon
    }, model_path_isolated)

    # 
    enhanced_standard_metrics.end_training()

    # final samples
    for path_idx in similar_group:
        target_path = targetPaths[path_idx]
        high_reward_samples = agent.replay_buffer.get_high_reward_samples(target_path, num_samples=20)
        for state_tuple, _, sim, triggered in high_reward_samples:
            enhanced_standard_metrics.record_final_output_sample(triggered, target_path)

    for path_idx in isolated_group:
        target_path = targetPaths[path_idx]
        high_reward_samples = agent_isolated.replay_buffer.get_high_reward_samples(target_path, num_samples=20)
        for state_tuple, _, sim, triggered in high_reward_samples:
            enhanced_standard_metrics.record_final_output_sample(triggered, target_path)

    # ===== : Excel =====
    combined_excel_path = os.path.join(results_save_dir, "__.xlsx")

    append_metrics_to_combined_excel(
        metrics_collector=enhanced_standard_metrics,
        agent_similar=agent,
        agent_isolated=agent_isolated,
        similar_group=similar_group,
        isolated_group=isolated_group,
        targetPaths=targetPaths,
        filepath=combined_excel_path,
        run_number=run_number
    )

    #  runMetric
    avg_similarity = np.mean(enhanced_standard_metrics.final_output_similarities)
    training_time = enhanced_standard_metrics.end_time - enhanced_standard_metrics.start_time
    print(f"\nRun  {run_number}  runcompleted:")
    print(f"  Average Similarity: {avg_similarity:.4f}")
    print(f"  Training Time: {training_time:.2f} seconds")
    print(f"  : {enhanced_standard_metrics.step_count}")


if __name__ == "__main__":
    results_save_dir = r"D:\Experiment\CNN\_"  # 
    os.makedirs(results_save_dir, exist_ok=True)

    # 20
    NUM_RUNS = 20

    print("=" * 80)
    print(f" {NUM_RUNS} DQN()")
    print(f": X[{X_MIN}, {X_MAX}], Y[{Y_MIN}, {Y_MAX}], Z[{Z_MIN}, {Z_MAX}]")
    print("=" * 80)

    for run in range(1, NUM_RUNS + 1):
        try:
            run_single_experiment(run, results_save_dir)
        except Exception as e:
            print(f"\nRun  {run}  run: {str(e)}")
            import traceback

            traceback.print_exc()
            continue

    print("\n" + "=" * 80)
    print(f" {NUM_RUNS}  runcompleted")
    print(f": {results_save_dir}")
    print("=" * 80)