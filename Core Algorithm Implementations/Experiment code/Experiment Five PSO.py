import numpy as np
import random
import time
from typing import List, Set
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def safe_divide(numerator, denominator, default=0.0):
    """, """
    if denominator == 0:
        return default
    return numerator / denominator

def execute_validation_rules(dx: int, dy: int, dz: int) -> Set[int]:
    """Path """
    # --- 1. constants and configuration ---
    MAX_GRID_SIZE = 50.0  # 50.0
    INITIAL_BATTERY = 500.0  # 
    BATTERY_PER_STEP = 1.0  # 
    SAFE_DISTANCE = 5.0  #  ()
    CRITICAL_BATTERY_LEVEL = 50.0  #  ()
    TARGET_X, TARGET_Y, TARGET_Z = 45.0, 45.0, 20.0  # 

    MIN_PLANNING_X = 5.0
    MIN_PLANNING_Y = 7.5
    MIN_PLANNING_Z = 4.0
    CRITICAL_X_VELOCITY = 10.0
    CRITICAL_Y_VELOCITY = 12.5
    CRITICAL_Z_VELOCITY = 7.5

    triggered = set()

    # , 1-50
    current_x = random.uniform(1.0, MAX_GRID_SIZE)
    current_y = random.uniform(1.0, MAX_GRID_SIZE)
    current_z = random.uniform(1.0, MAX_GRID_SIZE)

    # ''''
    simulated_y = current_y  #  current_y  self.y 

    # --- branch 1-4 ---
    if abs(dx) < MIN_PLANNING_X != abs(dy) < MIN_PLANNING_X:
        triggered.add(1)
    if abs(dx) < MIN_PLANNING_X != abs(dz) < MIN_PLANNING_X:
        triggered.add(2)
    if abs(dx) < MIN_PLANNING_X != abs(dx) < MIN_PLANNING_Y:
        triggered.add(3)
    if abs(dx) < MIN_PLANNING_X != abs(dx) < MIN_PLANNING_Z:
        triggered.add(4)

    # --- branch 5-9 ---
    if abs(dz) > MIN_PLANNING_Z * 2 != abs(dx) > MIN_PLANNING_Z * 2:
        triggered.add(5)
    if abs(dz) > MIN_PLANNING_Z * 2 != abs(dy) > MIN_PLANNING_Z * 2:
        triggered.add(6)
    if abs(dz) > MIN_PLANNING_Z * 2 != abs(dz) > MIN_PLANNING_X * 2:
        triggered.add(7)
    if abs(dz) > MIN_PLANNING_Z * 2 != abs(dz) > MIN_PLANNING_Y * 2:
        triggered.add(8)
    if abs(dz) > MIN_PLANNING_Z * 2 != abs(dz) > MIN_PLANNING_Z:
        triggered.add(9)

    # --- branch 10-15 --- ( simulated_y  self.y)
    if TARGET_Y > simulated_y and dy < 10 != TARGET_Y > simulated_y and dy < 5:
        triggered.add(10)
    if TARGET_Y > simulated_y and dy < 10 != TARGET_Y > simulated_y and dy < 15:
        triggered.add(11)
    if TARGET_Y > simulated_y and dy < 10 != TARGET_Y > simulated_y and dy < 20:
        triggered.add(12)
    if TARGET_Y > simulated_y and dy < 10 != TARGET_Y > simulated_y and dy < 25:
        triggered.add(13)
    if TARGET_Y > simulated_y and dy < 10 != TARGET_Y > simulated_y and dx < 10:
        triggered.add(14)
    if TARGET_Y > simulated_y and dy < 10 != TARGET_Y > simulated_y and dz < 10:
        triggered.add(15)

    # --- branch 16-21 ---
    if abs(dy) > CRITICAL_X_VELOCITY * 1.5 != abs(dx) > CRITICAL_X_VELOCITY * 1.5:
        triggered.add(16)
    if abs(dy) > CRITICAL_X_VELOCITY * 1.5 != abs(dz) > CRITICAL_X_VELOCITY * 1.5:
        triggered.add(17)
    if abs(dy) > CRITICAL_X_VELOCITY * 1.5 != abs(dy) > CRITICAL_X_VELOCITY:
        triggered.add(18)
    if abs(dy) > CRITICAL_X_VELOCITY * 1.5 != abs(dy) > CRITICAL_X_VELOCITY * 2:
        triggered.add(19)
    if abs(dy) > CRITICAL_X_VELOCITY * 1.5 != abs(dy) > CRITICAL_Z_VELOCITY * 1.5:
        triggered.add(20)
    if abs(dy) > CRITICAL_X_VELOCITY * 1.5 != abs(dy) > CRITICAL_Y_VELOCITY * 1.5:
        triggered.add(21)

    # --- branch 22-29 --- ( current_x, current_y, current_z )
    if TARGET_Z < current_z and dz > CRITICAL_Z_VELOCITY != TARGET_X < current_z and dz > CRITICAL_Z_VELOCITY:
        triggered.add(22)
    if TARGET_Z < current_z and dz > CRITICAL_Z_VELOCITY != TARGET_Y < current_z and dz > CRITICAL_Z_VELOCITY:
        triggered.add(23)
    if TARGET_Z < current_z and dz > CRITICAL_Z_VELOCITY != TARGET_Z < current_x and dz > CRITICAL_Z_VELOCITY:
        triggered.add(24)
    if TARGET_Z < current_z and dz > CRITICAL_Z_VELOCITY != TARGET_Z < current_y and dz > CRITICAL_Z_VELOCITY:
        triggered.add(25)
    if TARGET_Z < current_z and dz > CRITICAL_Z_VELOCITY != TARGET_Z < current_z and dx > CRITICAL_Z_VELOCITY:
        triggered.add(26)
    if TARGET_Z < current_z and dz > CRITICAL_Z_VELOCITY != TARGET_Z < current_z and dy > CRITICAL_Z_VELOCITY:
        triggered.add(27)
    if TARGET_Z < current_z and dz > CRITICAL_Z_VELOCITY != TARGET_Z < current_z and dz > CRITICAL_X_VELOCITY:
        triggered.add(28)
    if TARGET_Z < current_z and dz > CRITICAL_Z_VELOCITY != TARGET_Z < current_z and dz > CRITICAL_Y_VELOCITY:
        triggered.add(29)

    return triggered


def calculate_fitness(particle: List[float], target_path: Set[int]) -> float:
    """"""
    generated_path = execute_validation_rules(particle[0], particle[1], particle[2])

    if target_path.issubset(generated_path):
        return 1.0

    intersection = len(generated_path & target_path)
    union = len(generated_path | target_path)
    return intersection / union if union > 0 else 0.0


class BasicPSO:
    """"""

    def __init__(self, n_particles=20, max_iterations=10000, bounds=None):
        self.n_particles = n_particles
        self.max_iterations = max_iterations
        # : x:1-50, y:1-50, z:1-50
        self.bounds = bounds if bounds else [(1, 50), (1, 50), (1, 50)]
        self.dim = len(self.bounds)
        self.w = 0.7
        self.c1 = 1.5
        self.c2 = 1.5

    def initialize_particles(self):
        """"""
        particles = []
        velocities = []

        for _ in range(self.n_particles):
            particle = [random.randint(self.bounds[i][0], self.bounds[i][1])
                        for i in range(self.dim)]
            particles.append(particle)

            velocity = [random.uniform(-5, 5) for _ in range(self.dim)]
            velocities.append(velocity)

        return particles, velocities

    def update_velocity_and_position(self, particle, velocity, pbest, gbest):
        """"""
        new_velocity = []
        new_particle = []

        r1 = np.random.random(self.dim)
        r2 = np.random.random(self.dim)

        for i in range(self.dim):
            v = (self.w * velocity[i] +
                 self.c1 * r1[i] * (pbest[i] - particle[i]) +
                 self.c2 * r2[i] * (gbest[i] - particle[i]))

            v_max = 0.2 * (self.bounds[i][1] - self.bounds[i][0])
            v = max(-v_max, min(v, v_max))
            new_velocity.append(v)

            p = particle[i] + v
            p = round(p)
            p = max(self.bounds[i][0], min(p, self.bounds[i][1]))
            new_particle.append(p)

        return new_particle, new_velocity

    def optimize(self, target_path: Set[int]):
        """target pathsPSO"""
        start_time = time.time()

        particles, velocities = self.initialize_particles()

        pbest_particles = [p.copy() for p in particles]
        pbest_fitness = [calculate_fitness(p, target_path) for p in particles]

        gbest_idx = np.argmax(pbest_fitness)
        gbest_particle = particles[gbest_idx].copy()
        gbest_fitness = pbest_fitness[gbest_idx]

        fitness_history = []

        for iteration in range(self.max_iterations):
            for i in range(self.n_particles):
                fitness = calculate_fitness(particles[i], target_path)

                if fitness == 1.0:
                    result = {
                        'success': True,
                        'best_fitness': 1.0,
                        'best_particle': particles[i].copy(),
                        'best_path': execute_validation_rules(particles[i][0], particles[i][1], particles[i][2]),
                        'iterations': iteration,
                        'time': time.time() - start_time
                    }
                    return result

                if fitness > pbest_fitness[i]:
                    pbest_fitness[i] = fitness
                    pbest_particles[i] = particles[i].copy()

                    if fitness > gbest_fitness:
                        gbest_fitness = fitness
                        gbest_particle = particles[i].copy()

                particles[i], velocities[i] = self.update_velocity_and_position(
                    particles[i], velocities[i], pbest_particles[i], gbest_particle
                )

            fitness_history.append(gbest_fitness)

        result = {
            'success': gbest_fitness == 1.0,
            'best_fitness': gbest_fitness,
            'best_particle': gbest_particle,
            'best_path': execute_validation_rules(gbest_particle[0], gbest_particle[1], gbest_particle[2]),
            'iterations': self.max_iterations,
            'time': time.time() - start_time
        }

        return result


def run_pso_for_paths(target_paths: List[Set[int]], n_particles=20, max_iterations=10000):
    """Path PSO"""

    print(f"\n{'=' * 70}")
    print(f"baseline PSO - Path ")
    print(f"{'=' * 70}")
    print(f": {n_particles}, {max_iterations}iterations")
    print(f"Path : {len(target_paths)}")
    print(f"{'=' * 70}\n")

    results = {}
    total_start = time.time()

    pso = BasicPSO(n_particles=n_particles, max_iterations=max_iterations)

    for i, target_path in enumerate(target_paths):
        print(f"Path {i + 1}: ", end='')

        result = pso.optimize(target_path)
        results[i] = result

        status = "" if result['success'] else f"({result['best_fitness']:.3f})"
        print(f"{status} | {result['time']:.2f}s | iterations{result['iterations']}")

    total_time = time.time() - total_start
    results['total_time'] = total_time

    success_count = sum(1 for i in range(len(target_paths)) if results[i]['success'])
    success_rate = (success_count / len(target_paths)) * 100

    print(f"\n{'=' * 70}")
    print(f": {success_count}/{len(target_paths)} ({success_rate:.1f}%) | Total elapsed time{total_time:.2f}s")
    print(f"{'=' * 70}\n")

    return results


def run_multiple_experiments(target_paths: List[Set[int]], num_runs=20,
                             n_particles=20, max_iterations=3000):
    """"""

    print(f"\n{'=' * 70}")
    print(f"baseline PSO - {num_runs}")
    print(f"{'=' * 70}")
    print(f": {n_particles}, {max_iterations}iterations, {len(target_paths)}Path ")
    print(f"{'=' * 70}\n")

    all_results = []
    experiment_start = time.time()

    for run_idx in range(1, num_runs + 1):
        print(f"---  {run_idx}/{num_runs} ---")

        results = run_pso_for_paths(target_paths, n_particles, max_iterations)
        all_results.append(results)

        success_count = sum(1 for i in range(len(target_paths)) if results[i]['success'])
        print(f": {success_count}/{len(target_paths)}\n")

    total_time = time.time() - experiment_start

    print(f"{'=' * 70}")
    print(f"{num_runs} runcompleted | Total elapsed time{total_time:.2f}s")
    print(f"{'=' * 70}\n")

    return all_results


def export_to_excel(all_results, target_paths, filename=None):
    """Excel"""

    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"PSO_Results_{timestamp}.xlsx"

    wb = Workbook()

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name='Microsoft YaHei', size=11, bold=True, color="FFFFFF")
    success_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    alternate_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')

    # 1: 
    ws1 = wb.active
    ws1.title = ""
    ws1.sheet_view.showGridLines = False

    headers = ["Run", "", "", "", "Average Iterations", "(s)"]
    col_widths = [12, 12, 12, 14, 14, 14]

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws1.column_dimensions[get_column_letter(col)].width = width

    for run_idx, results in enumerate(all_results, 1):
        success_count = sum(1 for i in range(len(target_paths)) if results[i]['success'])
        success_rate = (success_count / len(target_paths)) * 100
        avg_fitness = np.mean([results[i]['best_fitness'] for i in range(len(target_paths))])
        avg_iterations = np.mean([results[i]['iterations'] for i in range(len(target_paths))])
        total_time = results.get('total_time', 0)

        row_data = [
            f" {run_idx}",
            f"{success_rate:.1f}%",
            f"{success_count}/{len(target_paths)}",
            f"{avg_fitness:.4f}",
            f"{avg_iterations:.1f}",
            f"{total_time:.2f}"
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws1.cell(row=run_idx + 1, column=col, value=value)
            cell.border = border
            cell.alignment = center_align

            if run_idx % 2 == 0:
                cell.fill = alternate_fill

            if col == 2:
                if success_rate == 100.0:
                    cell.fill = success_fill
                elif success_rate < 50.0:
                    cell.fill = fail_fill

    ws1.freeze_panes = 'A2'
    ws1.auto_filter.ref = f"A1:F{len(all_results) + 1}"

    # 2: Path 
    ws2 = wb.create_sheet(title="Path ")
    ws2.sheet_view.showGridLines = False

    headers2 = ["Path ID", "", "", "", "Average Iterations", "Minimum Iterations", "Maximum Iterations"]
    col_widths2 = [12, 12, 12, 14, 14, 14, 14]

    for col, (header, width) in enumerate(zip(headers2, col_widths2), 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws2.column_dimensions[get_column_letter(col)].width = width

    for path_idx in range(len(target_paths)):
        success_count = sum(1 for r in all_results if r[path_idx]['success'])
        success_rate = (success_count / len(all_results)) * 100
        avg_fitness = np.mean([r[path_idx]['best_fitness'] for r in all_results])

        iterations_list = [r[path_idx]['iterations'] for r in all_results]
        avg_iterations = np.mean(iterations_list)
        min_iterations = np.min(iterations_list)
        max_iterations = np.max(iterations_list)

        row_data = [
            f"Path  {path_idx + 1}",
            f"{success_count}/{len(all_results)}",
            f"{success_rate:.1f}%",
            f"{avg_fitness:.4f}",
            f"{avg_iterations:.1f}",
            f"{min_iterations}",
            f"{max_iterations}"
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws2.cell(row=path_idx + 2, column=col, value=value)
            cell.border = border
            cell.alignment = center_align

            if (path_idx + 1) % 2 == 0:
                cell.fill = alternate_fill

            if col == 3:
                if success_rate == 100.0:
                    cell.fill = success_fill
                elif success_rate < 50.0:
                    cell.fill = fail_fill

    ws2.freeze_panes = 'A2'
    ws2.auto_filter.ref = f"A1:G{len(target_paths) + 1}"

    # 3: 
    ws3 = wb.create_sheet(title="")
    ws3.sheet_view.showGridLines = False

    headers3 = ["Path ", "", "(x,y,z)", "", "Iterations", "Path "]
    col_widths3 = [10, 10, 22, 12, 12, 50]

    for col, (header, width) in enumerate(zip(headers3, col_widths3), 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws3.column_dimensions[get_column_letter(col)].width = width

    row_idx = 2
    for path_idx in range(len(target_paths)):
        for run_idx, results in enumerate(all_results, 1):
            best_particle = results[path_idx]['best_particle']
            best_fitness = results[path_idx]['best_fitness']
            best_path = results[path_idx]['best_path']
            iterations = results[path_idx]['iterations']

            particle_str = f"({best_particle[0]}, {best_particle[1]}, {best_particle[2]})"
            path_str = str(sorted(list(best_path)))

            row_data = [
                f"Path {path_idx + 1}",
                f"{run_idx}",
                particle_str,
                f"{best_fitness:.4f}",
                iterations,
                path_str
            ]

            for col, value in enumerate(row_data, 1):
                cell = ws3.cell(row=row_idx, column=col, value=value)
                cell.border = border

                if col == 6:
                    cell.alignment = left_align
                else:
                    cell.alignment = center_align

                if best_fitness == 1.0:
                    cell.fill = success_fill
                elif best_fitness < 0.5:
                    cell.fill = fail_fill
                elif row_idx % 2 == 0:
                    cell.fill = alternate_fill

            row_idx += 1

    ws3.freeze_panes = 'A2'
    ws3.auto_filter.ref = f"A1:F{row_idx - 1}"

    # 4: target paths
    ws4 = wb.create_sheet(title="target paths")
    ws4.sheet_view.showGridLines = False

    headers4 = ["Path ID", "target paths", ""]
    col_widths4 = [12, 60, 12]

    for col, (header, width) in enumerate(zip(headers4, col_widths4), 1):
        cell = ws4.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws4.column_dimensions[get_column_letter(col)].width = width

    for path_idx, target_path in enumerate(target_paths):
        path_str = str(sorted(list(target_path)))

        row_data = [
            f"Path  {path_idx + 1}",
            path_str,
            len(target_path)
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws4.cell(row=path_idx + 2, column=col, value=value)
            cell.border = border

            if col == 2:
                cell.alignment = left_align
            else:
                cell.alignment = center_align

            if (path_idx + 1) % 2 == 0:
                cell.fill = alternate_fill

    ws4.freeze_panes = 'A2'

    wb.save(filename)

    print(f"\n{'=' * 70}")
    print(f" : {filename}")
    print(f"{'=' * 70}")
    print(f":")
    print(f"  1.        - {len(all_results)} run")
    print(f"  2. Path        - Path ")
    print(f"  3.    -  runPath ")
    print(f"  4. target paths       - target paths")
    print(f"{'=' * 70}\n")

    return filename


def main():
    """"""

    target_paths = [
        {1, 2, 3, 4, 10, 11, 12, 13, 14, 15, 24, 25, 26, 27, 28, 29},
        {5, 6, 7, 8, 9, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25},
        {5, 6, 7, 8, 9, 17, 18, 19, 20, 21, 24, 25, 26, 27, 28, 29}
    ]

    print("=" * 70)
    print("baseline PSO")
    print("=" * 70)

    all_results = run_multiple_experiments(
        target_paths,
        num_runs=20,
        n_particles=20,
        max_iterations=3000
    )

    export_to_excel(all_results, target_paths)

    print("Program completed")


if __name__ == "__main__":
    main()