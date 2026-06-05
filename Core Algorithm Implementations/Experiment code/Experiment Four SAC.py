
import torch
import torch.nn as nn
import torch.optim as optim
import torch.nn.functional as F
from torch.distributions import Normal
import numpy as np
import random
import time
from collections import deque
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
import os

# device setup
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
print(f"Using device: {device}")

# ===  ===
EXPERIMENT_CONFIG = {
    'STATE_DIM': 3,
    'ACTION_DIM': 3,
    # : [co2, moisture, temp]
    'MIN_VALUES': np.array([800, 10, 1], dtype=np.float32),
    'MAX_VALUES': np.array([1500, 80, 40], dtype=np.float32),
    'SAMPLES_PER_PATH': 200,
    'BATCH_SIZE_SAMPLES': 50,
    'STEPS_PER_SAMPLE': 3,
    'REPLAY_BATCH_SIZE': 64,
    'SIMILARITY_WEIGHT': 10.0,
    'COVERAGE_BONUS': 5.0,
    'TRIGGER_BONUS': 1.0,
    'HIDDEN_DIM': 256,
    'ACTOR_LR': 3e-4,
    'CRITIC_LR': 3e-4,
    'ALPHA_LR': 3e-4,
    'GAMMA': 0.99,
    'TAU': 0.005,
    'NUM_RUNS': 20,
    'TOP_K_SAMPLES': 20,
    'TARGET_PATHS': [
        {1, 2, 6, 7, 8, 9, 14, 15, 16, 20, 24, 29, 30, 31, 35, 36, 40, 45, 47, 48, 54, 59, 60, 61, 68, 74, 75, 76, 86,
         87, 92, 93, 94, 103, 107, 109, 113, 114, 115, 116, 120, 121, 123, 124, 125, 127, 130, 131, 132, 137},
        {1, 2, 6, 7, 8, 9, 14, 16, 20, 24, 29, 30, 31, 35, 36, 40, 43, 44, 45, 47, 48, 54, 56, 57, 58, 68, 74, 75, 76,
         84, 85, 92, 93, 94, 103, 107, 109, 113, 114, 115, 116, 119, 122, 125, 130, 131, 132, 137},
        {1, 2, 6, 7, 8, 9, 14, 16, 20, 24, 29, 30, 31, 35, 36, 40, 43, 44, 45, 48, 54, 59, 64, 65, 68, 74, 76, 84, 85,
         93, 94, 102, 103, 107, 109, 113, 114, 115, 116, 119, 122, 125, 131, 136},
        {1, 2, 6, 7, 8, 9, 17, 18, 19, 20, 21, 29, 30, 31, 35, 36, 37, 38, 39, 42, 45, 47, 48, 54, 56, 57, 58, 68, 74,
         75, 76, 84, 85, 92, 93, 94, 111, 112, 115, 116, 124, 125, 131, 132, 137},
        {1, 2, 6, 7, 8, 9, 14, 15, 16, 21, 29, 30, 31, 35, 36, 40, 41, 42, 45, 47, 48, 54, 59, 60, 61, 68, 74, 75, 76,
         86, 87, 92, 93, 94, 110, 111, 112, 115, 116, 128, 129, 132, 137},
        {1, 2, 6, 14, 15, 16, 21, 36, 40, 41, 42, 45, 46, 47, 48, 59, 60, 61, 71, 72, 73, 74, 77, 78, 79, 86, 87, 92,
         103, 106, 108, 110, 111, 112, 120, 121, 123, 124, 129, 132, 137},
        {6, 7, 8, 9, 14, 15, 16, 20, 24, 29, 30, 31, 35, 36, 40, 48, 54, 59, 60, 61, 68, 74, 86, 87, 92, 93, 94, 98,
         107, 109, 113, 114, 115, 116, 120, 121, 123, 124, 127, 130, 131},
        {1, 2, 4, 9, 11, 14, 15, 16, 20, 24, 26, 27, 28, 29, 32, 40, 45, 47, 48, 54, 59, 68, 74, 75, 76, 86, 87, 91, 96,
         102, 103, 107, 109, 116, 118, 120, 121, 124, 129, 132, 137},
        {1, 2, 6, 7, 8, 9, 14, 15, 16, 24, 29, 30, 31, 35, 36, 40, 52, 55, 59, 68, 74, 75, 76, 86, 87, 93, 94, 103, 109,
         114, 115, 116, 120, 121, 123, 124, 125, 126, 130, 131, 136},
        {1, 2, 6, 8, 9, 14, 20, 24, 29, 30, 31, 35, 36, 40, 43, 44, 49, 50, 52, 59, 64, 65, 68, 74, 75, 76, 84, 85, 94,
         103, 107, 109, 113, 114, 116, 119, 122, 125, 131, 132, 137},
        {1, 2, 6, 8, 9, 17, 18, 19, 20, 21, 29, 30, 31, 35, 36, 37, 38, 39, 42, 49, 50, 52, 56, 57, 58, 68, 74, 75, 76,
         84, 85, 94, 99, 111, 112, 116, 124, 125, 131, 132, 137},
        {1, 2, 6, 10, 11, 17, 18, 19, 20, 21, 32, 33, 34, 35, 36, 40, 41, 42, 45, 47, 48, 59, 60, 61, 71, 72, 73, 74,
         75, 76, 86, 87, 88, 95, 96, 103, 106, 108, 117, 132, 137},
        {1, 2, 4, 5, 9, 11, 14, 15, 16, 24, 26, 27, 28, 29, 32, 40, 45, 46, 47, 48, 54, 59, 68, 74, 79, 86, 87, 91, 96,
         102, 103, 107, 109, 116, 118, 120, 121, 124, 129, 136},
        {2, 6, 7, 8, 9, 14, 16, 20, 24, 29, 30, 31, 35, 36, 40, 48, 54, 59, 60, 61, 68, 74, 86, 87, 92, 93, 94, 97, 98,
         107, 109, 113, 114, 115, 116, 120, 123, 124, 130, 131},
        {1, 2, 6, 10, 11, 14, 15, 16, 21, 32, 33, 34, 35, 36, 40, 41, 42, 45, 47, 48, 59, 60, 61, 71, 72, 73, 74, 79,
         86, 87, 88, 89, 95, 96, 103, 106, 108, 117, 132, 137},
        {1, 2, 6, 10, 11, 12, 13, 17, 32, 33, 34, 35, 36, 40, 42, 45, 46, 47, 48, 51, 59, 61, 71, 72, 73, 74, 77, 78,
         79, 87, 88, 89, 95, 96, 103, 106, 108, 117, 119, 122},
        {1, 2, 6, 8, 9, 14, 24, 29, 30, 31, 35, 36, 40, 43, 44, 49, 50, 52, 56, 57, 58, 68, 74, 79, 84, 85, 99, 103,
         107, 109, 113, 114, 116, 119, 122, 125, 126, 133, 134},
        {6, 7, 8, 9, 14, 15, 16, 20, 24, 29, 30, 31, 35, 36, 40, 48, 54, 59, 66, 67, 68, 74, 86, 87, 93, 94, 107, 109,
         113, 114, 115, 116, 120, 121, 123, 124, 130, 131},
        {1, 2, 6, 9, 14, 21, 22, 23, 24, 25, 29, 35, 36, 37, 38, 39, 42, 50, 52, 56, 57, 58, 66, 67, 69, 70, 71, 75, 76,
         84, 85, 112, 116, 124, 125, 126},
        {1, 2, 6, 10, 11, 14, 15, 16, 21, 32, 34, 35, 36, 40, 41, 42, 52, 53, 59, 71, 73, 74, 78, 79, 86, 87, 88, 89,
         96, 101, 103, 120, 121, 124, 135},
        {1, 2, 6, 14, 21, 36, 37, 38, 39, 42, 49, 50, 52, 56, 57, 58, 71, 72, 73, 74, 77, 78, 79, 84, 85, 99, 104, 112,
         119, 122, 125, 126, 127, 133},
        {1, 2, 4, 5, 9, 11, 12, 13, 17, 24, 26, 27, 28, 29, 32, 40, 45, 46, 47, 48, 51, 54, 59, 68, 74, 79, 88, 89, 96,
         100, 105, 116, 118, 119, 122},
        {1, 2, 6, 9, 14, 29, 35, 36, 40, 44, 59, 64, 65, 66, 67, 69, 70, 71, 75, 76, 80, 81, 84, 85, 104, 114, 116, 119,
         122, 125, 126},
        {1, 2, 6, 9, 14, 20, 24, 29, 30, 35, 36, 40, 44, 59, 64, 65, 68, 74, 75, 76, 80, 81, 82, 84, 85, 104, 114, 116,
         119, 122, 125},
        {1, 2, 6, 14, 16, 21, 36, 37, 38, 39, 42, 52, 53, 59, 62, 63, 71, 73, 74, 84, 85, 92, 101, 103, 111, 112, 119,
         122, 129, 135},
        {1, 2, 6, 10, 11, 14, 15, 16, 21, 32, 35, 36, 40, 41, 42, 52, 53, 59, 71, 74, 79, 86, 87, 90, 96, 103, 120, 121,
         124, 135},
        {1, 2, 6, 9, 14, 24, 29, 30, 35, 36, 40, 44, 59, 64, 65, 68, 74, 75, 76, 83, 84, 85, 104, 114, 116, 119, 122,
         125, 126},
        {3, 21, 22, 23, 24, 25, 35, 39, 40, 52, 58, 59, 66, 67, 69, 70, 71, 79, 94, 129}
    ],
}


# ===  ===
def clip_state(state):
    """"""
    min_vals = EXPERIMENT_CONFIG['MIN_VALUES']
    max_vals = EXPERIMENT_CONFIG['MAX_VALUES']
    return np.clip(state, min_vals, max_vals)


def normalize_state(state):
    """[-1, 1]"""
    min_vals = EXPERIMENT_CONFIG['MIN_VALUES']
    max_vals = EXPERIMENT_CONFIG['MAX_VALUES']
    return 2 * (state - min_vals) / (max_vals - min_vals) - 1


def denormalize_state(normalized_state):
    """"""
    min_vals = EXPERIMENT_CONFIG['MIN_VALUES']
    max_vals = EXPERIMENT_CONFIG['MAX_VALUES']
    return (normalized_state + 1) * (max_vals - min_vals) / 2 + min_vals


def coverage_similarity(triggered, target_path):
    if len(target_path) == 0:
        return 1.0 if len(triggered) == 0 else 0.0

    intersection = target_path.intersection(triggered)
    return len(intersection) / len(target_path)


def unified_reward_function(triggered, target_path):
    config = EXPERIMENT_CONFIG
    similarity = coverage_similarity(triggered, target_path)
    reward = similarity * config['SIMILARITY_WEIGHT']

    if target_path.issubset(triggered):
        reward += config['COVERAGE_BONUS']

    if len(triggered) > 0:
        reward += config['TRIGGER_BONUS']

    return reward


def safe_divide(a, b):
    """, """
    return a / b if b != 0 else 0


def section6_low_co2_extremes(moisture, co2, temp):
    """CO2branch"""
    triggered = set()
    b = [0] * 137  # branch

    if (co2 < 1150) != (co2 < 1000):
        b[0] = 1
        triggered.add(1)
    if (co2 < 1150) != (co2 < 950):
        b[1] = 2
        triggered.add(2)
    if (co2 < 1150) != (co2 < 1400):
        b[2] = 3
        triggered.add(3)

    # branch4-11: 
    if (co2 < 1150 and moisture > 45 and temp > 20) != (co2 < 1050 and moisture > 45 and temp > 20):
        b[3] = 4
        triggered.add(4)
    if (co2 < 1150 and moisture > 45 and temp > 20) != (co2 < 1100 and moisture > 45 and temp > 20):
        b[4] = 5
        triggered.add(5)
    if (co2 < 1150 and moisture > 45 and temp > 20) != (co2 < 1150 or moisture > 45 and temp > 20):
        b[5] = 6
        triggered.add(6)
    if (co2 < 1150 and moisture > 45 and temp > 20) != (co2 < 1150 and moisture > 35 and temp > 20):
        b[6] = 7
        triggered.add(7)
    if (co2 < 1150 and moisture > 45 and temp > 20) != (co2 < 1150 and moisture > 30 and temp > 20):
        b[7] = 8
        triggered.add(8)
    if (co2 < 1150 and moisture > 45 and temp > 20) != (co2 < 1150 and moisture < 45 and temp > 20):
        b[8] = 9
        triggered.add(9)
    if (co2 < 1150 and moisture > 45 and temp > 20) != (co2 < 1150 and moisture > 45 and temp > 5):
        b[9] = 10
        triggered.add(10)
    if (co2 < 1150 and moisture > 45 and temp > 20) != (co2 < 1150 and moisture > 45 and temp < 20):
        b[10] = 11
        triggered.add(11)

    # branch12-21: 
    if ((co2 < 1150 and moisture > 50) or (co2 < 1100 and temp < 22)) != (
            (co2 < 970 and moisture > 50) or (co2 < 1100 and temp < 22)):
        b[11] = 12
        triggered.add(12)
    if ((co2 < 1150 and moisture > 50) or (co2 < 1100 and temp < 22)) != (
            (co2 < 1000 and moisture > 50) or (co2 < 1100 and temp < 22)):
        b[12] = 13
        triggered.add(13)
    if ((co2 < 1150 and moisture > 50) or (co2 < 1100 and temp < 22)) != (
            (co2 < 1150 or moisture > 50) or (co2 < 1100 and temp < 22)):
        b[13] = 14
        triggered.add(14)
    if ((co2 < 1150 and moisture > 50) or (co2 < 1100 and temp < 22)) != (
            (co2 < 1150 and moisture > 40) or (co2 < 1100 and temp < 22)):
        b[14] = 15
        triggered.add(15)
    if ((co2 < 1150 and moisture > 50) or (co2 < 1100 and temp < 22)) != (
            (co2 < 1150 and moisture > 35) or (co2 < 1100 and temp < 22)):
        b[15] = 16
        triggered.add(16)
    if ((co2 < 1150 and moisture > 50) or (co2 < 1100 and temp < 22)) != (
            (co2 < 1150 and moisture > 50) and (co2 < 1100 and temp < 22)):
        b[16] = 17
        triggered.add(17)
    if ((co2 < 1150 and moisture > 50) or (co2 < 1100 and temp < 22)) != (
            (co2 < 1150 and moisture > 50) or (co2 < 960 and temp < 22)):
        b[17] = 18
        triggered.add(18)
    if ((co2 < 1150 and moisture > 50) or (co2 < 1100 and temp < 22)) != (
            (co2 < 1150 and moisture > 50) or (co2 < 1100 and temp < 18)):
        b[18] = 19
        triggered.add(19)
    if ((co2 < 1150 and moisture > 50) or (co2 < 1100 and temp < 22)) != (
            (co2 < 1150 and moisture > 50) or (co2 < 1100 and temp > 22)):
        b[19] = 20
        triggered.add(20)
    if ((co2 < 1150 and moisture > 50) or (co2 < 1100 and temp < 22)) != (
            (co2 < 1150 and moisture > 50) or (co2 > 1100 and temp < 22)):
        b[20] = 21
        triggered.add(21)

    # branch22-25: 
    if (safe_divide(moisture, co2 - 700) > 0.06 and temp < 22) != (safe_divide(moisture, co2 - 700) > 0.04 and temp < 22):
        b[21] = 22
        triggered.add(22)
    if (safe_divide(moisture, co2 - 700) > 0.06 and temp < 22) != (safe_divide(moisture, co2 - 700) > 0.03 and temp < 22):
        b[22] = 23
        triggered.add(23)
    if (safe_divide(moisture, co2 - 700) > 0.06 and temp < 22) != (safe_divide(moisture, co2 - 700) > 0.06 or temp < 22):
        b[23] = 24
        triggered.add(24)
    if (safe_divide(moisture, co2 - 700) > 0.06 and temp < 22) != (moisture + safe_divide(100, co2 - 700) > 0.06 and temp < 22):
        b[24] = 25
        triggered.add(25)

    # branch26-36: 
    if (co2 < 1150 and moisture > 45 and temp > 20) != (co2 < 1000 and moisture > 45 and temp > 20):
        b[25] = 26
        triggered.add(26)
    if (co2 < 1150 and moisture > 45 and temp > 20) != (co2 < 1070 and moisture > 45 and temp > 20):
        b[26] = 27
        triggered.add(27)
    if (co2 < 1150 and moisture > 45 and temp > 20) != (co2 > 1150 and moisture > 45 and temp > 20):
        b[27] = 28
        triggered.add(28)
    if (co2 < 1150 and moisture > 45 and temp > 20) != (co2 < 1150 and moisture < 45 and temp > 20):
        b[28] = 29
        triggered.add(29)
    if (co2 < 1150 and moisture > 45 and temp > 20) != (co2 < 1150 and moisture > 25 and temp > 20):
        b[29] = 30
        triggered.add(30)
    if (co2 < 1150 and moisture > 45 and temp > 20) != (co2 < 1150 and moisture > 30 and temp > 20):
        b[30] = 31
        triggered.add(31)
    if (co2 < 1150 and moisture > 45 and temp > 20) != (co2 < 1150 and moisture > 45 and temp < 20):
        b[31] = 32
        triggered.add(32)
    if (co2 < 1150 and moisture > 45 and temp > 20) != (co2 < 1150 and moisture > 45 and temp > 15):
        b[32] = 33
        triggered.add(33)
    if (co2 < 1150 and moisture > 45 and temp > 20) != (co2 < 1150 and moisture > 45 and temp > 12):
        b[33] = 34
        triggered.add(34)
    if (co2 < 1150 and moisture > 45 and temp > 20) != (co2 < 1150 and moisture > 45 or temp > 20):
        b[34] = 35
        triggered.add(35)
    if (co2 < 1150 and moisture > 45 and temp > 20) != (co2 < 1150 or moisture > 45 and temp > 20):
        b[35] = 36
        triggered.add(36)

    # branch37-44: 
    if (co2 < 1150 and moisture < 40 and temp < 22) != (co2 < 970 and moisture < 40 and temp < 22):
        b[36] = 37
        triggered.add(37)
    if (co2 < 1150 and moisture < 40 and temp < 22) != (co2 < 1020 and moisture < 40 and temp < 22):
        b[37] = 38
        triggered.add(38)
    if (co2 < 1150 and moisture < 40 and temp < 22) != (co2 > 1150 and moisture < 40 and temp < 22):
        b[38] = 39
        triggered.add(39)
    if (co2 < 1150 and moisture < 40 and temp < 22) != (co2 < 1150 or moisture < 40 and temp < 22):
        b[39] = 40
        triggered.add(40)
    if (co2 < 1150 and moisture < 40 and temp < 22) != (co2 < 1150 and moisture < 50 and temp < 22):
        b[40] = 41
        triggered.add(41)
    if (co2 < 1150 and moisture < 40 and temp < 22) != (co2 < 1150 and moisture > 40 and temp < 22):
        b[41] = 42
        triggered.add(42)
    if (co2 < 1150 and moisture < 40 and temp < 22) != (co2 < 1150 and moisture < 40 and temp < 27):
        b[42] = 43
        triggered.add(43)
    if (co2 < 1150 and moisture < 40 and temp < 22) != (co2 < 1150 and moisture < 40 and temp < 32):
        b[43] = 44
        triggered.add(44)

    # branch45-55: 
    if (co2 < 1150 and 35 < moisture < 60 and 15 < temp < 28) != (
            co2 < 1050 and 35 < moisture < 60 and 15 < temp < 28):
        b[44] = 45
        triggered.add(45)
    if (co2 < 1150 and 35 < moisture < 60 and 15 < temp < 28) != (
            co2 < 1120 and 35 < moisture < 60 and 15 < temp < 28):
        b[45] = 46
        triggered.add(46)
    if (co2 < 1150 and 35 < moisture < 60 and 15 < temp < 28) != (
            co2 < 1070 and 35 < moisture < 60 and 15 < temp < 28):
        b[46] = 47
        triggered.add(47)
    if (co2 < 1150 and 35 < moisture < 60 and 15 < temp < 28) != (
            co2 > 1150 and 35 < moisture < 60 and 15 < temp < 28):
        b[47] = 48
        triggered.add(48)
    if (co2 < 1150 and 35 < moisture < 60 and 15 < temp < 28) != (
            co2 < 1150 and 25 < moisture < 60 and 15 < temp < 28):
        b[48] = 49
        triggered.add(49)
    if (co2 < 1150 and 35 < moisture < 60 and 15 < temp < 28) != (
            co2 < 1150 and 20 < moisture < 60 and 15 < temp < 28):
        b[49] = 50
        triggered.add(50)
    if (co2 < 1150 and 35 < moisture < 60 and 15 < temp < 28) != (
            co2 < 1150 and 35 < moisture < 52 and 15 < temp < 28):
        b[50] = 51
        triggered.add(51)
    if (co2 < 1150 and 35 < moisture < 60 and 15 < temp < 28) != (
            co2 < 1150 and 35 < moisture < 60 or 15 < temp < 28):
        b[51] = 52
        triggered.add(52)
    if (co2 < 1150 and 35 < moisture < 60 and 15 < temp < 28) != (
            co2 < 1150 and 35 < moisture < 60 and 10 < temp < 28):
        b[52] = 53
        triggered.add(53)
    if (co2 < 1150 and 35 < moisture < 60 and 15 < temp < 28) != (
            co2 < 1150 and 35 < moisture < 60 and 15 < temp < 20):
        b[53] = 54
        triggered.add(54)
    if (co2 < 1150 and 35 < moisture < 60 and 15 < temp < 28) != (
            co2 < 1150 and 35 < moisture < 60 and 15 < temp < 33):
        b[54] = 55
        triggered.add(55)

    # branch56-65: 
    if (co2 < 1150 and moisture < 40 and 15 < temp < 25) != (co2 < 1020 and moisture < 40 and 15 < temp < 25):
        b[55] = 56
        triggered.add(56)
    if (co2 < 1150 and moisture < 40 and 15 < temp < 25) != (co2 < 1070 and moisture < 40 and 15 < temp < 25):
        b[56] = 57
        triggered.add(57)
    if (co2 < 1150 and moisture < 40 and 15 < temp < 25) != (co2 > 1150 and moisture < 40 and 15 < temp < 25):
        b[57] = 58
        triggered.add(58)
    if (co2 < 1150 and moisture < 40 and 15 < temp < 25) != (co2 < 1150 or moisture < 40 and 15 < temp < 25):
        b[58] = 59
        triggered.add(59)
    if (co2 < 1150 and moisture < 40 and 15 < temp < 25) != (co2 < 1150 and moisture < 50 and 15 < temp < 25):
        b[59] = 60
        triggered.add(60)
    if (co2 < 1150 and moisture < 40 and 15 < temp < 25) != (co2 < 1150 and moisture < 55 and 15 < temp < 25):
        b[60] = 61
        triggered.add(61)
    if (co2 < 1150 and moisture < 40 and 15 < temp < 25) != (co2 < 1150 and moisture < 40 and 10 < temp < 25):
        b[61] = 62
        triggered.add(62)
    if (co2 < 1150 and moisture < 40 and 15 < temp < 25) != (co2 < 1150 and moisture < 40 and 5 < temp < 25):
        b[62] = 63
        triggered.add(63)
    if (co2 < 1150 and moisture < 40 and 15 < temp < 25) != (co2 < 1150 and moisture < 40 and 15 < temp < 30):
        b[63] = 64
        triggered.add(64)
    if (co2 < 1150 and moisture < 40 and 15 < temp < 25) != (co2 < 1150 and moisture < 40 and 15 < temp < 30):
        b[64] = 65
        triggered.add(65)

    # branch66-74: temperature
    if (safe_divide(moisture, co2 - 700) > 0.06 and temp > 20) != (safe_divide(moisture, co2 - 850) > 0.06 and temp > 20):
        b[65] = 66
        triggered.add(66)
    if (safe_divide(moisture, co2 - 700) > 0.06 and temp > 20) != (safe_divide(moisture, co2 - 900) > 0.06 and temp > 20):
        b[66] = 67
        triggered.add(67)
    if (safe_divide(moisture, co2 - 700) > 0.06 and temp > 20) != (safe_divide(moisture, co2 + 200) > 0.06 and temp > 20):
        b[67] = 68
        triggered.add(68)
    if (safe_divide(moisture, co2 - 700) > 0.06 and temp > 20) != (safe_divide(moisture, co2 - 700) > 0.04 and temp > 20):
        b[68] = 69
        triggered.add(69)
    if (safe_divide(moisture, co2 - 700) > 0.06 and temp > 20) != (safe_divide(moisture, co2 - 700) > 0.03 and temp > 20):
        b[69] = 70
        triggered.add(70)
    if (safe_divide(moisture, co2 - 700) > 0.06 and temp > 20) != (safe_divide(moisture, co2 - 700) > 0.06 or temp > 20):
        b[70] = 71
        triggered.add(71)
    if (safe_divide(moisture, co2 - 700) > 0.06 and temp > 20) != (safe_divide(moisture, co2 - 700) > 0.06 and temp > 15):
        b[71] = 72
        triggered.add(72)
    if (safe_divide(moisture, co2 - 700) > 0.06 and temp > 20) != (safe_divide(moisture, co2 - 700) > 0.06 and temp > 12):
        b[72] = 73
        triggered.add(73)
    if (safe_divide(moisture, co2 - 700) > 0.06 and temp > 20) != (safe_divide(moisture, co2 - 700) > 0.06 and temp < 20):
        b[73] = 74
        triggered.add(74)

    if (co2 + moisture > 1150 and temp > 18) != (co2 + moisture > 1100 and temp > 18):
        b[74] = 75
        triggered.add(75)
    if (co2 + moisture > 1150 and temp > 18) != (co2 + moisture > 1050 and temp > 18):
        b[75] = 76
        triggered.add(76)
    if (co2 + moisture > 1150 and temp > 18) != (co2 + moisture > 1150 and temp > 15):
        b[76] = 77
        triggered.add(77)
    if (co2 + moisture > 1150 and temp > 18) != (co2 + moisture > 1150 and temp > 12):
        b[77] = 78
        triggered.add(78)
    if (co2 + moisture > 1150 and temp > 18) != (co2 + moisture > 1150 and temp < 18):
        b[78] = 79
        triggered.add(79)

    if (co2 < 1150 and abs(moisture - 45) > 15 and abs(temp - 20) > 7) != (
            co2 < 1000 and abs(moisture - 45) > 15 and abs(temp - 20) > 7):
        b[79] = 80
        triggered.add(80)
    if (co2 < 1150 and abs(moisture - 45) > 15 and abs(temp - 20) > 7) != (
            co2 < 1050 and abs(moisture - 45) > 15 and abs(temp - 20) > 7):
        b[80] = 81
        triggered.add(81)
    if (co2 < 1150 and abs(moisture - 45) > 15 and abs(temp - 20) > 7) != (
            co2 < 1150 and abs(moisture - 42) > 15 and abs(temp - 20) > 7):
        b[81] = 82
        triggered.add(82)
    if (co2 < 1150 and abs(moisture - 45) > 15 and abs(temp - 20) > 7) != (
            co2 < 1150 and abs(moisture - 45) > 12 and abs(temp - 20) > 7):
        b[82] = 83
        triggered.add(83)

    # branch84-87: 
    if (co2 < 1150 and moisture < 40) != (co2 < 1000 and moisture < 40):
        b[83] = 84
        triggered.add(84)
    if (co2 < 1150 and moisture < 40) != (co2 < 1050 and moisture < 40):
        b[84] = 85
        triggered.add(85)
    if (co2 < 1150 and moisture < 40) != (co2 < 1150 and moisture < 50):
        b[85] = 86
        triggered.add(86)
    if (co2 < 1150 and moisture < 40) != (co2 < 1150 and moisture < 55):
        b[86] = 87
        triggered.add(87)

    # branch88-92: 
    if (co2 < 1150 and 12 < temp < 25 and moisture > 45) != (co2 < 1030 and 12 < temp < 25 and moisture > 45):
        b[87] = 88
        triggered.add(88)
    if (co2 < 1150 and 12 < temp < 25 and moisture > 45) != (co2 < 1080 and 12 < temp < 25 and moisture > 45):
        b[88] = 89
        triggered.add(89)
    if (co2 < 1150 and 12 < temp < 25 and moisture > 45) != (co2 < 1150 and 8 < temp < 25 and moisture > 45):
        b[89] = 90
        triggered.add(90)
    if (co2 < 1150 and 12 < temp < 25 and moisture > 45) != (co2 < 1150 and 12 < temp < 30 and moisture > 45):
        b[90] = 91
        triggered.add(91)
    if (co2 < 1150 and 12 < temp < 25 and moisture > 45) != (co2 < 1150 and 12 < temp < 25 and moisture > 35):
        b[91] = 92
        triggered.add(92)

    # branch93-96: 
    if (moisture > 45 and temp > 20) != (moisture > 35 and temp > 20):
        b[92] = 93
        triggered.add(93)
    if (moisture > 45 and temp > 20) != (moisture > 32 and temp > 20):
        b[93] = 94
        triggered.add(94)
    if (moisture > 45 and temp > 20) != (moisture > 45 and temp > 15):
        b[94] = 95
        triggered.add(95)
    if (moisture > 45 and temp > 20) != (moisture > 45 and temp < 20):
        b[95] = 96
        triggered.add(96)

    # branch97-102: 
    if (1000 < co2 < 1200 and 35 < moisture < 55 and 15 < temp < 25) != (
            950 < co2 < 1250 and 35 < moisture < 55 and 15 < temp < 25):
        b[96] = 97
        triggered.add(97)
    if (1000 < co2 < 1200 and 35 < moisture < 55 and 15 < temp < 25) != (
            900 < co2 < 1300 and 35 < moisture < 55 and 15 < temp < 25):
        b[97] = 98
        triggered.add(98)
    if (1000 < co2 < 1200 and 35 < moisture < 55 and 15 < temp < 25) != (
            1000 < co2 < 1200 and 28 < moisture < 55 and 15 < temp < 25):
        b[98] = 99
        triggered.add(99)
    if (1000 < co2 < 1200 and 35 < moisture < 55 and 15 < temp < 25) != (
            1000 < co2 < 1200 and 35 < moisture < 62 and 15 < temp < 25):
        b[99] = 100
        triggered.add(100)
    if (1000 < co2 < 1200 and 35 < moisture < 55 and 15 < temp < 25) != (
            1000 < co2 < 1200 and 35 < moisture < 55 and 12 < temp < 25):
        b[100] = 101
        triggered.add(101)
    if (1000 < co2 < 1200 and 35 < moisture < 55 and 15 < temp < 25) != (
            1000 < co2 < 1200 and 35 < moisture < 55 and 15 < temp < 28):
        b[101] = 102
        triggered.add(102)

    # branch103-109: 
    if (co2 < 1150 and 30 < moisture < 55 and (temp < 20 or temp > 22)) != (
            co2 < 1050 and 30 < moisture < 55 and (temp < 20 or temp > 22)):
        b[102] = 103
        triggered.add(103)
    if (co2 < 1150 and 30 < moisture < 55 and (temp < 20 or temp > 22)) != (
            co2 < 1150 and 22 < moisture < 55 and (temp < 20 or temp > 22)):
        b[103] = 104
        triggered.add(104)
    if (co2 < 1150 and 30 < moisture < 55 and (temp < 20 or temp > 22)) != (
            co2 < 1150 and 30 < moisture < 62 and (temp < 20 or temp > 22)):
        b[104] = 105
        triggered.add(105)
    if (co2 < 1150 and 30 < moisture < 55 and (temp < 20 or temp > 22)) != (
            co2 < 1150 and 30 < moisture < 55 and (temp < 16 or temp > 22)):
        b[105] = 106
        triggered.add(106)
    if (co2 < 1150 and 30 < moisture < 55 and (temp < 20 or temp > 22)) != (
            co2 < 1150 and 30 < moisture < 55 and (temp < 20 or temp > 26)):
        b[106] = 107
        triggered.add(107)
    if (co2 < 1150 and 30 < moisture < 55 and (temp < 20 or temp > 22)) != (
            co2 < 1150 and 30 < moisture < 55 and (temp < 15 or temp > 22)):
        b[107] = 108
        triggered.add(108)
    if (co2 < 1150 and 30 < moisture < 55 and (temp < 20 or temp > 22)) != (
            co2 < 1150 and 30 < moisture < 55 and (temp < 20 or temp > 28)):
        b[108] = 109
        triggered.add(109)

    # branch110-114: 
    if (moisture < 45 and co2 < 1150 and temp < 22) != (moisture < 38 and co2 < 1150 and temp < 22):
        b[109] = 110
        triggered.add(110)
    if (moisture < 45 and co2 < 1150 and temp < 22) != (moisture < 35 and co2 < 1150 and temp < 22):
        b[110] = 111
        triggered.add(111)
    if (moisture < 45 and co2 < 1150 and temp < 22) != (moisture < 45 and co2 < 1000 and temp < 22):
        b[111] = 112
        triggered.add(112)
    if (moisture < 45 and co2 < 1150 and temp < 22) != (moisture < 45 and co2 < 1150 and temp < 27):
        b[112] = 113
        triggered.add(113)
    if (moisture < 45 and co2 < 1150 and temp < 22) != (moisture < 45 and co2 < 1150 and temp < 30):
        b[113] = 114
        triggered.add(114)

    # branch115-118: 
    if (co2 < 1200 and moisture > 45 and temp > 20 and co2 < 1150) != (
            co2 < 1200 and moisture > 35 and temp > 20 and co2 < 1150):
        b[114] = 115
        triggered.add(115)
    if (co2 < 1200 and moisture > 45 and temp > 20 and co2 < 1150) != (
            co2 < 1200 and moisture < 45 and temp > 20 and co2 < 1150):
        b[115] = 116
        triggered.add(116)
    if (co2 < 1200 and moisture > 45 and temp > 20 and co2 < 1150) != (
            co2 < 1200 and moisture > 45 and temp > 15 and co2 < 1150):
        b[116] = 117
        triggered.add(117)
    if (co2 < 1200 and moisture > 45 and temp > 20 and co2 < 1150) != (
            co2 < 1200 and moisture > 45 and temp > 20 and co2 < 1070):
        b[117] = 118
        triggered.add(118)

    # branch119-124: 
    if (co2 < 1150 and (moisture > 50 or moisture < 40) and (temp > 22 or temp < 18)) != (
            co2 < 1000 and (moisture > 50 or moisture < 40) and (temp > 22 or temp < 18)):
        b[118] = 119
        triggered.add(119)
    if (co2 < 1150 and (moisture > 50 or moisture < 40) and (temp > 22 or temp < 18)) != (
            co2 < 1150 and (moisture > 30 or moisture < 40) and (temp > 22 or temp < 18)):
        b[119] = 120
        triggered.add(120)
    if (co2 < 1150 and (moisture > 50 or moisture < 40) and (temp > 22 or temp < 18)) != (
            co2 < 1150 and (moisture > 40 or moisture < 40) and (temp > 22 or temp < 18)):
        b[120] = 121
        triggered.add(121)
    if (co2 < 1150 and (moisture > 50 or moisture < 40) and (temp > 22 or temp < 18)) != (
            co2 < 1150 and (moisture > 50 and moisture < 40) and (temp > 22 or temp < 18)):
        b[121] = 122
        triggered.add(122)
    if (co2 < 1150 and (moisture > 50 or moisture < 40) and (temp > 22 or temp < 18)) != (
            co2 < 1150 and (moisture > 50 or moisture < 45) and (temp > 22 or temp < 18)):
        b[122] = 123
        triggered.add(123)
    if (co2 < 1150 and (moisture > 50 or moisture < 40) and (temp > 22 or temp < 18)) != (
            co2 < 1150 and (moisture > 50 or moisture < 40) or (temp > 22 or temp < 18)):
        b[123] = 124
        triggered.add(124)

    # branch125-131: 
    if (co2 < 1150 and safe_divide(temp, moisture + 1) > 0.5 and moisture < 45) != (
            co2 < 1050 and safe_divide(temp, moisture + 1) > 0.5 and moisture < 45):
        b[124] = 125
        triggered.add(125)
    if (co2 < 1150 and safe_divide(temp, moisture + 1) > 0.5 and moisture < 45) != (
            co2 < 1100 and safe_divide(temp, moisture + 1) > 0.5 and moisture < 45):
        b[125] = 126
        triggered.add(126)
    if (co2 < 1150 and safe_divide(temp, moisture + 1) > 0.5 and moisture < 45) != (
            co2 < 1150 and safe_divide(temp, moisture + 5) > 0.5 and moisture < 45):
        b[126] = 127
        triggered.add(127)
    if (co2 < 1150 and safe_divide(temp, moisture + 1) > 0.5 and moisture < 45) != (
            co2 < 1150 and safe_divide(temp, moisture + 1) > 0.4 and moisture < 45):
        b[127] = 128
        triggered.add(128)
    if (co2 < 1150 and safe_divide(temp, moisture + 1) > 0.5 and moisture < 45) != (
            co2 < 1150 and safe_divide(temp, moisture + 1) > 0.5 or moisture < 45):
        b[128] = 129
        triggered.add(129)
    if (co2 < 1150 and safe_divide(temp, moisture + 1) > 0.5 and moisture < 45) != (
            co2 < 1150 and safe_divide(temp, moisture + 1) > 0.5 and moisture < 38):
        b[129] = 130
        triggered.add(130)
    if (co2 < 1150 and safe_divide(temp, moisture + 1) > 0.5 and moisture < 45) != (
            co2 < 1150 and safe_divide(temp, moisture + 1) > 0.5 and moisture < 35):
        b[130] = 131
        triggered.add(131)

    # branch132-137: 
    if (co2 < 1150 and 35 <= moisture <= 50 and 15 <= temp <= 25) != (
            co2 < 1000 and 35 <= moisture <= 50 and 15 <= temp <= 25):
        b[131] = 132
        triggered.add(132)
    if (co2 < 1150 and 35 <= moisture <= 50 and 15 <= temp <= 25) != (
            co2 < 1150 and 28 <= moisture <= 50 and 15 <= temp <= 25):
        b[132] = 133
        triggered.add(133)
    if (co2 < 1150 and 35 <= moisture <= 50 and 15 <= temp <= 25) != (
            co2 < 1150 and 32 <= moisture <= 50 and 15 <= temp <= 25):
        b[133] = 134
        triggered.add(134)
    if (co2 < 1150 and 35 <= moisture <= 50 and 15 <= temp <= 25) != (
            co2 < 1150 and 35 <= moisture <= 50 and 12 <= temp <= 25):
        b[134] = 135
        triggered.add(135)
    if (co2 < 1150 and 35 <= moisture <= 50 and 15 <= temp <= 25) != (
            co2 < 1150 and 35 <= moisture <= 50 and 15 <= temp <= 28):
        b[135] = 136
        triggered.add(136)
    if (co2 < 1150 and 35 <= moisture <= 50 and 15 <= temp <= 25) != (
            co2 < 1050 and 35 <= moisture <= 50 and 15 <= temp <= 25):
        b[136] = 137
        triggered.add(137)

    return triggered


# === SAC Actor ===
class GaussianPolicy(nn.Module):
    def __init__(self, state_dim, action_dim, hidden_dim=256):
        super(GaussianPolicy, self).__init__()

        self.fc1 = nn.Linear(state_dim, hidden_dim)
        self.fc2 = nn.Linear(hidden_dim, hidden_dim)

        self.mean = nn.Linear(hidden_dim, action_dim)
        self.log_std = nn.Linear(hidden_dim, action_dim)

        self.action_scale = 8.0
        self.action_bias = 0.0

    def forward(self, state):
        x = F.relu(self.fc1(state))
        x = F.relu(self.fc2(x))

        mean = self.mean(x)
        log_std = self.log_std(x)
        log_std = torch.clamp(log_std, min=-20, max=2)

        return mean, log_std

    def sample(self, state):
        mean, log_std = self.forward(state)
        std = log_std.exp()
        normal = Normal(mean, std)

        x_t = normal.rsample()
        action = torch.tanh(x_t) * self.action_scale + self.action_bias

        log_prob = normal.log_prob(x_t)
        log_prob -= torch.log(self.action_scale * (1 - torch.tanh(x_t).pow(2)) + 1e-6)
        log_prob = log_prob.sum(1, keepdim=True)

        mean = torch.tanh(mean) * self.action_scale + self.action_bias

        return action, log_prob, mean


# === SAC Critic ===
class QNetwork(nn.Module):
    def __init__(self, state_dim, action_dim, hidden_dim=256):
        super(QNetwork, self).__init__()

        self.fc1 = nn.Linear(state_dim + action_dim, hidden_dim)
        self.fc2 = nn.Linear(hidden_dim, hidden_dim)
        self.fc3 = nn.Linear(hidden_dim, 1)

        self.fc4 = nn.Linear(state_dim + action_dim, hidden_dim)
        self.fc5 = nn.Linear(hidden_dim, hidden_dim)
        self.fc6 = nn.Linear(hidden_dim, 1)

    def forward(self, state, action):
        xu = torch.cat([state, action], 1)

        x1 = F.relu(self.fc1(xu))
        x1 = F.relu(self.fc2(x1))
        q1 = self.fc3(x1)

        x2 = F.relu(self.fc4(xu))
        x2 = F.relu(self.fc5(x2))
        q2 = self.fc6(x2)

        return q1, q2


# ===  ===
class EnhancedReplayBuffer:
    def __init__(self, capacity=50000):
        self.buffer = deque(maxlen=capacity)
        self.experience_info = deque(maxlen=capacity)

    def push(self, state, action, reward, next_state, done, path_idx, similarity):
        self.buffer.append((state, action, reward, next_state, done))
        self.experience_info.append({'path_idx': path_idx, 'similarity': similarity})

    def sample(self, batch_size):
        if len(self.buffer) < batch_size:
            return None

        batch = random.sample(self.buffer, batch_size)
        state, action, reward, next_state, done = map(np.stack, zip(*batch))

        return (
            torch.FloatTensor(state).to(device),
            torch.FloatTensor(action).to(device),
            torch.FloatTensor(reward).unsqueeze(1).to(device),
            torch.FloatTensor(next_state).to(device),
            torch.FloatTensor(done).unsqueeze(1).to(device)
        )

    def get_top_k_per_path(self, num_paths, k=20):
        path_samples = {i: [] for i in range(num_paths)}

        for idx, info in enumerate(self.experience_info):
            path_idx = info['path_idx']
            similarity = info['similarity']
            path_samples[path_idx].append((idx, similarity, self.buffer[idx]))

        top_k_results = {}
        for path_idx in range(num_paths):
            samples = path_samples[path_idx]
            if len(samples) == 0:
                top_k_results[path_idx] = []
                continue

            samples.sort(key=lambda x: x[1], reverse=True)
            top_k = samples[:k]

            top_k_results[path_idx] = []
            for sample in top_k:
                normalized_state = sample[2][0]
                original_state = denormalize_state(normalized_state)
                original_state_int = np.round(original_state).astype(int)

                # : 
                co2, moisture, temp = original_state_int
                triggered = section6_low_co2_extremes(moisture, co2, temp)

                top_k_results[path_idx].append({
                    'state': original_state_int,
                    'similarity': sample[1],
                    'triggered': triggered
                })

        return top_k_results

    def __len__(self):
        return len(self.buffer)


# === SAC ===
class SACAgent:
    def __init__(self, state_dim=3, action_dim=3):
        self.state_dim = state_dim
        self.action_dim = action_dim

        # : 
        self.policy = GaussianPolicy(state_dim, action_dim, EXPERIMENT_CONFIG['HIDDEN_DIM']).to(device)
        self.policy_optimizer = optim.Adam(self.policy.parameters(), lr=EXPERIMENT_CONFIG['ACTOR_LR'])

        self.critic = QNetwork(state_dim, action_dim, EXPERIMENT_CONFIG['HIDDEN_DIM']).to(device)
        self.critic_target = QNetwork(state_dim, action_dim, EXPERIMENT_CONFIG['HIDDEN_DIM']).to(device)
        self.critic_target.load_state_dict(self.critic.state_dict())
        self.critic_optimizer = optim.Adam(self.critic.parameters(), lr=EXPERIMENT_CONFIG['CRITIC_LR'])

        self.target_entropy = -action_dim
        self.log_alpha = torch.zeros(1, requires_grad=True, device=device)
        self.alpha_optimizer = optim.Adam([self.log_alpha], lr=EXPERIMENT_CONFIG['ALPHA_LR'])

        self.replay_buffer = EnhancedReplayBuffer()
        self.replay_train_count = 0

    def get_action(self, state, deterministic=False):
        normalized_state = normalize_state(state)
        state_tensor = torch.FloatTensor(normalized_state).unsqueeze(0).to(device)

        with torch.no_grad():
            if deterministic:
                _, _, action = self.policy.sample(state_tensor)
            else:
                action, _, _ = self.policy.sample(state_tensor)

        action = action.cpu().numpy()[0]
        return action

    def store_experience(self, state, action, reward, next_state, done, path_idx, similarity):
        normalized_state = normalize_state(state)
        normalized_next_state = normalize_state(next_state)

        self.replay_buffer.push(
            normalized_state, action, reward,
            normalized_next_state, done, path_idx, similarity
        )

    def replay_train(self):
        batch = self.replay_buffer.sample(EXPERIMENT_CONFIG['REPLAY_BATCH_SIZE'])

        if batch is None:
            return

        state, action, reward, next_state, done = batch

        with torch.no_grad():
            next_action, next_log_prob, _ = self.policy.sample(next_state)

            q1_next, q2_next = self.critic_target(next_state, next_action)
            q_next = torch.min(q1_next, q2_next)
            target_q = reward + (1 - done) * EXPERIMENT_CONFIG['GAMMA'] * (
                    q_next - self.log_alpha.exp() * next_log_prob)

        q1, q2 = self.critic(state, action)
        critic_loss = F.mse_loss(q1, target_q) + F.mse_loss(q2, target_q)

        self.critic_optimizer.zero_grad()
        critic_loss.backward()
        self.critic_optimizer.step()

        new_action, log_prob, _ = self.policy.sample(state)
        q1_new, q2_new = self.critic(state, new_action)
        q_new = torch.min(q1_new, q2_new)

        policy_loss = (self.log_alpha.exp() * log_prob - q_new).mean()

        self.policy_optimizer.zero_grad()
        policy_loss.backward()
        self.policy_optimizer.step()

        alpha_loss = -(self.log_alpha * (log_prob + self.target_entropy).detach()).mean()

        self.alpha_optimizer.zero_grad()
        alpha_loss.backward()
        self.alpha_optimizer.step()

        for param, target_param in zip(self.critic.parameters(), self.critic_target.parameters()):
            target_param.data.copy_(
                EXPERIMENT_CONFIG['TAU'] * param.data + (1 - EXPERIMENT_CONFIG['TAU']) * target_param.data)

        self.replay_train_count += 1

        if self.replay_train_count % 2 == 0:
            alpha_value = self.log_alpha.exp().item()
            print(f"  ->  (Run {self.replay_train_count}), Alpha={alpha_value:.4f}")


# === Metric ===
def calculate_run_performance(run_idx, sac_results, training_time, total_steps, update_count, agent):
    """ runMetric"""
    target_paths = EXPERIMENT_CONFIG['TARGET_PATHS']
    num_paths = len(target_paths)

    # Similarity
    all_similarities = []

    # Metric
    total_samples = 0
    all_rewards = []
    total_reward = 0

    for path_idx in range(num_paths):
        samples = sac_results[path_idx]
        for sample in samples:
            triggered = sample['triggered']
            target_path = target_paths[path_idx]
            reward = unified_reward_function(triggered, target_path)
            similarity = sample['similarity']

            total_reward += reward
            all_rewards.append(reward)
            all_similarities.append(similarity)
            total_samples += 1

    # 1. 
    total_reward = total_reward

    # 2. 
    if total_samples > 0:
        average_reward = total_reward / total_samples
    else:
        average_reward = 0

    # 5. (Average Similarity)
    if all_similarities:
        convergence = np.mean(all_similarities)
    else:
        convergence = 0

    # 12. (Similarity)
    if len(all_similarities) > 1:
        environment_adaptability = 1 / (np.std(all_similarities) + 1e-8)
    else:
        environment_adaptability = 0

    # 13. (Average Similarity)
    generalization_ability = convergence

    # 15. (/ seconds)
    if training_time > 0:
        computational_efficiency = total_steps / training_time
    else:
        computational_efficiency = 0

    # 16. 
    if training_time > 0:
        policy_update_frequency = update_count / training_time
    else:
        policy_update_frequency = 0

    # Similarity
    avg_similarity = np.mean(all_similarities) if all_similarities else 0
    max_similarity = np.max(all_similarities) if all_similarities else 0
    min_similarity = np.min(all_similarities) if all_similarities else 0

    return {
        '': run_idx + 1,

        # Metric
        '': round(total_reward, 2),
        '': round(average_reward, 4),
        '': round(convergence, 4),
        '': round(environment_adaptability, 4),
        '': round(generalization_ability, 4),
        '': round(computational_efficiency, 2),
        '': round(policy_update_frequency, 4),

        # Similarity
        'Average Similarity': round(avg_similarity, 4),
        'Similarity': round(max_similarity, 4),
        'Similarity': round(min_similarity, 4),
    }


# === Excel ===
def export_to_excel(all_sac_results, all_performance_data, target_paths, output_path="SAC_20 run.xlsx"):
    """20 runSACExcel"""
    print("\nExcel...")

    # 
    all_sac_summary_data = []
    all_sac_detailed_data = []

    #  run
    for run_idx, (sac_results, performance_data) in enumerate(zip(all_sac_results, all_performance_data)):
        # ===== Sheet1: SACPath  =====
        sac_summary_data = []
        for path_idx in range(len(target_paths)):
            target_path = target_paths[path_idx]
            samples = sac_results[path_idx]

            if len(samples) == 0:
                sac_summary_data.append({
                    '': run_idx + 1,
                    'Path ID': path_idx + 1,
                    '': len(target_path),
                    '': 0,
                    'Average Similarity': 0,
                    'Similarity': 0,
                    'Similarity': 0,
                    'SimilarityStandard deviation': 0,
                    '': '',
                    'target paths': ', '.join(map(str, sorted(target_path)))
                })
                continue

            similarities = [s['similarity'] for s in samples]
            perfect_count = sum(1 for s in similarities if abs(s - 1.0) < 0.001)
            is_perfect = '' if perfect_count > 0 else ''

            sac_summary_data.append({
                '': run_idx + 1,
                'Path ID': path_idx + 1,
                '': len(target_path),
                '': len(samples),
                'Average Similarity': round(np.mean(similarities), 4),
                'Similarity': round(max(similarities), 4),
                'Similarity': round(min(similarities), 4),
                'SimilarityStandard deviation': round(np.std(similarities), 4),
                '': is_perfect,
                'target paths': ', '.join(map(str, sorted(target_path)))
            })

        all_sac_summary_data.extend(sac_summary_data)

        # ===== Sheet2: SACDetailed Sample Data =====
        sac_detailed_data = []
        for path_idx in range(len(target_paths)):
            target_path = target_paths[path_idx]
            samples = sac_results[path_idx]

            for sample_idx, sample in enumerate(samples):
                state = sample['state']
                similarity = sample['similarity']
                triggered = sample['triggered']

                sac_detailed_data.append({
                    '': run_idx + 1,
                    'Path ID': path_idx + 1,
                    'Sample ID': sample_idx + 1,
                    'CO2': int(state[0]),
                    'moisture': int(state[1]),
                    'temperature': int(state[2]),
                    'Similarity': round(similarity, 4),
                    '': '' if abs(similarity - 1.0) < 0.001 else '',
                    'target paths': ', '.join(map(str, sorted(target_path))),
                    '': ', '.join(map(str, sorted(triggered))),
                    '': len(target_path.intersection(triggered)),
                    '': len(target_path)
                })

        all_sac_detailed_data.extend(sac_detailed_data)

    # Excel
    sac_summary_df = pd.DataFrame(all_sac_summary_data)
    sac_detailed_df = pd.DataFrame(all_sac_detailed_data)
    performance_df = pd.DataFrame(all_performance_data)

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Sheet1: SACPath 
        sac_summary_df.to_excel(writer, sheet_name='SACPath ', index=False)

        # Sheet2: SACDetailed Sample Data
        sac_detailed_df.to_excel(writer, sheet_name='SACDetailed Sample Data', index=False)

        # Sheet3: Metric - 
        selected_columns = [
            '',
            '', '', '', '',
            '', '', '',
            'Average Similarity', 'Similarity', 'Similarity'
        ]
        performance_df_selected = performance_df[selected_columns]
        performance_df_selected.to_excel(writer, sheet_name='Metric', index=False)

        # 
        workbook = writer.book

        # 
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(name='Microsoft YaHei', size=11, bold=True, color='FFFFFF')
        perfect_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # 

        # === Sheet1 ===
        ws1 = writer.sheets['SACPath ']
        for cell in ws1[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 
        for row_idx in range(2, ws1.max_row + 1):
            if ws1.cell(row_idx, 9).value == '':  # Run 9""
                for col_idx in range(1, ws1.max_column + 1):
                    ws1.cell(row_idx, col_idx).fill = perfect_fill

        ws1.column_dimensions['A'].width = 12
        ws1.column_dimensions['B'].width = 12
        ws1.column_dimensions['C'].width = 12
        ws1.column_dimensions['D'].width = 12
        ws1.column_dimensions['E'].width = 15
        ws1.column_dimensions['F'].width = 15
        ws1.column_dimensions['G'].width = 15
        ws1.column_dimensions['H'].width = 15
        ws1.column_dimensions['I'].width = 15
        ws1.column_dimensions['J'].width = 50

        # === Sheet2 ===
        ws2 = writer.sheets['SACDetailed Sample Data']
        for cell in ws2[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        ws2.column_dimensions['A'].width = 12
        ws2.column_dimensions['B'].width = 12
        ws2.column_dimensions['C'].width = 12
        ws2.column_dimensions['D'].width = 10
        ws2.column_dimensions['E'].width = 10
        ws2.column_dimensions['F'].width = 10
        ws2.column_dimensions['G'].width = 12
        ws2.column_dimensions['H'].width = 15
        ws2.column_dimensions['I'].width = 40
        ws2.column_dimensions['J'].width = 40
        ws2.column_dimensions['K'].width = 15
        ws2.column_dimensions['L'].width = 15

        # === Sheet3 ===
        ws3 = writer.sheets['Metric']
        for cell in ws3[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 
        columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']
        for col in columns:
            ws3.column_dimensions[col].width = 18

    print(f"Excel: {output_path}")
    print(f"  - Sheet1: SACPath  ({len(all_sac_summary_data)})")
    print(f"  - Sheet2: SACDetailed Sample Data ({len(all_sac_detailed_data)})")
    print(f"  - Sheet3: Metric ({len(all_performance_data)})")


# ===  ===
def train_sac_workflow():
    print("=" * 80)
    print("SAC")
    print("Similarity:  / target paths")
    print("=" * 80)

    agent = SACAgent()
    target_paths = EXPERIMENT_CONFIG['TARGET_PATHS']
    num_paths = len(target_paths)

    start_time = time.time()
    total_steps = 0

    print(f"\n: Path {EXPERIMENT_CONFIG['SAMPLES_PER_PATH']}")
    path_samples = {}
    min_vals = EXPERIMENT_CONFIG['MIN_VALUES']
    max_vals = EXPERIMENT_CONFIG['MAX_VALUES']

    for path_idx in range(num_paths):
        samples = []
        for _ in range(EXPERIMENT_CONFIG['SAMPLES_PER_PATH']):
            # 
            state = np.random.uniform(min_vals, max_vals).astype(np.float32)
            samples.append(state)
        path_samples[path_idx] = samples
        print(f"  Path  {path_idx + 1}/{num_paths}:  {len(samples)} ")

    batch_size = EXPERIMENT_CONFIG['BATCH_SIZE_SAMPLES']
    num_batches = EXPERIMENT_CONFIG['SAMPLES_PER_PATH'] // batch_size

    print(f"\n: {batch_size},{EXPERIMENT_CONFIG['STEPS_PER_SAMPLE']}")
    print(f": {num_batches} /Path  x {num_paths} Path  = {num_batches * num_paths} ")
    print("-" * 80)

    for batch_idx in range(num_batches):
        print(f"\n {batch_idx + 1}/{num_batches}")

        for path_idx in range(num_paths):
            target_path = target_paths[path_idx]
            batch_samples = path_samples[path_idx][batch_idx * batch_size:(batch_idx + 1) * batch_size]

            batch_rewards = []
            batch_similarities = []

            for sample_idx, initial_state in enumerate(batch_samples):
                state = initial_state.copy()
                episode_reward = 0
                final_similarity = 0

                for step in range(EXPERIMENT_CONFIG['STEPS_PER_SAMPLE']):
                    action = agent.get_action(state, deterministic=False)

                    next_state = state + action
                    next_state = clip_state(next_state)

                    # : 
                    co2, moisture, temp = next_state
                    triggered = section6_low_co2_extremes(moisture, co2, temp)
                    reward = unified_reward_function(triggered, target_path)
                    similarity = coverage_similarity(triggered, target_path)

                    done = (step == EXPERIMENT_CONFIG['STEPS_PER_SAMPLE'] - 1)

                    agent.store_experience(
                        state, action, reward, next_state, done,
                        path_idx, similarity
                    )

                    state = next_state
                    episode_reward += reward
                    final_similarity = similarity
                    total_steps += 1

                batch_rewards.append(episode_reward)
                batch_similarities.append(final_similarity)

            avg_reward = np.mean(batch_rewards)
            avg_similarity = np.mean(batch_similarities)
            print(f"  Path {path_idx + 1}: ={avg_reward:.2f}, Average Similarity={avg_similarity:.4f}")

        print(f"\n  ...")
        agent.replay_train()
        print(f"  : {len(agent.replay_buffer)}")

    training_time = time.time() - start_time

    print("\n" + "=" * 80)
    print(f"SACcompleted! Total elapsed time: {training_time:.2f} seconds, : {total_steps}")
    print(f": {len(agent.replay_buffer)}")
    print(f": {agent.replay_train_count}")
    print("=" * 80)

    print(f"\nPath SimilarityMaximum{EXPERIMENT_CONFIG['TOP_K_SAMPLES']}...")
    top_k_results = agent.replay_buffer.get_top_k_per_path(num_paths, EXPERIMENT_CONFIG['TOP_K_SAMPLES'])

    return agent, top_k_results, training_time, total_steps, agent.replay_train_count


# ===  ===
def main():
    print("\n" + "=" * 80)
    print("SAC - 20 run")
    print(": CO2(800-1500), moisture(10-80), temperature(1-40)")
    print("Metric")
    print("=" * 80)

    all_sac_results = []
    all_performance_data = []
    target_paths = EXPERIMENT_CONFIG['TARGET_PATHS']

    # 20
    for run_idx in range(EXPERIMENT_CONFIG['NUM_RUNS']):
        print(f"\n{'=' * 80}")
        print(f"Start run  {run_idx + 1}/{EXPERIMENT_CONFIG['NUM_RUNS']}  run")
        print(f"{'=' * 80}")

        # SAC
        sac_agent, sac_results, training_time, total_steps, update_count = train_sac_workflow()

        # Metric
        performance_data = calculate_run_performance(
            run_idx, sac_results, training_time, total_steps, update_count, sac_agent
        )

        # 
        all_sac_results.append(sac_results)
        all_performance_data.append(performance_data)

        print(f"\nRun  {run_idx + 1}  runcompleted!")
        print(f"  : {performance_data['']}")
        print(f"  : {performance_data['']}")
        print(f"  : {performance_data['']}")

    # Excel(20 run)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = f"SAC_20 run_{timestamp}.xlsx"
    export_to_excel(all_sac_results, all_performance_data, target_paths, output_path)

    # 
    print("\n" + "=" * 80)
    print("20 run")
    print("=" * 80)

    # Metric
    total_rewards = [p[''] for p in all_performance_data]
    average_rewards = [p[''] for p in all_performance_data]
    convergences = [p[''] for p in all_performance_data]
    environment_adaptabilities = [p[''] for p in all_performance_data]
    generalization_abilities = [p[''] for p in all_performance_data]
    computational_efficiencies = [p[''] for p in all_performance_data]
    policy_update_frequencies = [p[''] for p in all_performance_data]
    avg_similarities = [p['Average Similarity'] for p in all_performance_data]

    print(f":")
    print(f"  : {np.mean(total_rewards):.2f}")
    print(f"  Standard deviation: {np.std(total_rewards):.2f}")

    print(f"\n:")
    print(f"  : {np.mean(average_rewards):.4f}")
    print(f"  Standard deviation: {np.std(average_rewards):.4f}")

    print(f"\n:")
    print(f"  : {np.mean(convergences):.4f}")
    print(f"  Standard deviation: {np.std(convergences):.4f}")

    print(f"\n:")
    print(f"  : {np.mean(environment_adaptabilities):.4f}")
    print(f"  Standard deviation: {np.std(environment_adaptabilities):.4f}")

    print(f"\n:")
    print(f"  : {np.mean(generalization_abilities):.4f}")
    print(f"  Standard deviation: {np.std(generalization_abilities):.4f}")

    print(f"\n:")
    print(f"  : {np.mean(computational_efficiencies):.2f}")
    print(f"  Standard deviation: {np.std(computational_efficiencies):.2f}")

    print(f"\n:")
    print(f"  : {np.mean(policy_update_frequencies):.4f}")
    print(f"  Standard deviation: {np.std(policy_update_frequencies):.4f}")

    print(f"\nAverage similarity statistics:")
    print(f"  : {np.mean(avg_similarities):.4f}")
    print(f"  Standard deviation: {np.std(avg_similarities):.4f}")

    print("\n" + "=" * 80)
    print(f" {EXPERIMENT_CONFIG['NUM_RUNS']} completed!")
    print("=" * 80)


if __name__ == "__main__":
    main()