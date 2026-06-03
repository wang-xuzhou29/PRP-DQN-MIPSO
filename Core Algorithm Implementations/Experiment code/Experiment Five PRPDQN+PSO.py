import torch.nn as nn
import os
import torch.optim as optim
import random
from collections import deque
import numpy as np
import torch
from datetime import datetime
import time
import psutil
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sys

NUM_RUNS = 20  # 实验运行次数，可以在这里修改

# === 设备设置 ===
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"

# === 新的三维范围设置 ===
LIGHT_MIN = 1
LIGHT_MAX = 50
MOISTURE_MIN = 1
MOISTURE_MAX = 50
TEMP_MIN = 1
TEMP_MAX = 50

BOUNDS = {
    'light': (LIGHT_MIN, LIGHT_MAX),  # x: 光照强度
    'temp': (TEMP_MIN, TEMP_MAX),    # y: 温度
    'moisture': (MOISTURE_MIN, MOISTURE_MAX)  # z: 湿度
}

# === 动作增量设置（按照最大值的正负5%10%20%50%70%的近似整数）===
# light (1-50, 范围49): 5%≈2, 10%≈5, 20%≈10, 50%≈25, 70%≈35
DELTA_LIGHT = [35, 25, 10, 5, 2, -2, -5, -10, -25, -35]
# temp (1-50, 范围49): 5%≈2, 10%≈5, 20%≈10, 50%≈25, 70%≈35
DELTA_TEMP = [35, 25, 10, 5, 2, -2, -5, -10, -25, -35]
# moisture (1-50, 范围49): 5%≈2, 10%≈5, 20%≈10, 50%≈25, 70%≈35
DELTA_MOISTURE = [35, 25, 10, 5, 2, -2, -5, -10, -25, -35]


# ========================================
# ========== 状态归一化函数 ==========
# ========================================
def normalize_state(state):
    """
    将状态归一化到[0, 1]区间
    state: (light, temp, moisture)
    """
    normalized = np.array([
        (state[0] - BOUNDS['light'][0]) / (BOUNDS['light'][1] - BOUNDS['light'][0]),
        (state[1] - BOUNDS['temp'][0]) / (BOUNDS['temp'][1] - BOUNDS['temp'][0]),
        (state[2] - BOUNDS['moisture'][0]) / (BOUNDS['moisture'][1] - BOUNDS['moisture'][0])
    ], dtype=np.float32)
    return normalized


def denormalize_state(normalized_state):
    """
    将归一化的状态还原到原始范围
    normalized_state: (norm_light, norm_temp, norm_moisture) 在[0,1]区间
    """
    state = np.array([
        normalized_state[0] * (BOUNDS['light'][1] - BOUNDS['light'][0]) + BOUNDS['light'][0],
        normalized_state[1] * (BOUNDS['temp'][1] - BOUNDS['temp'][0]) + BOUNDS['temp'][0],
        normalized_state[2] * (BOUNDS['moisture'][1] - BOUNDS['moisture'][0]) + BOUNDS['moisture'][0]
    ])
    return state


# ========================================


def generate_random_state():
    """生成符合各维度范围的随机状态"""
    light = np.random.randint(BOUNDS['light'][0], BOUNDS['light'][1] + 1)
    temp = np.random.randint(BOUNDS['temp'][0], BOUNDS['temp'][1] + 1)
    moisture = np.random.randint(BOUNDS['moisture'][0], BOUNDS['moisture'][1] + 1)
    return np.array([light, temp, moisture])


def clip_state(state):
    """将状态限制在各维度的边界内"""
    return np.array([
        np.clip(state[0], BOUNDS['light'][0], BOUNDS['light'][1]),
        np.clip(state[1], BOUNDS['temp'][0], BOUNDS['temp'][1]),
        np.clip(state[2], BOUNDS['moisture'][0], BOUNDS['moisture'][1])
    ])


def is_state_valid(state):
    """检查状态是否在所有维度的有效范围内"""
    return (BOUNDS['light'][0] <= state[0] <= BOUNDS['light'][1] and
            BOUNDS['temp'][0] <= state[1] <= BOUNDS['temp'][1] and
            BOUNDS['moisture'][0] <= state[2] <= BOUNDS['moisture'][1])


def execute_Tr(position):
    """执行目标函数并返回触发的路径"""
    x = int(np.clip(position[0], BOUNDS['light'][0], BOUNDS['light'][1]))
    temp = int(np.clip(position[1], BOUNDS['temp'][0], BOUNDS['temp'][1]))
    z = int(np.clip(position[2], BOUNDS['moisture'][0], BOUNDS['moisture'][1]))
    return category1_multivariable_control(x, temp, z)


# === 增强版指标收集器 ===
class MetricsCollector:
    def __init__(self):
        self.start_time = None
        self.end_time = None
        self.total_reward = 0
        self.td_errors = []
        self.final_output_similarities = []
        self.action_improvements = []
        self.total_memory_usage = 0
        self.memory_check_count = 0
        self.step_count = 0

        # === PSO阶段统计 ===
        self.pso_start_time = None
        self.pso_end_time = None
        self.perfect_solutions_count = 0
        self.total_paths_count = 0
        self.final_fitness_scores = []
        self.pso_convergence_iterations = []
        self.pso_perfect_solutions = []
        self.pso_reset_counts = []
        self.path_execution_times = []

    def start_training(self):
        self.start_time = time.time()

    def end_training(self):
        self.end_time = time.time()

    def start_pso_phase(self):
        self.pso_start_time = time.time()

    def end_pso_phase(self):
        self.pso_end_time = time.time()

    def record_pso_result(self, fitness, is_perfect_match, convergence_iter=None, path_id=None, method='PSO',
                          reset_count=0, execution_time=0):
        self.final_fitness_scores.append(fitness)
        self.total_paths_count += 1
        self.pso_reset_counts.append(reset_count)
        self.path_execution_times.append(execution_time)

        if is_perfect_match:
            self.perfect_solutions_count += 1
            if method == 'PSO' and convergence_iter is not None:
                self.pso_perfect_solutions.append({
                    'path_id': path_id,
                    'convergence_iteration': convergence_iter,
                    'fitness': fitness,
                    'reset_count': reset_count
                })
        if convergence_iter is not None:
            self.pso_convergence_iterations.append(convergence_iter)

    def record_step_metrics(self, reward, td_error, triggered, target_path):
        self.step_count += 1
        self.total_reward += reward
        self.td_errors.append(td_error)

        try:
            process = psutil.Process(os.getpid())
            current_memory = process.memory_info().rss / 1024 / 1024
            self.total_memory_usage += current_memory
            self.memory_check_count += 1
        except:
            pass

    def record_final_output_sample(self, triggered, target_path):
        similarity = jaccard_similarity(triggered, target_path)
        self.final_output_similarities.append(similarity)

    def record_action_improvement(self, current_reward, prev_reward):
        if prev_reward is not None:
            improvement = current_reward - prev_reward
            self.action_improvements.append(1 if improvement > 0 else 0)


# === 奖励函数 ===
def compute_reward(state, target_path, triggered, prev_triggered=None, prev_state=None):
    sim = jaccard_similarity(triggered, target_path)
    reward = sim * 10

    if target_path.issubset(triggered):
        reward += 1

    return reward


def category1_multivariable_control(dx, dy, dz):
    """执行目标函数并返回触发的路径"""
    # --- 1. 常量与配置 ---
    MAX_GRID_SIZE = 500.0  # 模拟网格的最大边界，扩大到 500.0
    INITIAL_BATTERY = 1000.0  # 初始电量相应扩大，以支撑更长的路径
    BATTERY_PER_STEP = 1.0  # 每走一步消耗的电量降低，避免一步耗尽
    SAFE_DISTANCE = 5.0  # 安全距离阈值 (固定)
    CRITICAL_BATTERY_LEVEL = 100.0  # 危险电量阈值 (固定)
    TARGET_X, TARGET_Y, TARGET_Z = 450.0, 450.0, 200.0  # 目标坐标相应扩大 (固定值)

    MIN_PLANNING_X = 10.0
    MIN_PLANNING_Y = 15.0
    MIN_PLANNING_Z = 8.0
    CRITICAL_X_VELOCITY = 20.0
    CRITICAL_Y_VELOCITY = 25.0
    CRITICAL_Z_VELOCITY = 15.0

    triggered = set()

    # 模拟环境状态变量，以修复原始代码中的语法错误
    # 这些变量在原始代码中未定义，这里随机生成以确保代码可执行
    current_x = random.uniform(0.0, MAX_GRID_SIZE)
    current_y = random.uniform(0.0, MAX_GRID_SIZE)
    current_z = random.uniform(0.0, MAX_GRID_SIZE)

    # 为了模仿原始代码可能试图比较的'当前位置'与'速度'的关系，
    # 针对第10-15分支中的 'self.y' 采用一个模拟值。
    simulated_y = current_y  # 使用 current_y 作为 self.y 的模拟

    # --- 分支 1-4 ---
    if abs(dx) < MIN_PLANNING_X != abs(dy) < MIN_PLANNING_X:
        triggered.add(1)
    if abs(dx) < MIN_PLANNING_X != abs(dz) < MIN_PLANNING_X:
        triggered.add(2)
    if abs(dx) < MIN_PLANNING_X != abs(dx) < MIN_PLANNING_Y:
        triggered.add(3)
    if abs(dx) < MIN_PLANNING_X != abs(dx) < MIN_PLANNING_Z:
        triggered.add(4)

    # --- 分支 5-9 ---
    if abs(dz) > MIN_PLANNING_Z * 2 != abs(dx) > MIN_PLANNING_Z * 2:
        triggered.add(5)
    if abs(dz) > MIN_PLANNING_Z * 2 != abs(dy) > MIN_PLANNING_Z * 2:
        triggered.add(6)
    if abs(dz) > MIN_PLANNING_Z * 2 != abs(dz) > MIN_PLANNING_X * 2:
        triggered.add(7)
    if abs(dz) > MIN_PLANNING_Z * 2 != abs(dz) > MIN_PLANNING_Y * 2:
        triggered.add(8)
    if abs(dz) > MIN_PLANNING_Z * 2 != abs(dz) > MIN_PLANNING_Z:
        triggered.add(9)

    # --- 分支 10-15 --- (使用 simulated_y 替代 self.y)
    if TARGET_Y > simulated_y and dy < 20 != TARGET_Y > simulated_y and dy < 10:
        triggered.add(10)
    if TARGET_Y > simulated_y and dy < 20 != TARGET_Y > simulated_y and dy < 30:
        triggered.add(11)
    if TARGET_Y > simulated_y and dy < 20 != TARGET_Y > simulated_y and dy < 40:
        triggered.add(12)
    if TARGET_Y > simulated_y and dy < 20 != TARGET_Y > simulated_y and dy < 50:
        triggered.add(13)
    if TARGET_Y > simulated_y and dy < 20 != TARGET_Y > simulated_y and dx < 20:
        triggered.add(14)
    if TARGET_Y > simulated_y and dy < 20 != TARGET_Y > simulated_y and dz < 20:
        triggered.add(15)

    # --- 分支 16-21 ---
    if abs(dy) > CRITICAL_X_VELOCITY * 1.5 != abs(dx) > CRITICAL_X_VELOCITY * 1.5:
        triggered.add(16)
    if abs(dy) > CRITICAL_X_VELOCITY * 1.5 != abs(dz) > CRITICAL_X_VELOCITY * 1.5:
        triggered.add(17)
    if abs(dy) > CRITICAL_X_VELOCITY * 1.5 != abs(dy) > CRITICAL_X_VELOCITY:
        triggered.add(18)
    if abs(dy) > CRITICAL_X_VELOCITY * 1.5 != abs(dy) > CRITICAL_X_VELOCITY * 2:
        triggered.add(19)
    if abs(dy) > CRITICAL_X_VELOCITY * 1.5 != abs(dy) > CRITICAL_Z_VELOCITY * 1.5:
        triggered.add(20)
    if abs(dy) > CRITICAL_X_VELOCITY * 1.5 != abs(dy) > CRITICAL_Y_VELOCITY * 1.5:
        triggered.add(21)

    # --- 分支 22-29 --- (使用 current_x, current_y, current_z 替代未定义的变量)
    if TARGET_Z < current_z and dz > CRITICAL_Z_VELOCITY != TARGET_X < current_z and dz > CRITICAL_Z_VELOCITY:
        triggered.add(22)
    if TARGET_Z < current_z and dz > CRITICAL_Z_VELOCITY != TARGET_Y < current_z and dz > CRITICAL_Z_VELOCITY:
        triggered.add(23)
    if TARGET_Z < current_z and dz > CRITICAL_Z_VELOCITY != TARGET_Z < current_x and dz > CRITICAL_Z_VELOCITY:
        triggered.add(24)
    if TARGET_Z < current_z and dz > CRITICAL_Z_VELOCITY != TARGET_Z < current_y and dz > CRITICAL_Z_VELOCITY:
        triggered.add(25)
    if TARGET_Z < current_z and dz > CRITICAL_Z_VELOCITY != TARGET_Z < current_z and dx > CRITICAL_Z_VELOCITY:
        triggered.add(26)
    if TARGET_Z < current_z and dz > CRITICAL_Z_VELOCITY != TARGET_Z < current_z and dy > CRITICAL_Z_VELOCITY:
        triggered.add(27)
    if TARGET_Z < current_z and dz > CRITICAL_Z_VELOCITY != TARGET_Z < current_z and dz > CRITICAL_X_VELOCITY:
        triggered.add(28)
    if TARGET_Z < current_z and dz > CRITICAL_Z_VELOCITY != TARGET_Z < current_z and dz > CRITICAL_Y_VELOCITY:
        triggered.add(29)

    return triggered


# 目标路径定义
targetPaths = [
    {1, 2, 3, 4, 10, 11, 12, 13, 14, 15, 24, 25, 26, 27, 28, 29},
    {5, 6, 7, 8, 9, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25},
    {5, 6, 7, 8, 9, 17, 18, 19, 20, 21, 24, 25, 26, 27, 28, 29}
]


def jaccard_similarity(set1, set2):
    """
    计算Jaccard相似度
    **如果set1包含set2（即set2是set1的子集），返回1.0**
    """
    # 如果set1覆盖了set2（目标路径），返回相似度1
    if set2.issubset(set1):
        return 1.0

    # 否则计算标准Jaccard相似度
    intersection = len(set1 & set2)
    union = len(set1 | set2)
    return intersection / union if union != 0 else 0.0


def calculate_fitness(position, target_path):
    """计算适应度（Jaccard相似度）"""
    triggered = execute_Tr(position)

    # 如果生成路径完全包含目标路径，返回最高适应度
    if target_path.issubset(triggered):
        return 1.0

    # 否则计算Jaccard相似度
    intersection = len(triggered & target_path)
    union = len(triggered | target_path)
    return intersection / union if union > 0 else 0.0


# === 路径相似度矩阵计算 ===
def compute_path_similarity_matrix(paths):
    n = len(paths)
    matrix = np.zeros((n, n))
    for i in range(n):
        for j in range(n):
            inter = len(paths[i] & paths[j])
            union = len(paths[i] | paths[j])
            matrix[i][j] = inter / union if union > 0 else 0.0
    return matrix


# === 路径分组 ===
def group_paths_by_similarity(paths):
    sim_matrix = compute_path_similarity_matrix(paths)
    avg_sim_scores = np.mean(sim_matrix, axis=1)
    threshold = np.mean(avg_sim_scores)

    center_idx = np.argmax(avg_sim_scores)
    similar_group = [center_idx]
    for i in range(len(paths)):
        if i != center_idx and sim_matrix[center_idx][i] > threshold:
            similar_group.append(i)

    isolated_group = [i for i in range(len(paths)) if i not in similar_group]
    return similar_group, isolated_group


# === 样本生成===
def generate_samples_for_similar_paths(similar_group_indices, num_total=2000, top_k=200):
    def jaccard_similarity_local(a, b):
        if not a and not b:
            return 1.0
        return len(a & b) / len(a | b) if a | b else 0.0

    def compute_robustness(state, path):
        base = execute_Tr(state)
        if not base:
            return 0.0
        rob, neighbors = 0.0, 0
        for dx in [-1, 0, 1]:
            for dy in [-1, 0, 1]:
                for dz in [-1, 0, 1]:
                    if dx == dy == dz == 0:
                        continue
                    neighbor = clip_state(state + np.array([dx, dy, dz]))
                    n_trig = execute_Tr(neighbor)
                    if not n_trig:
                        continue
                    rob += jaccard_similarity_local(base, n_trig)
                    neighbors += 1
        return rob / neighbors if neighbors > 0 else 0.0

    def save_samples(path_id, samples, base_dir):
        os.makedirs(base_dir, exist_ok=True)
        filepath = os.path.join(base_dir, f"path{path_id}.txt")
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(f"Path {path_id}\n")
            f.write("light temp moisture\tScore\tSimilarity\tLengthDiff\tRobustness\n")
            for s in samples:
                light, temp, moisture = int(s[0][0]), int(s[0][1]), int(s[0][2])
                f.write(f"{light} {temp} {moisture}\t{s[1]:.4f}\t{s[2]:.4f}\t{s[3]:.4f}\t{s[4]:.4f}\n")

    print("正在为相似路径组生成样本数据...")
    base_dir = os.path.join(os.getcwd(), "path_samples")
    for path_idx in similar_group_indices:
        path = targetPaths[path_idx]
        samples = []
        attempts = 0
        while len(samples) < top_k and attempts < num_total * 5:
            attempts += 1
            state = generate_random_state()
            triggered = execute_Tr(state)
            if not triggered:
                continue
            sim = jaccard_similarity_local(triggered, path)
            len_diff = 1 - abs(len(triggered) - len(path)) / max(len(triggered), len(path))
            rob = compute_robustness(state, path)
            score = 0.55 * sim + 0.25 * len_diff + 0.2 * rob
            samples.append((state, score, sim, len_diff, rob))
        if samples:
            samples.sort(key=lambda x: x[1], reverse=True)
            save_samples(path_id=path_idx + 1, samples=samples[:top_k], base_dir=base_dir)


# === 经验回放池===
class SharedExperienceReplay:
    def __init__(self, capacity=10000):
        self.capacity = capacity
        self.buffer = deque(maxlen=self.capacity)
        self.priorities = deque(maxlen=self.capacity)

    def append(self, experience):
        self.buffer.append(experience)
        self.priorities.append(experience[-1])

    def sample(self, batch_size, alpha=0.6):
        if len(self.buffer) == 0:
            return [], [], []

        priorities = np.array(self.priorities) ** alpha
        sum_priorities = np.sum(priorities)
        if sum_priorities == 0:
            probabilities = np.ones(len(self.buffer)) / len(self.buffer)
        else:
            probabilities = priorities / sum_priorities

        batch_indices = np.random.choice(len(self.buffer), batch_size, p=probabilities)
        batch = [self.buffer[idx] for idx in batch_indices]
        return batch, batch_indices, probabilities[batch_indices]

    def __len__(self):
        return len(self.buffer)

    def get_high_reward_samples(self, target_path, num_samples=20):
        """从经验池中获取与目标路径相似度最高的样本"""
        if len(self.buffer) == 0:
            return []

        samples_with_similarity = []
        for experience in self.buffer:
            state_tensor = experience[0]
            state_tuple = tuple(state_tensor.cpu().numpy().flatten().astype(int))
            triggered = execute_Tr(state_tuple)
            sim = jaccard_similarity(triggered, target_path)

            # 如果发现完美匹配（相似度为1），直接返回该样本
            if sim >= 1.0:
                return [(state_tuple, 0, sim, triggered)]

            samples_with_similarity.append((state_tuple, 0, sim, triggered))

        # 按相似度从大到小排序，返回前num_samples个
        samples_with_similarity.sort(key=lambda x: x[2], reverse=True)
        return samples_with_similarity[:num_samples]


def load_path_data(file_path):
    path_data = []
    if not os.path.exists(file_path):
        return path_data

    with open(file_path, "r", encoding="utf-8") as f:
        lines = f.readlines()
        for line in lines[2:]:
            parts = line.strip().split("\t")
            if parts:
                state = tuple(map(int, parts[0].split()))
                path_data.append(state)
    return path_data


# === DQN网络 ===
class DQN(nn.Module):
    def __init__(self, state_dim, action_dim):
        super(DQN, self).__init__()
        self.fc1 = nn.Linear(state_dim, 128)
        self.fc2 = nn.Linear(128, 64)
        self.fc3 = nn.Linear(64, action_dim)

    def forward(self, state):
        x = torch.relu(self.fc1(state))
        x = torch.relu(self.fc2(x))
        return self.fc3(x)


# === DQN Agent===
class DQNAgentWithPER:
    def __init__(self, state_dim, action_dim, replay_buffer, gamma=0.99, epsilon=1.0, epsilon_decay=0.995,
                 epsilon_min=0.1, learning_rate=0.001, alpha=0.6, beta=0.4):
        self.state_dim = state_dim
        self.action_dim = action_dim
        self.gamma = gamma
        self.epsilon = epsilon
        self.epsilon_decay = epsilon_decay
        self.epsilon_min = epsilon_min
        self.learning_rate = learning_rate
        self.replay_buffer = replay_buffer
        self.alpha = alpha
        self.beta = beta

        self.model = DQN(state_dim, action_dim).to(device)
        self.target_model = DQN(state_dim, action_dim).to(device)
        self.optimizer = optim.Adam(self.model.parameters(), lr=self.learning_rate)
        self.target_model.load_state_dict(self.model.state_dict())

    def decode_action(self, action_idx):
        """解码动作索引为具体的维度增量"""
        dim = action_idx // 10  # 维度: 0=light, 1=temp, 2=moisture
        delta_idx = action_idx % 10  # 增量索引: 0-9

        if dim == 0:  # light维度
            delta = DELTA_LIGHT[delta_idx]
            return (delta, 0, 0)
        elif dim == 1:  # temp维度
            delta = DELTA_TEMP[delta_idx]
            return (0, delta, 0)
        elif dim == 2:  # moisture维度
            delta = DELTA_MOISTURE[delta_idx]
            return (0, 0, delta)

    def act(self, state):
        if random.random() < self.epsilon:
            return random.randrange(self.action_dim)
        state = torch.tensor(state, dtype=torch.float32).unsqueeze(0).to(device)
        with torch.no_grad():
            q_values = self.model(state)
        return torch.argmax(q_values, dim=1).item()

    def store_transition(self, state, action, reward, next_state, done):
        state = torch.tensor(state, dtype=torch.float32).unsqueeze(0).to(device)
        next_state = torch.tensor(next_state, dtype=torch.float32).unsqueeze(0).to(device)

        with torch.no_grad():
            q_values = self.model(state)
            next_q_values = self.target_model(next_state)
            max_next_q_values = next_q_values.max(1)[0]
            target_q_values = reward + (self.gamma * max_next_q_values * (1 - done))
            td_error = torch.abs(q_values[0][action] - target_q_values).item()

        self.replay_buffer.append((state, action, reward, next_state, done, td_error))
        return td_error

    def train(self, batch_size=32):
        if len(self.replay_buffer) < batch_size:
            return

        batch, batch_indices, probabilities = self.replay_buffer.sample(batch_size, alpha=self.alpha)
        states, actions, rewards, next_states, dones, _ = zip(*batch)

        states = torch.tensor(np.array([s.cpu().numpy().flatten() for s in states]), dtype=torch.float32).to(device)
        actions = torch.tensor(actions, dtype=torch.long).to(device)
        rewards = torch.tensor(rewards, dtype=torch.float32).to(device)
        next_states = torch.tensor(np.array([ns.cpu().numpy().flatten() for ns in next_states]),
                                   dtype=torch.float32).to(device)
        dones = torch.tensor(dones, dtype=torch.float32).to(device)

        current_q_values = self.model(states).gather(1, actions.unsqueeze(1)).squeeze(1)
        next_max_q_values = self.target_model(next_states).max(1)[0].detach()
        target_q_values = rewards + (self.gamma * next_max_q_values * (1 - dones))

        loss = nn.MSELoss()(current_q_values, target_q_values)

        self.optimizer.zero_grad()
        loss.backward()
        self.optimizer.step()

        if self.epsilon > self.epsilon_min:
            self.epsilon *= self.epsilon_decay

    def update_target_model(self):
        self.target_model.load_state_dict(self.model.state_dict())


# === 第一阶段训练函数===
def generate_and_train_for_similar_paths(agent, similar_group, path_documents, run_metrics, episodes=500, batch_size=32,
                                         steps_per_test=5, replay_times=10, is_isolated=False):
    trained_paths = set()

    for episode in range(episodes):
        for path_idx in similar_group:
            if path_idx in trained_paths:
                continue

            file_path = os.path.join(path_documents, f"path{path_idx + 1}{'_isolated' if is_isolated else ''}.txt")
            path_data = load_path_data(file_path)
            if not path_data:
                trained_paths.add(path_idx)
                continue

            target_path = targetPaths[path_idx]

            # === 训练流程参数 ===
            BATCH_SIZE = 50  # 每组50个样本
            N_SAMPLES = 200  # 总共200个样本
            N_STEPS = 3  # 每个样本走3步
            N_EPOCHS = 5  # 重复训练5次

            replay_count = 0  # 回放计数器，用于控制模型更新

            # 重复训练5次
            for epoch in range(N_EPOCHS):
                print(f"  路径{path_idx + 1} - 第{epoch + 1}/{N_EPOCHS}轮训练")

                # 将200个样本分成4组，每组50个
                for batch_start in range(0, N_SAMPLES, BATCH_SIZE):
                    batch_end = min(batch_start + BATCH_SIZE, N_SAMPLES)

                    # 处理当前批次的50个样本
                    for test_data_idx in range(batch_start, batch_end):
                        if test_data_idx >= len(path_data):
                            break

                        state = path_data[test_data_idx]
                        prev_state = None
                        prev_triggered = None
                        prev_reward = None

                        # 每个样本走3步
                        for step in range(N_STEPS):
                            # 获取合法动作
                            legal_actions = []
                            for a in range(agent.action_dim):
                                dx, dy, dz = agent.decode_action(a)
                                cand_next = (state[0] + dx, state[1] + dy, state[2] + dz)
                                if is_state_valid(cand_next):
                                    legal_actions.append(a)

                            if not legal_actions:
                                break

                            # 选择动作
                            if random.random() < agent.epsilon:
                                action = random.choice(legal_actions)
                            else:
                                state_tensor = torch.tensor(state, dtype=torch.float32).unsqueeze(0).to(device)
                                with torch.no_grad():
                                    q_values = agent.model(state_tensor)[0]
                                action = legal_actions[torch.argmax(q_values[legal_actions]).item()]

                            # 执行动作
                            dx, dy, dz = agent.decode_action(action)
                            next_state = (state[0] + dx, state[1] + dy, state[2] + dz)

                            # 计算奖励
                            triggered = execute_Tr(next_state)
                            reward = compute_reward(next_state, target_path, triggered, prev_triggered, prev_state)
                            done = (step == N_STEPS - 1)

                            # 存储经验
                            td_error = agent.store_transition(state, action, reward, next_state, done)
                            run_metrics.record_step_metrics(reward, td_error, triggered, target_path)

                            if prev_reward is not None:
                                run_metrics.record_action_improvement(reward, prev_reward)

                            # 更新状态
                            prev_state = state
                            prev_triggered = triggered
                            prev_reward = reward
                            state = next_state

                    # 当前批次的50个样本都走完后，进行一次回放训练
                    if len(agent.replay_buffer) >= batch_size:
                        agent.train(batch_size)
                        replay_count += 1

                        # 每回放2次后更新一次目标模型
                        if replay_count % 2 == 0:
                            agent.update_target_model()
                            print(
                                f"    批次 {batch_start // BATCH_SIZE + 1}/4 完成 | 已回放 {replay_count} 次 | 模型已更新")

            trained_paths.add(path_idx)
            print(f"  路径{path_idx + 1} 训练完成\n")

        if len(trained_paths) == len(similar_group):
            break

    return agent


def generate_samples_for_isolated_paths(agent_similar, isolated_group_indices, num_total=2000, top_k=200):
    def compute_q_value(state, agent):
        state_tensor = torch.tensor(state, dtype=torch.float32).unsqueeze(0).to(device)
        with torch.no_grad():
            q_values = agent.model(state_tensor)
        return q_values.max().item()

    def compute_robustness(state, path):
        base = execute_Tr(state)
        if not base:
            return 0.0
        rob, neighbors = 0.0, 0
        for dx in [-1, 0, 1]:
            for dy in [-1, 0, 1]:
                for dz in [-1, 0, 1]:
                    if dx == dy == dz == 0:
                        continue
                    neighbor = clip_state(state + np.array([dx, dy, dz]))
                    n_trig = execute_Tr(neighbor)
                    if not n_trig:
                        continue
                    rob += jaccard_similarity(base, n_trig)
                    neighbors += 1
        return rob / neighbors if neighbors > 0 else 0.0

    def save_samples(path_id, samples, base_dir):
        os.makedirs(base_dir, exist_ok=True)
        filepath = os.path.join(base_dir, f"path{path_id}_isolated.txt")
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(f"Isolated Path {path_id}\n")
            f.write("light temp moisture\tScore\tSimilarity\tLengthDiff\tRobustness\tQ_value_normalized_complement\n")
            for s in samples:
                light, temp, moisture = int(s[0][0]), int(s[0][1]), int(s[0][2])
                f.write(f"{light} {temp} {moisture}\t{s[1]:.4f}\t{s[2]:.4f}\t{s[3]:.4f}\t{s[4]:.4f}\t{s[5]:.4f}\n")

    base_dir = os.path.join(os.getcwd(), "path_samples")

    for path_idx in isolated_group_indices:
        path = targetPaths[path_idx]
        candidate_samples = []
        attempts = 0

        while len(candidate_samples) < num_total and attempts < num_total * 5:
            attempts += 1
            state = generate_random_state()
            triggered = execute_Tr(state)
            if not triggered:
                continue

            sim = jaccard_similarity(triggered, path)
            len_diff = 1 - abs(len(triggered) - len(path)) / max(len(triggered), len(path))
            rob = compute_robustness(state, path)
            q_value = compute_q_value(state, agent_similar)

            candidate_samples.append((state, sim, len_diff, rob, q_value))

        if not candidate_samples:
            continue

        q_values = [sample[4] for sample in candidate_samples]
        q_min = min(q_values)
        q_max = max(q_values)

        normalized_samples = []
        for state, sim, len_diff, rob, q_value in candidate_samples:
            if q_max - q_min > 0:
                q_normalized = (q_value - q_min) / (q_max - q_min)
            else:
                q_normalized = 0.5

            q_complement = 1.0 - q_normalized
            score = 0.28 * sim + 0.1 * len_diff + 0.19 * rob + 0.43 * q_complement

            normalized_samples.append((state, score, sim, len_diff, rob, q_complement))

        normalized_samples.sort(key=lambda x: x[1], reverse=True)
        top_samples = normalized_samples[:top_k]

        save_samples(path_id=path_idx + 1, samples=top_samples, base_dir=base_dir)


# === 第二阶段增强训练函数===
def generate_and_train_for_isolated_paths_enhanced(agent_similar, agent_isolated, similar_group, isolated_group,
                                                   path_documents, run_metrics, episodes=500, batch_size=32,
                                                   is_isolated=True):
    trained_paths = set()

    for episode in range(episodes):
        for path_idx in isolated_group:
            if path_idx in trained_paths:
                continue

            file_path = os.path.join(path_documents, f"path{path_idx + 1}_isolated.txt")
            stage2_path_data = load_path_data(file_path)
            if not stage2_path_data:
                trained_paths.add(path_idx)
                continue

            target_path = targetPaths[path_idx]

            # === 训练流程参数 ===
            BATCH_SIZE = 50  # 每组50个样本
            N_SAMPLES = 200  # 总共200个样本
            N_STEPS = 3  # 每个样本走3步
            N_EPOCHS = 5  # 重复训练5次

            replay_count = 0  # 回放计数器

            # 重复训练5次
            for epoch in range(N_EPOCHS):
                print(f"  孤岛路径{path_idx + 1} - 第{epoch + 1}/{N_EPOCHS}轮训练")

                # 将200个样本分成4组,每组50个
                for batch_start in range(0, N_SAMPLES, BATCH_SIZE):
                    batch_end = min(batch_start + BATCH_SIZE, N_SAMPLES)

                    # 处理当前批次的50个样本
                    for test_data_idx in range(batch_start, batch_end):
                        if test_data_idx >= len(stage2_path_data):
                            break

                        state = stage2_path_data[test_data_idx]
                        prev_state = None
                        prev_triggered = None
                        prev_reward = None

                        # 每个样本走3步
                        for step in range(N_STEPS):
                            # 获取合法动作
                            legal_actions = []
                            for a in range(agent_isolated.action_dim):
                                dx, dy, dz = agent_isolated.decode_action(a)
                                cand_next = (state[0] + dx, state[1] + dy, state[2] + dz)
                                if is_state_valid(cand_next):
                                    legal_actions.append(a)

                            if not legal_actions:
                                break

                            # 选择动作
                            if random.random() < agent_isolated.epsilon:
                                action = random.choice(legal_actions)
                            else:
                                state_tensor = torch.tensor(state, dtype=torch.float32).unsqueeze(0).to(device)
                                with torch.no_grad():
                                    q_values = agent_isolated.model(state_tensor)[0]
                                action = legal_actions[torch.argmax(q_values[legal_actions]).item()]

                            # 执行动作
                            dx, dy, dz = agent_isolated.decode_action(action)
                            next_state = (state[0] + dx, state[1] + dy, state[2] + dz)

                            # 计算奖励
                            triggered = execute_Tr(next_state)
                            reward = compute_reward(next_state, target_path, triggered, prev_triggered, prev_state)

                            # 对孤岛路径给予额外奖励
                            if target_path.issubset(triggered):
                                reward += 2.0

                            done = (step == N_STEPS - 1)

                            # 存储经验
                            td_error = agent_isolated.store_transition(state, action, reward, next_state, done)
                            run_metrics.record_step_metrics(reward, td_error, triggered, target_path)

                            if prev_reward is not None:
                                run_metrics.record_action_improvement(reward, prev_reward)

                            # 更新状态
                            prev_state = state
                            prev_triggered = triggered
                            prev_reward = reward
                            state = next_state

                    # 当前批次的50个样本都走完后，进行一次回放训练
                    if len(agent_isolated.replay_buffer) >= batch_size:
                        agent_isolated.train(batch_size)
                        replay_count += 1

                        # 每回放2次后更新一次目标模型
                        if replay_count % 2 == 0:
                            agent_isolated.update_target_model()
                            print(
                                f"    批次 {batch_start // BATCH_SIZE + 1}/4 完成 | 已回放 {replay_count} 次 | 模型已更新")

            trained_paths.add(path_idx)
            print(f"  孤岛路径{path_idx + 1} 训练完成\n")

        if len(trained_paths) == len(isolated_group):
            break

    return agent_isolated


# === PSO粒子类 ===
class Particle:
    def __init__(self, initial_position=None):
        if initial_position is not None:
            self.position = np.array(initial_position, dtype=float)
        else:
            self.position = np.array([
                np.random.uniform(BOUNDS['light'][0], BOUNDS['light'][1]),
                np.random.uniform(BOUNDS['temp'][0], BOUNDS['temp'][1]),
                np.random.uniform(BOUNDS['moisture'][0], BOUNDS['moisture'][1])
            ])

        self.velocity = np.array([
            np.random.uniform(-5, 5),
            np.random.uniform(-3, 3),
            np.random.uniform(-5, 5)
        ])

        self.best_position = self.position.copy()
        self.best_fitness = 0
        self.fitness = 0


# === 普通PSO优化器类 ===
class PSO:
    def __init__(self, target_path, swarm_size=20, dqn_samples=None):
        self.target_path = target_path
        self.swarm_size = swarm_size
        self.particles = []
        self.global_best_position = None
        self.global_best_fitness = 0
        self.reset_count = 0

        if dqn_samples is not None and len(dqn_samples) > 0:
            num_direct = min(len(dqn_samples), swarm_size)
            for i in range(num_direct):
                state_tuple, reward, sim, triggered = dqn_samples[i]
                particle = Particle(initial_position=state_tuple)
                self.particles.append(particle)

            if len(self.particles) < swarm_size:
                remaining = swarm_size - len(self.particles)
                for i in range(remaining):
                    base_idx = i % len(dqn_samples)
                    state_tuple, _, _, _ = dqn_samples[base_idx]
                    perturbed = np.array(state_tuple) + np.random.randint(-10, 11, size=3)
                    perturbed = clip_state(perturbed)
                    particle = Particle(initial_position=perturbed.tolist())
                    self.particles.append(particle)
        else:
            self.particles = [Particle() for _ in range(swarm_size)]

        for particle in self.particles:
            particle.fitness = self.fitness_function(particle.position)
            if particle.fitness > particle.best_fitness:
                particle.best_fitness = particle.fitness
                particle.best_position = particle.position.copy()
            if particle.fitness > self.global_best_fitness:
                self.global_best_fitness = particle.fitness
                self.global_best_position = particle.position.copy()

    def fitness_function(self, position):
        try:
            triggered = execute_Tr(position)
            # 使用子集判断：如果目标路径是触发路径的子集，适应度为1
            if self.target_path.issubset(triggered):
                return 1.0

            # 否则计算Jaccard相似度
            intersection = len(triggered & self.target_path)
            union = len(triggered | self.target_path)
            return intersection / union if union > 0 else 0.0
        except:
            return 0.0

    def update(self, iteration, max_iterations):
        # 标准PSO参数
        w = 0.7  # 惯性权重
        c1 = 1.5  # 个体学习因子
        c2 = 1.5  # 社会学习因子

        for particle in self.particles:
            r1 = np.random.random(3)
            r2 = np.random.random(3)

            # 更新速度
            particle.velocity = (w * particle.velocity +
                                 c1 * r1 * (particle.best_position - particle.position) +
                                 c2 * r2 * (self.global_best_position - particle.position))

            # 限制最大速度
            max_velocity = np.array([
                (BOUNDS['light'][1] - BOUNDS['light'][0]) * 0.2,
                (BOUNDS['temp'][1] - BOUNDS['temp'][0]) * 0.2,
                (BOUNDS['moisture'][1] - BOUNDS['moisture'][0]) * 0.2
            ])
            particle.velocity = np.clip(particle.velocity, -max_velocity, max_velocity)

            # 更新位置
            particle.position += particle.velocity
            particle.position = clip_state(particle.position)

            # 计算适应度
            particle.fitness = self.fitness_function(particle.position)

            # 更新个体最优
            if particle.fitness > particle.best_fitness:
                particle.best_fitness = particle.fitness
                particle.best_position = particle.position.copy()

            # 更新全局最优
            if particle.fitness > self.global_best_fitness:
                self.global_best_fitness = particle.fitness
                self.global_best_position = particle.position.copy()


# === 改进的Excel导出函数 ===
def export_multiple_runs_to_excel(all_run_results, all_run_metrics, num_runs, filename=None):
    """导出多次运行结果到Excel"""

    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"DQN_PSO_{num_runs}Runs_{timestamp}.xlsx"

    wb = Workbook()

    # 样式定义
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name='微软雅黑', size=11, bold=True, color="FFFFFF")

    success_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    alternate_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    dqn_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")

    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')

    # ========== 工作表1: 运行汇总 ==========
    ws1 = wb.active
    ws1.title = "运行汇总"
    ws1.sheet_view.showGridLines = False

    # 表头
    headers = ["运行次数", "成功率", "成功数量", "平均适应度", "平均迭代次数", "运行时间(s)", "DQN解决数"]
    col_widths = [12, 12, 12, 14, 14, 14, 12]

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws1.column_dimensions[get_column_letter(col)].width = width

    # 填充数据
    for run_idx, (results, run_metrics) in enumerate(zip(all_run_results, all_run_metrics), start=1):
        success_count = sum(1 for r in results if r['perfect_match'])
        success_rate = (success_count / len(targetPaths)) * 100
        avg_fitness = np.mean([r['fitness'] for r in results])

        # 计算平均迭代次数
        iterations_list = []
        for r in results:
            if r.get('method') == 'DQN':
                iterations_list.append(0)
            elif r.get('convergence_iteration') is not None:
                iterations_list.append(r['convergence_iteration'])
            else:
                iterations_list.append(10000)
        avg_iterations = np.mean(iterations_list)

        # 统计DQN解决的数量
        dqn_solved_count = sum(1 for r in results if r.get('method') == 'DQN')

        # 计算总运行时间
        total_time = run_metrics.pso_end_time - run_metrics.pso_start_time if run_metrics.pso_end_time else 0

        row_data = [
            f"运行 {run_idx}",
            f"{success_rate:.1f}%",
            f"{success_count}/{len(targetPaths)}",
            f"{avg_fitness:.4f}",
            f"{avg_iterations:.1f}",
            f"{total_time:.2f}",
            f"{dqn_solved_count}/{len(targetPaths)}"
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws1.cell(row=run_idx + 1, column=col, value=value)
            cell.border = border
            cell.alignment = center_align

            if run_idx % 2 == 0:
                cell.fill = alternate_fill

            if col == 2:
                if success_rate == 100.0:
                    cell.fill = success_fill
                elif success_rate < 50.0:
                    cell.fill = fail_fill

            if col == 7 and dqn_solved_count > 0:
                cell.fill = dqn_fill

    ws1.freeze_panes = 'A2'
    ws1.auto_filter.ref = f"A1:G{len(all_run_results) + 1}"

    # ========== 工作表2: 路径统计 ==========
    ws2 = wb.create_sheet(title="路径统计")
    ws2.sheet_view.showGridLines = False

    headers2 = ["路径编号", "成功次数", "成功率", "平均适应度", "平均迭代次数", "最小迭代次数", "最大迭代次数",
                "DQN解决次数"]
    col_widths2 = [12, 12, 12, 14, 14, 14, 14, 14]

    for col, (header, width) in enumerate(zip(headers2, col_widths2), 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws2.column_dimensions[get_column_letter(col)].width = width

    num_paths = len(targetPaths)
    for path_idx in range(num_paths):
        success_count = sum(1 for results in all_run_results if results[path_idx]['perfect_match'])
        success_rate = (success_count / num_runs) * 100
        avg_fitness = np.mean([results[path_idx]['fitness'] for results in all_run_results])

        iterations_list = []
        for results in all_run_results:
            r = results[path_idx]
            if r.get('method') == 'DQN':
                iterations_list.append(0)
            elif r.get('convergence_iteration') is not None:
                iterations_list.append(r['convergence_iteration'])
            else:
                iterations_list.append(10000)

        avg_iterations = np.mean(iterations_list) if iterations_list else 0
        min_iterations = np.min(iterations_list) if iterations_list else 0
        max_iterations = np.max(iterations_list) if iterations_list else 0

        dqn_solved_count = sum(1 for results in all_run_results
                               if results[path_idx].get('method') == 'DQN')

        row_data = [
            f"路径 {path_idx + 1}",
            f"{success_count}/{num_runs}",
            f"{success_rate:.1f}%",
            f"{avg_fitness:.4f}",
            f"{avg_iterations:.1f}",
            f"{min_iterations}",
            f"{max_iterations}",
            f"{dqn_solved_count}/{num_runs}"
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws2.cell(row=path_idx + 2, column=col, value=value)
            cell.border = border
            cell.alignment = center_align

            if (path_idx + 1) % 2 == 0:
                cell.fill = alternate_fill

            if col == 3:
                if success_rate == 100.0:
                    cell.fill = success_fill
                elif success_rate < 50.0:
                    cell.fill = fail_fill

            if col == 8 and dqn_solved_count > 0:
                cell.fill = dqn_fill

    ws2.freeze_panes = 'A2'
    ws2.auto_filter.ref = f"A1:H{len(targetPaths) + 1}"

    # ========== 工作表3: 最佳粒子详情 ==========
    ws3 = wb.create_sheet(title="最佳粒子详情")
    ws3.sheet_view.showGridLines = False

    headers3 = ["路径", "运行", "最佳粒子(light,temp,moisture)", "适应度", "迭代次数", "求解方法", "生成路径"]
    col_widths3 = [10, 10, 25, 12, 12, 12, 50]

    for col, (header, width) in enumerate(zip(headers3, col_widths3), 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws3.column_dimensions[get_column_letter(col)].width = width

    row_idx = 2
    for path_idx in range(num_paths):
        for run_idx, results in enumerate(all_run_results, start=1):
            result = results[path_idx]
            best_position = result['best_position']
            fitness = result['fitness']
            triggered = result['triggered']
            method = result.get('method', 'PSO')

            if method == 'DQN':
                convergence_iter = 0
            elif result.get('convergence_iteration') is not None:
                convergence_iter = result['convergence_iteration']
            else:
                convergence_iter = 10000

            particle_str = f"({int(best_position[0])}, {int(best_position[1])}, {int(best_position[2])})"
            path_str = str(sorted(list(triggered)))

            row_data = [
                f"路径{path_idx + 1}",
                f"运行{run_idx}",
                particle_str,
                f"{fitness:.4f}",
                convergence_iter if convergence_iter < 10000 else "-",
                method,
                path_str
            ]

            for col, value in enumerate(row_data, 1):
                cell = ws3.cell(row=row_idx, column=col, value=value)
                cell.border = border

                if col == 7:
                    cell.alignment = left_align
                else:
                    cell.alignment = center_align

                if fitness == 1.0:
                    if method == 'DQN':
                        cell.fill = dqn_fill
                    else:
                        cell.fill = success_fill
                elif fitness < 0.5:
                    cell.fill = fail_fill
                elif row_idx % 2 == 0:
                    cell.fill = alternate_fill

            row_idx += 1

    ws3.freeze_panes = 'A2'
    ws3.auto_filter.ref = f"A1:G{row_idx - 1}"

    # ========== 工作表4: 目标路径 ==========
    ws4 = wb.create_sheet(title="目标路径")
    ws4.sheet_view.showGridLines = False

    headers4 = ["路径编号", "目标路径", "规则数量"]
    col_widths4 = [12, 60, 12]

    for col, (header, width) in enumerate(zip(headers4, col_widths4), 1):
        cell = ws4.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        ws4.column_dimensions[get_column_letter(col)].width = width

    for path_idx, target_path in enumerate(targetPaths):
        path_str = str(sorted(list(target_path)))

        row_data = [
            f"路径 {path_idx + 1}",
            path_str,
            len(target_path)
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws4.cell(row=path_idx + 2, column=col, value=value)
            cell.border = border

            if col == 2:
                cell.alignment = left_align
            else:
                cell.alignment = center_align

            if (path_idx + 1) % 2 == 0:
                cell.fill = alternate_fill

    ws4.freeze_panes = 'A2'

    # 保存文件
    output_dir = os.path.join(os.getcwd(), "results")
    os.makedirs(output_dir, exist_ok=True)
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)

    print(f"\n{'=' * 70}")
    print(f"✓ 结果已导出到: {filepath}")
    print(f"{'=' * 70}")
    print(f"包含以下工作表:")
    print(f"  1. 运行汇总       - {num_runs}次运行的整体统计（含DQN解决数）")
    print(f"  2. 路径统计       - 每条路径的详细统计（含DQN解决次数）")
    print(f"  3. 最佳粒子详情   - 每次运行每条路径的最佳解（标注求解方法）")
    print(f"  4. 目标路径       - 所有目标路径的定义")
    print(f"{'=' * 70}\n")

    return filepath


def run_single_experiment(run_num, similar_group, isolated_group):
    """运行单次DQN-PSO实验"""
    print(f"\n{'=' * 100}")
    print(f"开始第 {run_num} 次运行")
    print(f"{'=' * 100}")

    run_metrics = MetricsCollector()
    run_metrics.start_training()

    path_documents = os.path.join(os.getcwd(), "path_samples")

    if run_num == 1:
        print("生成相似路径样本数据...")
        generate_samples_for_similar_paths(similar_group, num_total=2000, top_k=200)

    replay_buffer = SharedExperienceReplay(capacity=10000)
    state_dim = 3
    action_dim = 30
    agent = DQNAgentWithPER(state_dim, action_dim, replay_buffer)

    print(f"运行{run_num} - 第一阶段：相似路径训练")
    generate_and_train_for_similar_paths(agent, similar_group, path_documents, run_metrics, episodes=500, batch_size=32,
                                         is_isolated=False)

    if run_num == 1:
        print("生成孤岛路径增强样本...")
        generate_samples_for_isolated_paths(agent, isolated_group, num_total=2000, top_k=200)

    print(f"运行{run_num} - 第二阶段：增强版孤岛路径训练")
    isolated_replay_buffer = SharedExperienceReplay(capacity=15000)
    agent_isolated = DQNAgentWithPER(state_dim, action_dim, isolated_replay_buffer)

    agent_isolated.model.load_state_dict(agent.model.state_dict())
    agent_isolated.target_model.load_state_dict(agent.model.state_dict())

    agent_isolated = generate_and_train_for_isolated_paths_enhanced(
        agent_similar=agent,
        agent_isolated=agent_isolated,
        similar_group=similar_group,
        isolated_group=isolated_group,
        path_documents=path_documents,
        run_metrics=run_metrics,
        episodes=500,
        batch_size=32,
        is_isolated=True
    )

    run_metrics.end_training()

    # 从经验池提取最终样本
    dqn_best_samples = {}

    for path_idx in similar_group:
        target_path = targetPaths[path_idx]
        high_reward_samples = agent.replay_buffer.get_high_reward_samples(target_path, num_samples=20)
        dqn_best_samples[path_idx] = high_reward_samples
        for state_tuple, _, sim, triggered in high_reward_samples:
            run_metrics.record_final_output_sample(triggered, target_path)

    for path_idx in isolated_group:
        target_path = targetPaths[path_idx]
        high_reward_samples = agent_isolated.replay_buffer.get_high_reward_samples(target_path, num_samples=20)
        dqn_best_samples[path_idx] = high_reward_samples
        for state_tuple, _, sim, triggered in high_reward_samples:
            run_metrics.record_final_output_sample(triggered, target_path)

    # PSO优化阶段
    print(f"运行{run_num} - PSO优化阶段")
    run_metrics.start_pso_phase()

    max_iterations = 3000
    pso_results = []

    for i, target_path in enumerate(targetPaths):
        path_start_time = time.time()
        dqn_samples_for_path = dqn_best_samples.get(i, [])

        perfect_solution_found = False
        perfect_solution_state = None

        if dqn_samples_for_path:
            for sample in dqn_samples_for_path:
                state_tuple, reward, sim, triggered = sample
                if target_path.issubset(triggered):
                    perfect_solution_found = True
                    perfect_solution_state = state_tuple
                    break

        if perfect_solution_found:
            path_execution_time = time.time() - path_start_time
            pso_results.append({
                'target_path': target_path,
                'best_position': np.array(perfect_solution_state),
                'fitness': 1.0,
                'triggered': triggered,
                'perfect_match': True,
                'method': 'DQN',
                'convergence_iteration': 0,
                'early_stopped': False,
                'reset_count': 0
            })
            run_metrics.record_pso_result(1.0, True, convergence_iter=0, path_id=i + 1,
                                          method='DQN', reset_count=0, execution_time=path_execution_time)
            status = "✓完美(DQN)"
        else:
            pso = PSO(target_path, swarm_size=20, dqn_samples=dqn_samples_for_path)

            converged_at_iteration = max_iterations
            early_stop = False

            for iteration in range(max_iterations):
                pso.update(iteration, max_iterations)

                if pso.global_best_fitness >= 1.0:
                    converged_at_iteration = iteration + 1
                    early_stop = True
                    break

            path_execution_time = time.time() - path_start_time
            best_position = pso.global_best_position
            triggered = execute_Tr(best_position)

            is_perfect = target_path.issubset(triggered)

            pso_results.append({
                'target_path': target_path,
                'best_position': best_position,
                'fitness': pso.global_best_fitness,
                'triggered': triggered,
                'perfect_match': is_perfect,
                'method': 'PSO',
                'convergence_iteration': converged_at_iteration,
                'early_stopped': early_stop,
                'reset_count': pso.reset_count
            })

            run_metrics.record_pso_result(
                fitness=pso.global_best_fitness,
                is_perfect_match=is_perfect,
                convergence_iter=converged_at_iteration if early_stop else None,
                path_id=i + 1,
                method='PSO',
                reset_count=pso.reset_count,
                execution_time=path_execution_time
            )

            status = "✓完美(PSO)" if is_perfect else f"○部分({pso.global_best_fitness:.3f})"

        print(f"  路径{i + 1}: {status} | 耗时 {path_execution_time:.2f}s")

    run_metrics.end_pso_phase()

    success_count = sum(1 for r in pso_results if r['perfect_match'])
    success_rate = (success_count / len(targetPaths)) * 100
    pso_time = run_metrics.pso_end_time - run_metrics.pso_start_time

    print(f"\n第{run_num}次运行完成: 成功率 {success_rate:.1f}% ({success_count}/{len(targetPaths)}) | "
          f"PSO耗时 {pso_time:.2f}秒")

    return pso_results, run_metrics


def run_multiple_dqn_pso_experiments(num_runs):
    """运行多次DQN-PSO实验"""
    print("\n" + "=" * 100)
    print(f"DQN-PSO算法（普通PSO） - {num_runs}次独立运行实验")
    print("=" * 100)
    print(f"目标路径: 路径1 至 路径{len(targetPaths)} (共{len(targetPaths)}条)")
    print(f"运行次数: {num_runs}")
    print(f"三维取值范围：Light: {BOUNDS['light']}, Temp: {BOUNDS['temp']}, Moisture: {BOUNDS['moisture']}")
    print("=" * 100)

    similar_group, isolated_group = group_paths_by_similarity(targetPaths)

    similar_paths_display = [idx + 1 for idx in similar_group]
    isolated_paths_display = [idx + 1 for idx in isolated_group]

    print(f"相似路径组: {similar_paths_display}")
    print(f"孤岛路径组: {isolated_paths_display}\n")

    all_run_results = []
    all_run_metrics = []
    total_experiment_start = time.time()

    for run_num in range(1, num_runs + 1):
        pso_results, run_metrics = run_single_experiment(run_num, similar_group, isolated_group)
        all_run_results.append(pso_results)
        all_run_metrics.append(run_metrics)

    total_experiment_time = time.time() - total_experiment_start

    print(f"\n{'=' * 100}")
    print(f"全部{num_runs}次运行完成!")
    print(f"总实验耗时: {total_experiment_time:.2f}秒 ({total_experiment_time / 60:.2f}分钟)")
    print(f"{'=' * 100}\n")

    return all_run_results, all_run_metrics


if __name__ == "__main__":
    print("=" * 100)
    print("DQN-PSO算法（普通PSO） - 多次独立运行实验系统")
    print("=" * 100)
    print(f"当前配置：运行次数 = {NUM_RUNS}")
    print(f"路径范围：路径1 - 路径{len(targetPaths)}")
    print("=" * 100)

    # 可以通过命令行参数修改运行次数
    if len(sys.argv) > 1:
        try:
            NUM_RUNS = int(sys.argv[1])
            print(f"从命令行读取：运行次数已修改为 {NUM_RUNS}")
        except ValueError:
            print(f"命令行参数无效，使用默认运行次数 {NUM_RUNS}")

    print("\n启动实验...")

    all_run_results, all_run_metrics = run_multiple_dqn_pso_experiments(num_runs=NUM_RUNS)

    print("\n正在导出结果到Excel...")
    excel_filename = export_multiple_runs_to_excel(all_run_results, all_run_metrics, NUM_RUNS)

    print(f"\n程序执行完成!")
    print(f"Excel文件: {excel_filename}")
    print("=" * 100)